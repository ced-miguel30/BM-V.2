"""Pruebas Fase 3 — registro de comida.

Ejecutar:

    py -m unittest tests.test_multi_fase3_comida -v
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.models import (
    AppData,
    CategoriaReceta,
    IngredienteReceta,
    LoteStock,
    OrigenConsumo,
    Producto,
    Receta,
    UnidadProducto,
    Usuario,
)
from app.core.models.enums import RolUsuario
from app.core.services import comida_service
from app.core.services.exportacion_semanal_service import exportar_periodo
from app.core.services.excel_bloques import nombre_hoja_dia
from tests.demo_isolation import EXPORT_SESSION_MODULES, isolated_persist


def _datos() -> AppData:
    return AppData(
        productos=[
            Producto("p01", "Arroz", UnidadProducto.KG),
            Producto("p02", "Zumo", UnidadProducto.L, es_bebida=True),
            Producto("p03", "Pan", UnidadProducto.KG),
        ],
        lotes=[
            LoteStock("l01", "p01", 20.0, 2.0, 2.0, fecha_compra=date(2026, 7, 1)),
            LoteStock("l02", "p02", 4.0, 2.0, 2.0, fecha_compra=date(2026, 7, 1)),
            LoteStock("l03", "p03", 5.0, 1.0, 1.0, fecha_compra=date(2026, 7, 1)),
        ],
        recetas=[
            Receta("r_com", "Paella", [IngredienteReceta("p01", 0.3)], CategoriaReceta.COMIDA, porciones_estandar=1.0),
            Receta("r_beb", "Zumo natural", [IngredienteReceta("p02", 0.2)], CategoriaReceta.BEBIDAS, porciones_estandar=1.0),
            Receta("r_des", "Tostada", [IngredienteReceta("p03", 0.1)], CategoriaReceta.DESAYUNO, porciones_estandar=1.0),
            Receta("r_cen", "Sopa", [IngredienteReceta("p01", 0.2)], CategoriaReceta.CENA, porciones_estandar=1.0),
        ],
        usuarios=[Usuario("u01", "Ana", RolUsuario.ADMIN, True)],
        usuario_actual_id="u01",
    )


class _FakeSession(dict):
    def __init__(self, store: dict):
        super().__init__()
        self._store = store

    def __setitem__(self, key, value):
        self._store[key] = value
        super().__setitem__(key, value)

    def get(self, key, default=None):
        return self._store.get(key, default)

    def __contains__(self, key):
        return key in self._store


class TestComidaRegistro(unittest.TestCase):
    def setUp(self) -> None:
        self.data = _datos()
        self.state: dict = {}
        fake_st = MagicMock()
        fake_st.session_state = _FakeSession(self.state)

        self.patches = [
            patch("app.core.services.servicio_registro_service.get_data", return_value=self.data),
            patch("app.core.services.servicio_registro_service.persist_data"),
            patch("app.core.services.cesta_service.get_data", return_value=self.data),
            patch("app.core.services.alert_service.sincronizar_alertas"),
            patch.dict("sys.modules", {"streamlit": fake_st}),
        ]
        for p in self.patches:
            p.start()
        import streamlit as st
        st.session_state = fake_st.session_state

        # Recrear el servicio para que use un MotorCesta limpio
        from app.core.services.servicio_registro_service import crear_servicio
        from app.core.models import CategoriaReceta
        comida_service.servicio = crear_servicio(
            "comida", "comida",
            [CategoriaReceta.COMIDA, CategoriaReceta.BEBIDAS],
            titulo_documento="Registro de Comida",
        )
        for nombre in (
            "anadir_a_cesta", "anadir_receta_a_cesta", "registrar",
            "historial_ordenado", "fecha_mas_antigua", "registros_exportables",
            "configuracion_exportacion", "limpiar_cesta",
        ):
            setattr(comida_service, nombre, getattr(comida_service.servicio, nombre))

    def tearDown(self) -> None:
        for p in self.patches:
            p.stop()

    def test_acepta_receta_comida_y_bebidas(self) -> None:
        self.assertTrue(comida_service.anadir_receta_a_cesta("r_com", 1.0).ok)
        self.assertTrue(comida_service.anadir_receta_a_cesta("r_beb", 1.0).ok)

    def test_rechaza_desayuno_y_cena(self) -> None:
        self.assertFalse(comida_service.anadir_receta_a_cesta("r_des", 1.0).ok)
        self.assertFalse(comida_service.anadir_receta_a_cesta("r_cen", 1.0).ok)

    def test_registrar_descuento_e_origen(self) -> None:
        self.assertTrue(comida_service.anadir_receta_a_cesta("r_com", 1.0).ok)
        self.assertTrue(comida_service.anadir_a_cesta("p03", 0.2).ok)
        resultado = comida_service.registrar(date(2026, 7, 21))
        self.assertTrue(resultado.ok, resultado.mensaje)

        self.assertEqual(len(self.data.registros_servicio), 1)
        reg = self.data.registros_servicio[0]
        self.assertEqual(reg.tipo_servicio, "comida")

        # Stock: paella 0.3 arroz + 0.2 pan suelto
        arroz = next(l for l in self.data.lotes if l.id == "l01")
        pan = next(l for l in self.data.lotes if l.id == "l03")
        self.assertAlmostEqual(arroz.cantidad_restante, 1.7, places=4)
        self.assertAlmostEqual(pan.cantidad_restante, 0.8, places=4)

        origenes = {d.origen for d in reg.lineas_detalle}
        self.assertIn(OrigenConsumo.INGREDIENTE_RECETA.value, origenes)
        self.assertIn(OrigenConsumo.PRODUCTO_DIRECTO.value, origenes)
        self.assertTrue(all(d.tipo_servicio == "comida" for d in reg.lineas_detalle))

        # Actividad registrada
        self.assertTrue(any("comida" in a.accion.lower() for a in self.data.actividades))

    def test_historial_filtra_por_tipo(self) -> None:
        self.assertTrue(comida_service.anadir_receta_a_cesta("r_com", 1.0).ok)
        self.assertTrue(comida_service.registrar(date(2026, 7, 21)).ok)
        historial = comida_service.historial_ordenado()
        self.assertEqual(len(historial), 1)
        self.assertEqual(historial[0].tipo_servicio, "comida")

    def test_exportacion_genera_hoja_por_dia(self) -> None:
        self.assertTrue(comida_service.anadir_receta_a_cesta("r_com", 1.0).ok)
        self.assertTrue(comida_service.registrar(date(2026, 7, 21)).ok)

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            meta = carpeta / "_meta.json"
            with isolated_persist(*EXPORT_SESSION_MODULES, data=self.data):
                resultado = exportar_periodo(
                    comida_service.configuracion_exportacion(),
                    date(2026, 7, 20),
                    datetime(2026, 7, 26, 23, 59, 59),
                    automatica=True,
                    carpeta_exports=carpeta,
                    archivo_meta=meta,
                )
            self.assertTrue(resultado.ok, resultado.mensaje)
            self.assertIsNotNone(resultado.ruta)
            from openpyxl import load_workbook
            libro = load_workbook(resultado.ruta)
            self.assertIn("Info", libro.sheetnames)
            self.assertIn(nombre_hoja_dia(date(2026, 7, 21)), libro.sheetnames)
            self.assertTrue(resultado.nombre_archivo.startswith("Registro_de_Comida_"))


if __name__ == "__main__":
    unittest.main()
