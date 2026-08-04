"""Pruebas del motor central de exportación semanal.

Ejecutar desde la raíz del proyecto con:

    python -m unittest discover -s tests

Estas pruebas no dependen de una sesión de Streamlit ni de los datos reales
de la app: usan carpetas temporales y configuraciones sintéticas inyectadas
vía los parámetros `carpeta_exports` / `archivo_meta` / `config.carpeta`.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime, time
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.models import AppData
from app.core.services.exportacion_semanal_service import (
    ConfiguracionExportacionModulo,
    exportar_periodo,
    exportar_semana_actual,
    limite_semana,
    procesar_pendientes,
    rango_manual_actual,
    rango_semana_actual,
    semanas_pendientes,
    ultima_semana_exportada,
)
from app.core.services.excel_bloques import RegistroExportable
from tests.demo_isolation import EXPORT_SESSION_MODULES, isolated_persist


def _start_export_isolation(testcase: unittest.TestCase, data: AppData | None = None) -> None:
    cm = isolated_persist(*EXPORT_SESSION_MODULES, data=data or AppData())
    testcase._export_isolation_cm = cm  # type: ignore[attr-defined]
    cm.__enter__()


def _stop_export_isolation(testcase: unittest.TestCase) -> None:
    cm = getattr(testcase, "_export_isolation_cm", None)
    if cm is not None:
        cm.__exit__(None, None, None)
        testcase._export_isolation_cm = None  # type: ignore[attr-defined]


def _registro(fecha: date, identificador: str, hora: time | None = time(8, 30)) -> RegistroExportable:
    return RegistroExportable(
        fecha=fecha,
        hora=hora,
        tipo="Prueba",
        identificador=identificador,
        usuario="Tester",
        columnas=["Producto", "Cantidad", "Coste"],
        filas=[["Café", 1.5, 2.3]],
        resumen=[("Coste total", "2,30 €")],
    )


def _config_vacia(tipo: str = "prueba", titulo: str = "Registro de Prueba") -> ConfiguracionExportacionModulo:
    return ConfiguracionExportacionModulo(tipo=tipo, titulo_documento=titulo, obtener_registros=lambda i, h: [])


class TestCalculoSemana(unittest.TestCase):
    def test_semana_normal_lunes_a_domingo(self) -> None:
        # Miércoles 15/07/2026 -> semana lunes 13/07 a domingo 19/07/2026.
        lunes, domingo = limite_semana(date(2026, 7, 15))
        self.assertEqual(lunes, date(2026, 7, 13))
        self.assertEqual(domingo, date(2026, 7, 19))

    def test_fecha_ya_es_lunes(self) -> None:
        lunes, domingo = limite_semana(date(2026, 7, 20))
        self.assertEqual(lunes, date(2026, 7, 20))
        self.assertEqual(domingo, date(2026, 7, 26))

    def test_fecha_ya_es_domingo(self) -> None:
        lunes, domingo = limite_semana(date(2026, 7, 26))
        self.assertEqual(lunes, date(2026, 7, 20))
        self.assertEqual(domingo, date(2026, 7, 26))

    def test_cambio_de_mes(self) -> None:
        # Sábado 1 de agosto de 2026 -> semana de julio->agosto (27 jul - 2 ago).
        lunes, domingo = limite_semana(date(2026, 8, 1))
        self.assertEqual(lunes, date(2026, 7, 27))
        self.assertEqual(domingo, date(2026, 8, 2))

    def test_cambio_de_anio(self) -> None:
        # Jueves 1 de enero de 2026 -> semana 29 dic 2025 - 4 ene 2026.
        lunes, domingo = limite_semana(date(2026, 1, 1))
        self.assertEqual(lunes, date(2025, 12, 29))
        self.assertEqual(domingo, date(2026, 1, 4))

    def test_rango_semana_actual_incluye_domingo_completo(self) -> None:
        ahora = datetime(2026, 7, 22, 9, 0, 0)
        inicio, hasta = rango_semana_actual(ahora)
        self.assertEqual(inicio, date(2026, 7, 20))
        self.assertEqual(hasta, datetime(2026, 7, 26, 23, 59, 59))

    def test_rango_manual_actual_hasta_el_momento_exacto(self) -> None:
        ahora = datetime(2026, 7, 22, 9, 30, 15)
        inicio, hasta = rango_manual_actual(ahora)
        self.assertEqual(inicio, date(2026, 7, 20))
        self.assertEqual(hasta, ahora)


class TestSemanasPendientes(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.archivo_meta = Path(self._tmp.name) / "_meta.json"

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_sin_exportacion_previa_y_sin_datos_no_hay_pendientes(self) -> None:
        pendientes = semanas_pendientes(
            "prueba", datetime(2026, 7, 22, 10, 0), archivo_meta=self.archivo_meta,
        )
        self.assertEqual(pendientes, [])

    def test_sin_exportacion_previa_usa_fecha_mas_antigua(self) -> None:
        pendientes = semanas_pendientes(
            "prueba",
            datetime(2026, 7, 22, 10, 0),
            fecha_mas_antigua=date(2026, 7, 1),
            archivo_meta=self.archivo_meta,
        )
        # Semana actual empieza el lunes 20/07 -> semanas completas pendientes: 29/06 y 06/07 y 13/07.
        self.assertEqual(pendientes, [date(2026, 6, 29), date(2026, 7, 6), date(2026, 7, 13)])

    def test_semana_actual_nunca_esta_pendiente(self) -> None:
        pendientes = semanas_pendientes(
            "prueba",
            datetime(2026, 7, 22, 10, 0),
            fecha_mas_antigua=date(2026, 7, 20),
            archivo_meta=self.archivo_meta,
        )
        self.assertEqual(pendientes, [])


class TestNombreArchivo(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        _start_export_isolation(self)

    def tearDown(self) -> None:
        _stop_export_isolation(self)
        self._tmp.cleanup()

    def test_formato_de_nombre_semanal(self) -> None:
        config = _config_vacia(titulo="Registro de Desayuno")
        resultado = exportar_periodo(
            config,
            date(2026, 7, 20),
            datetime(2026, 7, 26, 23, 59, 59),
            automatica=True,
            fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta,
            archivo_meta=self.carpeta / "_meta.json",
        )
        self.assertTrue(resultado.ok)
        self.assertEqual(
            resultado.nombre_archivo,
            "Registro_de_Desayuno_2026-07-27_2026-07-20_a_2026-07-26.xlsx",
        )

    def test_no_sobrescribe_evita_colision_con_sufijo(self) -> None:
        config = _config_vacia(titulo="Registro de Desayuno")
        kwargs = dict(
            inicio=date(2026, 7, 20),
            hasta=datetime(2026, 7, 23, 12, 0, 0),
            automatica=False,
            fecha_exportacion=date(2026, 7, 23),
            carpeta_exports=self.carpeta,
            archivo_meta=self.carpeta / "_meta.json",
        )
        primero = exportar_periodo(config, **kwargs)
        segundo = exportar_periodo(config, **kwargs)

        self.assertTrue(primero.ok and segundo.ok)
        self.assertNotEqual(primero.nombre_archivo, segundo.nombre_archivo)
        self.assertTrue(primero.ruta.exists())
        self.assertTrue(segundo.ruta.exists())


class TestGeneracionHojas(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        _start_export_isolation(self)

    def tearDown(self) -> None:
        _stop_export_isolation(self)
        self._tmp.cleanup()

    def test_dias_sin_registros_no_generan_hoja(self) -> None:
        config = ConfiguracionExportacionModulo(
            tipo="prueba",
            titulo_documento="Registro de Prueba",
            obtener_registros=lambda i, h: [_registro(date(2026, 7, 21), "r01")],
        )
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        libro = load_workbook(resultado.ruta)
        # Solo "Info" + la hoja del martes 21/07 con datos; el resto de días no existen.
        self.assertEqual(set(libro.sheetnames), {"Info", "Mar 21-07"})

    def test_varios_registros_mismo_dia_en_una_sola_hoja(self) -> None:
        registros = [
            _registro(date(2026, 7, 21), "r01", hora=time(8, 0)),
            _registro(date(2026, 7, 21), "r02", hora=time(9, 30)),
        ]
        config = ConfiguracionExportacionModulo(
            tipo="prueba",
            titulo_documento="Registro de Prueba",
            obtener_registros=lambda i, h: registros,
        )
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        libro = load_workbook(resultado.ruta)
        self.assertEqual(set(libro.sheetnames), {"Info", "Mar 21-07"})

        hoja = libro["Mar 21-07"]
        valores = [cell.value for row in hoja.iter_rows() for cell in row if cell.value is not None]
        texto = " | ".join(str(v) for v in valores)
        self.assertIn("r01", texto)
        self.assertIn("r02", texto)
        # Ambos identificadores aparecen en la MISMA hoja, no en hojas separadas.

    def test_exportacion_vacia_solo_tiene_hoja_info(self) -> None:
        config = _config_vacia()
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        libro = load_workbook(resultado.ruta)
        self.assertEqual(libro.sheetnames, ["Info"])


class TestExportacionManualParcial(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        _start_export_isolation(self)

    def tearDown(self) -> None:
        _stop_export_isolation(self)
        self._tmp.cleanup()

    def test_exportacion_manual_no_espera_fin_de_semana(self) -> None:
        config = _config_vacia(titulo="Registro de Merma")
        ahora = datetime(2026, 7, 23, 16, 45, 0)  # jueves, semana incompleta
        resultado = exportar_semana_actual(
            config, ahora, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertEqual(
            resultado.nombre_archivo,
            "Registro_de_Merma_2026-07-23_2026-07-20_a_2026-07-23.xlsx",
        )

    def test_exportacion_manual_no_marca_la_semana_como_exportada(self) -> None:
        config = _config_vacia()
        ahora = datetime(2026, 7, 23, 16, 45, 0)
        exportar_semana_actual(config, ahora, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta)
        self.assertIsNone(ultima_semana_exportada("prueba", archivo_meta=self.archivo_meta))


class TestNoDuplicacion(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"

    def tearDown(self) -> None:
        self._tmp.cleanup()

    @patch("app.core.services.exportacion_semanal_service.get_data")
    def test_procesar_pendientes_dos_veces_no_duplica_archivos_ni_actividad(self, mock_get_data) -> None:
        mock_get_data.side_effect = RuntimeError("sin sesión de Streamlit en pruebas")

        config = _config_vacia(titulo="Registro de Actividad")
        ahora = datetime(2026, 7, 22, 8, 0, 0)

        primera_pasada = procesar_pendientes(
            config, ahora, fecha_mas_antigua=date(2026, 7, 6),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        segunda_pasada = procesar_pendientes(
            config, ahora, fecha_mas_antigua=date(2026, 7, 6),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )

        self.assertEqual(len(primera_pasada), 2)  # semanas del 06/07 y 13/07
        self.assertTrue(all(r.ok for r in primera_pasada))
        self.assertEqual(segunda_pasada, [])  # nada pendiente ya: idempotente

        archivos = list((self.carpeta / "prueba").glob("*.xlsx"))
        self.assertEqual(len(archivos), 2)

    @patch("app.core.services.exportacion_semanal_service.persist_data")
    @patch("app.core.services.exportacion_semanal_service.get_data")
    def test_una_exportacion_correcta_registra_una_sola_actividad(
        self, mock_get_data, mock_persist_data,
    ) -> None:
        class _DatosFalsos:
            usuarios: list = []
            usuario_actual_id = ""
            actividades: list = []

        datos = _DatosFalsos()
        mock_get_data.return_value = datos

        config = _config_vacia()
        exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )

        self.assertEqual(len(datos.actividades), 1)
        self.assertEqual(datos.actividades[0].accion, "Exportación")
        mock_persist_data.assert_called_once()

    @patch("app.core.services.exportacion_semanal_service.persist_data")
    @patch("app.core.services.exportacion_semanal_service.get_data")
    def test_fallo_al_obtener_registros_tambien_registra_una_actividad(
        self, mock_get_data, mock_persist_data,
    ) -> None:
        """Fase 7: si `obtener_registros()` lanza una excepción, el intento
        fallido debe quedar igualmente registrado en el Registro de
        actividad (con resultado "con error"), no solo los fallos al guardar
        el archivo."""
        class _DatosFalsos:
            usuarios: list = []
            usuario_actual_id = ""
            actividades: list = []

        datos = _DatosFalsos()
        mock_get_data.return_value = datos

        config = ConfiguracionExportacionModulo(
            tipo="prueba",
            titulo_documento="Registro de Prueba",
            obtener_registros=lambda i, h: (_ for _ in ()).throw(RuntimeError("boom")),
        )
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )

        self.assertFalse(resultado.ok)
        self.assertEqual(len(datos.actividades), 1)
        self.assertEqual(datos.actividades[0].accion, "Exportación")
        self.assertEqual(datos.actividades[0].resultado, "Error")
        mock_persist_data.assert_called_once()


if __name__ == "__main__":
    unittest.main()
