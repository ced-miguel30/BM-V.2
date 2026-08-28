"""Tests exportacion_registro_operativo_service."""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime, time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.application.actor import Actor
from app.core.application.clock import FixedClock
from app.core.application.context import build_app_context
from app.core.application.unit_of_work import InMemoryUnitOfWork
from app.core.models import AppData
from app.core.models.buffet import LineaRegistroBuffet, RegistroBuffetDiario
from app.core.models.desayuno import LineaDesayuno, RegistroDesayuno
from app.core.services.exportacion_registro_operativo_service import (
    exportar_semana_registro_operativo,
)
from tests.demo_isolation import EXPORT_SESSION_MODULES, isolated_persist


class TestExportacionRegistroOperativo(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData(
            desayunos=[
                RegistroDesayuno(
                    "d1",
                    date(2026, 8, 11),
                    [LineaDesayuno("p1", 1.0, 1.0, True)],
                    1.0,
                    "Tester",
                    1,
                    [],
                    time(8, 0),
                )
            ],
            registros_buffet=[
                RegistroBuffetDiario(
                    "bf1",
                    date(2026, 8, 11),
                    [LineaRegistroBuffet("cb1", "Kiwi", 1.0, "consumo", coste_snapshot=0.5)],
                    0.5,
                )
            ],
        )
        self._iso = isolated_persist(*EXPORT_SESSION_MODULES, data=self.data)
        self._iso.__enter__()
        self.addCleanup(self._iso.__exit__, None, None, None)
        self.ctx = build_app_context(
            uow=InMemoryUnitOfWork(self.data),
            clock=FixedClock(datetime(2026, 8, 15, 10, 0, 0)),
            actor=Actor(id="u1", nombre="Tester", rol="Admin"),
        )
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)

    def test_libro_unificado_contiene_hojas(self) -> None:
        import app.core.services.exportacion_registro_operativo_service as svc

        out = Path(self._tmp.name)
        with unittest.mock.patch.object(svc, "_dir_salida", return_value=out):
            res = exportar_semana_registro_operativo(
                date(2026, 8, 10), date(2026, 8, 16), ctx=self.ctx,
            )
        self.assertTrue(res.ok, res.mensaje)
        assert res.ruta is not None
        wb = load_workbook(res.ruta)
        for hoja in ("Info", "Desayuno", "Comida", "Cena", "ConsumoBuffet"):
            self.assertIn(hoja, wb.sheetnames)
        wb.close()


if __name__ == "__main__":
    unittest.main()
