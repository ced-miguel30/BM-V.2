"""Pruebas de la Fase 4: exportación semanal de stock y bebidas.

Ejecutar desde la raíz del proyecto con:

    py -m unittest discover -s tests -v
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.models import AppData, LoteStock, Producto, UnidadProducto
from app.core.services import stock_service
from app.core.services.excel_bloques import nombre_hoja_dia
from app.core.services.exportacion_semanal_service import exportar_periodo


def _catalogo() -> list[Producto]:
    return [
        Producto("p01", "Pan", UnidadProducto.UD, es_bebida=False),
        Producto("b01", "Zumo", UnidadProducto.L, es_bebida=True),
    ]


def _lotes() -> list[LoteStock]:
    return [
        LoteStock("l01", "p01", 10.0, 20.0, 20.0, date(2026, 7, 20), None, "Panadería"),
        LoteStock("l02", "p01", 5.0, 10.0, 8.0, date(2026, 7, 22), date(2026, 8, 1), "Proveedor A"),
        LoteStock("l03", "b01", 12.0, 6.0, 6.0, date(2026, 7, 21), date(2026, 9, 1), "Bebidas SL"),
        LoteStock("l04", "p01", 3.0, 5.0, 5.0, None, None, "Sin fecha"),
    ]


class TestRegistrosExportablesStock(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData(productos=_catalogo(), lotes=_lotes())
        self._patcher = patch("app.core.services.stock_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()

    def test_stock_excluye_bebidas_y_lotes_sin_fecha(self) -> None:
        registros = stock_service.registros_exportables(
            date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59), es_bebida=False,
        )
        ids = [r.identificador for r in registros]
        self.assertEqual(ids, ["l01", "l02"])
        self.assertNotIn("l03", ids)
        self.assertNotIn("l04", ids)

    def test_bebidas_solo_incluye_catalogo_de_bebidas(self) -> None:
        registros = stock_service.registros_exportables(
            date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59), es_bebida=True,
        )
        self.assertEqual(len(registros), 1)
        self.assertEqual(registros[0].identificador, "l03")
        self.assertEqual(registros[0].tipo, "Bebida")
        self.assertEqual(registros[0].columnas[0], "Bebida")

    def test_columnas_y_desglose_de_un_lote(self) -> None:
        registros = stock_service.registros_exportables(
            date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59), es_bebida=False,
        )
        registro = next(r for r in registros if r.identificador == "l01")
        self.assertEqual(registro.tipo, "Stock")
        self.assertEqual(registro.fecha, date(2026, 7, 20))
        fila = registro.filas[0]
        self.assertEqual(fila[0], "Pan")
        self.assertEqual(fila[1], "Panadería")
        self.assertEqual(fila[2], "l01")
        self.assertEqual(fila[3], 20.0)
        self.assertEqual(fila[4], "Ud")
        self.assertEqual(fila[8], "Compra")

    def test_filtro_por_rango_parcial(self) -> None:
        registros = stock_service.registros_exportables(
            date(2026, 7, 20), datetime(2026, 7, 21, 23, 59, 59), es_bebida=False,
        )
        self.assertEqual([r.identificador for r in registros], ["l01"])

    def test_fecha_mas_antigua_por_tipo(self) -> None:
        self.assertEqual(stock_service.fecha_mas_antigua(es_bebida=False), date(2026, 7, 20))
        self.assertEqual(stock_service.fecha_mas_antigua(es_bebida=True), date(2026, 7, 21))


class TestIntegracionExportacionStockBebidas(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        self.data = AppData(productos=_catalogo(), lotes=_lotes())
        self._patcher = patch("app.core.services.stock_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()
        self._tmp.cleanup()

    def test_exportacion_stock_genera_hojas_por_dia(self) -> None:
        config = stock_service.configuracion_exportacion(es_bebida=False)
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertIn("Registro_de_Stock", resultado.nombre_archivo)

        libro = load_workbook(resultado.ruta)
        self.assertEqual(
            set(libro.sheetnames),
            {"Info", nombre_hoja_dia(date(2026, 7, 20)), nombre_hoja_dia(date(2026, 7, 22))},
        )

    def test_exportacion_bebidas_separada_de_stock(self) -> None:
        config = stock_service.configuracion_exportacion(es_bebida=True)
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertIn("Registro_de_Bebidas", resultado.nombre_archivo)

        libro = load_workbook(resultado.ruta)
        hoja = nombre_hoja_dia(date(2026, 7, 21))
        self.assertEqual(set(libro.sheetnames), {"Info", hoja})
        valores = " | ".join(
            str(c.value) for row in libro[hoja].iter_rows() for c in row if c.value is not None
        )
        self.assertIn("l03", valores)
        self.assertIn("Zumo", valores)


if __name__ == "__main__":
    unittest.main()
