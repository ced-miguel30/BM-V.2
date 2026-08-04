"""Pruebas de la Fase 6: exportación semanal del Registro de actividad.

Ejecutar desde la raíz del proyecto con:

    py -m unittest discover -s tests -v

No dependen de una sesión de Streamlit real: se inyecta un `AppData`
sintético parcheando `get_data` en cada servicio.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.models import Actividad, AppData
from app.core.services import actividad_service
from app.core.services.exportacion_semanal_service import exportar_periodo
from app.data.serializers import appdata_to_dict, dict_to_appdata


class TestRegistrosExportablesActividad(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData()
        self._patcher = patch("app.core.services.actividad_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()

    def test_columnas_del_registro_de_actividad(self) -> None:
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 21, 8, 30), "Ana", "Registro desayuno", "Desayuno del 21/07 — 4.90 €"),
        ]
        registros = actividad_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 1)
        registro = registros[0]
        self.assertEqual(registro.tipo, "Registro desayuno")
        self.assertEqual(registro.identificador, "act01")
        self.assertEqual(registro.usuario, "Ana")
        self.assertEqual(
            registro.columnas,
            ["Módulo", "Descripción", "Resultado", "Tipo de exportación", "Periodo afectado", "Archivo generado"],
        )

    def test_actividad_normal_deja_vacios_los_campos_solo_de_exportacion(self) -> None:
        # "cuando esté disponible": una acción normal (no una exportación) no
        # tiene módulo/resultado/tipo de exportación/periodo/archivo.
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 21, 8, 30), "Ana", "Registro desayuno", "Desayuno del 21/07 — 4.90 €"),
        ]
        registros = actividad_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        fila = registros[0].filas[0]
        self.assertEqual(fila, ["", "Desayuno del 21/07 — 4.90 €", "", "", "", ""])

    def test_actividad_de_exportacion_muestra_los_campos_estructurados(self) -> None:
        self.data.actividades = [
            Actividad(
                "act01", datetime(2026, 7, 21, 0, 0), "Sistema", "Exportación",
                "Exportación automática de desayuno — periodo 2026-07-13 a 2026-07-19 — "
                "archivo Registro_de_Desayuno_x.xlsx — resultado correcto",
                modulo="desayuno",
                resultado="Correcto",
                tipo_exportacion="Automática",
                periodo_afectado="2026-07-13 a 2026-07-19",
                archivo_generado="Registro_de_Desayuno_x.xlsx",
            ),
        ]
        registros = actividad_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        fila = registros[0].filas[0]
        self.assertEqual(fila[0], "desayuno")
        self.assertEqual(fila[2], "Correcto")
        self.assertEqual(fila[3], "Automática")
        self.assertEqual(fila[4], "2026-07-13 a 2026-07-19")
        self.assertEqual(fila[5], "Registro_de_Desayuno_x.xlsx")

    def test_varias_actividades_del_mismo_dia_no_se_fusionan(self) -> None:
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 21, 8, 0), "Ana", "Registro desayuno", "..."),
            Actividad("act02", datetime(2026, 7, 21, 9, 0), "Luis", "Registro merma", "..."),
        ]
        registros = actividad_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 2)
        self.assertEqual({r.identificador for r in registros}, {"act01", "act02"})

    def test_dias_sin_actividad_no_generan_registros(self) -> None:
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 21, 8, 0), "Ana", "Registro desayuno", "..."),
        ]
        registros = actividad_service.registros_exportables(date(2026, 7, 27), datetime(2026, 8, 2, 23, 59, 59))
        self.assertEqual(registros, [])

    def test_fecha_mas_antigua(self) -> None:
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 6, 8, 0), "Ana", "Registro desayuno", "..."),
            Actividad("act02", datetime(2026, 7, 21, 8, 0), "Ana", "Registro desayuno", "..."),
        ]
        self.assertEqual(actividad_service.fecha_mas_antigua(), date(2026, 7, 6))

    def test_fecha_mas_antigua_sin_datos_es_none(self) -> None:
        self.assertIsNone(actividad_service.fecha_mas_antigua())


class TestSinBucleDeExportacion(unittest.TestCase):
    """La exportación del propio Registro de actividad no debe incluirse a sí
    misma: `exportar_periodo` obtiene los registros ANTES de registrar la
    actividad de "Exportación" que documenta el resultado."""

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        self.data = AppData()
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 21, 8, 0), "Ana", "Registro desayuno", "Desayuno registrado."),
        ]
        self._p_actividad = patch("app.core.services.actividad_service.get_data", return_value=self.data)
        self._p_motor = patch("app.core.services.exportacion_semanal_service.get_data", return_value=self.data)
        self._p_persist = patch("app.core.services.exportacion_semanal_service.persist_data")
        self._p_actividad.start()
        self._p_motor.start()
        self._p_persist.start()

    def tearDown(self) -> None:
        self._p_actividad.stop()
        self._p_motor.stop()
        self._p_persist.stop()
        self._tmp.cleanup()

    def test_exportacion_no_se_incluye_a_si_misma_y_se_registra_una_sola_vez(self) -> None:
        config = actividad_service.configuracion_exportacion()
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertEqual(resultado.filas_exportadas, 1)  # solo "act01", no la propia exportación

        nuevas_exportaciones = [a for a in self.data.actividades if a.accion == "Exportación"]
        self.assertEqual(len(nuevas_exportaciones), 1)
        self.assertEqual(nuevas_exportaciones[0].modulo, "actividad")
        self.assertEqual(nuevas_exportaciones[0].resultado, "Correcto")

        libro = load_workbook(resultado.ruta)
        hoja = libro["Mar 21-07"]
        valores = " | ".join(
            str(c.value) for row in hoja.iter_rows() for c in row if c.value is not None
        )
        self.assertIn("act01", valores)
        self.assertNotIn("act02", valores)  # la nueva actividad de exportación sería "act02"


class TestIntegracionExportacionActividad(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        self.data = AppData()
        self._patcher = patch("app.core.services.actividad_service.get_data", return_value=self.data)
        self._p_motor = patch(
            "app.core.services.exportacion_semanal_service.get_data", return_value=self.data
        )
        self._p_persist = patch(
            "app.core.services.exportacion_semanal_service.persist_data",
            side_effect=lambda d=None: d if d is not None else self.data,
        )
        self._patcher.start()
        self._p_motor.start()
        self._p_persist.start()

    def tearDown(self) -> None:
        self._p_persist.stop()
        self._p_motor.stop()
        self._patcher.stop()
        self._tmp.cleanup()

    def test_exportacion_semanal_genera_una_hoja_por_dia(self) -> None:
        self.data.actividades = [
            Actividad("act01", datetime(2026, 7, 20, 8, 0), "Ana", "Registro desayuno", "..."),
            Actividad("act02", datetime(2026, 7, 23, 9, 0), "Luis", "Registro merma", "..."),
        ]
        config = actividad_service.configuracion_exportacion()
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertEqual(
            resultado.nombre_archivo,
            "Registro_de_Actividad_2026-07-27_2026-07-20_a_2026-07-26.xlsx",
        )
        libro = load_workbook(resultado.ruta)
        self.assertEqual(set(libro.sheetnames), {"Info", "Lun 20-07", "Jue 23-07"})


class TestCompatibilidadSerializador(unittest.TestCase):
    """Las actividades antiguas (sin los campos estructurados de la Fase 6)
    deben seguir cargando correctamente, con esos campos en `None`."""

    def test_actividad_antigua_sin_campos_nuevos_se_deserializa(self) -> None:
        payload = {
            "productos": [], "lotes": [], "recetas": [], "desayunos": [], "mermas": [],
            "alertas": [], "alertas_descartadas": [], "usuarios": [], "configuracion": None,
            "actividades": [
                {
                    "id": "act01", "fecha_hora": "2026-07-21T08:00:00",
                    "usuario": "Ana", "accion": "Registro desayuno", "detalle": "...",
                },
            ],
            "meta": {},
        }
        data = dict_to_appdata(payload)
        self.assertEqual(len(data.actividades), 1)
        actividad = data.actividades[0]
        self.assertIsNone(actividad.modulo)
        self.assertIsNone(actividad.resultado)
        self.assertIsNone(actividad.tipo_exportacion)

    def test_ida_y_vuelta_conserva_los_campos_estructurados(self) -> None:
        data = AppData(actividades=[
            Actividad(
                "act01", datetime(2026, 7, 21, 8, 0), "Sistema", "Exportación", "...",
                modulo="stock", resultado="Correcto", tipo_exportacion="Automática",
                periodo_afectado="2026-07-13 a 2026-07-19", archivo_generado="Registro_de_Stock_x.xlsx",
            ),
        ])
        payload = appdata_to_dict(data)
        data2 = dict_to_appdata(payload)
        actividad = data2.actividades[0]
        self.assertEqual(actividad.modulo, "stock")
        self.assertEqual(actividad.resultado, "Correcto")
        self.assertEqual(actividad.tipo_exportacion, "Automática")
        self.assertEqual(actividad.periodo_afectado, "2026-07-13 a 2026-07-19")
        self.assertEqual(actividad.archivo_generado, "Registro_de_Stock_x.xlsx")


if __name__ == "__main__":
    unittest.main()
