"""Pruebas de la Fase 5: rankings de consumo (productos, recetas, bebidas) y
exportación del «Registro de Consumo».

Ejecutar desde la raíz del proyecto con:

    py -m unittest discover -s tests -v

No dependen de una sesión de Streamlit real: se inyecta un `AppData`
sintético parcheando `get_data` en cada servicio.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime, time
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.models import (
    AppData,
    IngredienteReceta,
    LineaDesayuno,
    Producto,
    Receta,
    RegistroDesayuno,
    RegistroRecetaDesayuno,
    UnidadProducto,
)
from app.core.services import consumo_service
from app.core.services.exportacion_semanal_service import exportar_periodo


def _productos() -> list[Producto]:
    return [
        Producto("p01", "Café", UnidadProducto.L, es_bebida=True),
        Producto("p02", "Leche", UnidadProducto.L, es_bebida=True),
        Producto("p03", "Pan", UnidadProducto.UD),
        Producto("p04", "Mantequilla", UnidadProducto.KG),
        Producto("p05", "Zumo de naranja", UnidadProducto.L, es_bebida=True),
    ]


def _recetas() -> list[Receta]:
    return [
        Receta("r01", "Café con leche", ingredientes=[
            IngredienteReceta("p01", 0.1),
            IngredienteReceta("p02", 0.05),
        ]),
    ]


class TestRankingProductosYBebidas(unittest.TestCase):
    """Un mismo producto consumido directamente y a la vez como ingrediente
    de receta no debe contabilizarse dos veces: `lineas` ya trae la cantidad
    neta fusionada, así que el ranking debe usar exclusivamente esa fuente."""

    def setUp(self) -> None:
        self.data = AppData(productos=_productos(), recetas=_recetas())
        self._patcher_repo = patch("app.core.services.data_service.get_data", return_value=self.data)
        self._patcher_repo.start()

    def tearDown(self) -> None:
        self._patcher_repo.stop()

    def test_producto_directo_y_via_receta_no_se_duplica(self) -> None:
        # "Pan" se consume 3 veces de forma directa (sin receta) en la misma
        # cesta ya fusionada por `_aplanar_cesta`: una única línea neta.
        self.data.desayunos = [
            RegistroDesayuno(
                "d01", date(2026, 7, 21),
                lineas=[LineaDesayuno("p03", 3.0, 1.0, False)],
            ),
        ]
        ranking = consumo_service.ranking_productos_periodo(date(2026, 7, 20), date(2026, 7, 26))
        self.assertEqual(len(ranking), 1)
        self.assertEqual(ranking[0]["nombre"], "Pan")
        self.assertEqual(ranking[0]["cantidad"], 3.0)
        self.assertEqual(ranking[0]["usos"], 1)

    def test_bebida_directa_y_bebida_dentro_de_receta_se_agrega_una_vez_por_registro(self) -> None:
        # "Café con leche" (receta) aporta café + leche; en `lineas` ya viene
        # la cantidad neta de cada ingrediente (base receta), sin volver a
        # sumarla aparte al reconstruir el desglose de receta.
        self.data.desayunos = [
            RegistroDesayuno(
                "d01", date(2026, 7, 21),
                lineas=[
                    LineaDesayuno("p01", 1.0, 0.5, False),  # café (vía receta, ×10 porciones)
                    LineaDesayuno("p02", 0.5, 0.3, False),  # leche (vía receta)
                ],
                registros_recetas=[RegistroRecetaDesayuno("r01", "Café con leche", 10.0)],
            ),
        ]
        bebidas = consumo_service.ranking_productos_periodo(date(2026, 7, 20), date(2026, 7, 26), es_bebida=True)
        nombres = {b["nombre"] for b in bebidas}
        self.assertEqual(nombres, {"Café", "Leche"})
        cafe = next(b for b in bebidas if b["nombre"] == "Café")
        self.assertEqual(cafe["cantidad"], 1.0)

        # Y la receta en sí no debe aparecer como una "bebida" independiente.
        self.assertNotIn("Café con leche", nombres)

    def test_bebida_directa_mas_bebida_vía_receta_del_mismo_producto_se_suma_una_sola_vez(self) -> None:
        # El café se toma directo (sin receta) Y también aparece dentro de una
        # receta en OTRO registro: cada registro aporta su propia línea neta
        # (ya fusionada), así que la suma total no duplica ninguna cantidad.
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 21), lineas=[LineaDesayuno("p01", 1.0, 0.5, False)]),
            RegistroDesayuno(
                "d02", date(2026, 7, 22),
                lineas=[LineaDesayuno("p01", 2.0, 1.0, False)],
                registros_recetas=[RegistroRecetaDesayuno("r01", "Café con leche", 20.0)],
            ),
        ]
        bebidas = consumo_service.ranking_productos_periodo(date(2026, 7, 20), date(2026, 7, 26), es_bebida=True)
        cafe = next(b for b in bebidas if b["nombre"] == "Café")
        self.assertEqual(cafe["cantidad"], 3.0)
        self.assertEqual(cafe["usos"], 2)

    def test_productos_menos_consumidos_excluye_los_de_consumo_cero(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 21), lineas=[
                LineaDesayuno("p03", 10.0, 3.0, False),
                LineaDesayuno("p04", 0.2, 1.0, False),
            ]),
        ]
        # "Café", "Leche" y "Zumo" no tuvieron consumo: no deben aparecer.
        menos = consumo_service.ranking_productos_periodo(
            date(2026, 7, 20), date(2026, 7, 26), es_bebida=False, ascendente=True,
        )
        nombres = [f["nombre"] for f in menos]
        self.assertEqual(nombres, ["Mantequilla", "Pan"])  # ascendente por cantidad

    def test_varias_unidades_compatibles_se_muestran_en_unidad_legible(self) -> None:
        # Mantequilla nativa en kg; cantidades pequeñas se muestran en gramos.
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 21), lineas=[LineaDesayuno("p04", 0.05, 0.5, False)]),
            RegistroDesayuno("d02", date(2026, 7, 22), lineas=[LineaDesayuno("p04", 0.03, 0.3, False)]),
        ]
        ranking = consumo_service.ranking_productos_periodo(date(2026, 7, 20), date(2026, 7, 26))
        self.assertEqual(len(ranking), 1)
        self.assertAlmostEqual(ranking[0]["cantidad"], 80.0)
        self.assertEqual(ranking[0]["unidad"], "gr")

    def test_periodo_filtra_por_fecha(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 6), lineas=[LineaDesayuno("p03", 5.0, 1.0, False)]),
            RegistroDesayuno("d02", date(2026, 7, 21), lineas=[LineaDesayuno("p03", 3.0, 1.0, False)]),
        ]
        ranking = consumo_service.ranking_productos_periodo(date(2026, 7, 20), date(2026, 7, 26))
        self.assertEqual(ranking[0]["cantidad"], 3.0)


class TestRankingRecetas(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData(productos=_productos(), recetas=_recetas())
        self._patcher_repo = patch("app.core.services.data_service.get_data", return_value=self.data)
        self._patcher_repo.start()

    def tearDown(self) -> None:
        self._patcher_repo.stop()

    def test_receta_con_varios_ingredientes_cuenta_como_una_sola_unidad_de_consumo(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno(
                "d01", date(2026, 7, 21),
                lineas=[LineaDesayuno("p01", 1.0, 0.5, False), LineaDesayuno("p02", 0.5, 0.3, False)],
                registros_recetas=[RegistroRecetaDesayuno("r01", "Café con leche", 10.0)],
            ),
            RegistroDesayuno(
                "d02", date(2026, 7, 22),
                lineas=[LineaDesayuno("p01", 2.0, 1.0, False), LineaDesayuno("p02", 1.0, 0.6, False)],
                registros_recetas=[RegistroRecetaDesayuno("r01", "Café con leche", 20.0)],
            ),
        ]
        ranking = consumo_service.ranking_recetas_periodo(date(2026, 7, 20), date(2026, 7, 26))
        self.assertEqual(len(ranking), 1)
        self.assertEqual(ranking[0]["nombre"], "Café con leche")
        self.assertEqual(ranking[0]["cantidad"], 30.0)
        self.assertEqual(ranking[0]["usos"], 2)


class TestExportacionConsumo(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData(productos=_productos(), recetas=_recetas())
        self._patcher = patch("app.core.services.consumo_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()

    def test_columnas_y_tipos_del_registro_de_consumo(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno(
                "d01", date(2026, 7, 21),
                lineas=[
                    LineaDesayuno("p01", 1.0, 0.5, False),
                    LineaDesayuno("p02", 0.5, 0.3, False),
                    LineaDesayuno("p03", 5.0, 1.5, False),
                ],
                registros_recetas=[RegistroRecetaDesayuno("r01", "Café con leche", 10.0)],
                registrado_por="Ana",
                hora=time(8, 30),
            ),
        ]
        registros = consumo_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 1)
        registro = registros[0]
        self.assertEqual(registro.tipo, "Consumo")
        self.assertEqual(
            registro.columnas,
            ["Tipo", "Producto o receta", "Cantidad visible", "Unidad visible",
             "Cantidad interna", "Unidad interna", "Coste", "Relación con receta"],
        )

        filas_por_nombre = {fila[1]: fila for fila in registro.filas}
        self.assertEqual(filas_por_nombre["Café con leche"][0], "Receta")
        self.assertEqual(filas_por_nombre["Café"][0], "Bebida")
        self.assertEqual(filas_por_nombre["Café"][7], "Café con leche")
        self.assertEqual(filas_por_nombre["Pan"][0], "Producto")
        self.assertEqual(filas_por_nombre["Pan"][7], "Directo")

    def test_dias_sin_registros_no_generan_filas(self) -> None:
        self.data.desayunos = [RegistroDesayuno("d01", date(2026, 7, 21))]
        registros = consumo_service.registros_exportables(date(2026, 7, 27), datetime(2026, 8, 2, 23, 59, 59))
        self.assertEqual(registros, [])

    def test_fecha_mas_antigua_reutiliza_la_de_desayuno(self) -> None:
        with patch("app.core.services.desayuno_service.get_data", return_value=self.data):
            self.data.desayunos = [
                RegistroDesayuno("d01", date(2026, 7, 6), lineas=[LineaDesayuno("p03", 1.0, 1.0)]),
                RegistroDesayuno("d02", date(2026, 7, 21), lineas=[LineaDesayuno("p03", 1.0, 1.0)]),
            ]
            self.assertEqual(consumo_service.fecha_mas_antigua(), date(2026, 7, 6))


class TestIntegracionExportacionConsumo(unittest.TestCase):
    """De extremo a extremo: registros de desayuno -> Excel real generado por
    el motor central (Fase 1), usando el builder de consumo (Fase 5)."""

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        self.data = AppData(productos=_productos(), recetas=_recetas())
        self._patcher = patch("app.core.services.consumo_service.get_data", return_value=self.data)
        self._patcher.start()
        from tests.demo_isolation import EXPORT_SESSION_MODULES, isolated_persist

        self._export_iso = isolated_persist(*EXPORT_SESSION_MODULES, data=self.data)
        self._export_iso.__enter__()

    def tearDown(self) -> None:
        self._export_iso.__exit__(None, None, None)
        self._patcher.stop()
        self._tmp.cleanup()

    def test_exportacion_semanal_genera_una_hoja_por_dia(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 20), lineas=[LineaDesayuno("p03", 1.0, 1.0)], hora=time(8, 0)),
            RegistroDesayuno("d02", date(2026, 7, 23), lineas=[LineaDesayuno("p04", 0.2, 1.0)], hora=time(8, 30)),
        ]
        config = consumo_service.configuracion_exportacion()
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertEqual(
            resultado.nombre_archivo,
            "Registro_de_Consumo_2026-07-27_2026-07-20_a_2026-07-26.xlsx",
        )

        libro = load_workbook(resultado.ruta)
        self.assertEqual(set(libro.sheetnames), {"Info", "Lun 20-07", "Jue 23-07"})


if __name__ == "__main__":
    unittest.main()
