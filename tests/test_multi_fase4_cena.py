"""Pruebas Fase 4 — registro de cena.

Ejecutar:

    py -m unittest tests.test_multi_fase4_cena -v
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.models import (
    AppData,
    CategoriaReceta,
    IngredienteReceta,
    LoteStock,
    OrigenConsumo,
    Producto,
    Receta,
    UnidadProducto,
    Usuario,
)
from app.core.models.enums import RolUsuario
from app.core.services import cena_service
from app.core.services.excel_bloques import nombre_hoja_dia
from app.core.services.exportacion_semanal_service import exportar_periodo
from app.core.services.servicio_registro_service import crear_servicio
from tests.demo_isolation import EXPORT_SESSION_MODULES, isolated_persist


def _datos() -> AppData:
    return AppData(
        productos=[
            Producto("p01", "Pescado", UnidadProducto.KG),
            Producto("p02", "Vino", UnidadProducto.L, es_bebida=True),
            Producto("p03", "Pan", UnidadProducto.KG),
        ],
        lotes=[
            LoteStock("l01", "p01", 30.0, 2.0, 2.0, fecha_compra=date(2026, 7, 1)),
            LoteStock("l02", "p02", 10.0, 2.0, 2.0, fecha_compra=date(2026, 7, 1)),
            LoteStock("l03", "p03", 5.0, 1.0, 1.0, fecha_compra=date(2026, 7, 1)),
        ],
        recetas=[
            Receta("r_cen", "Merluza", [IngredienteReceta("p01", 0.25)], CategoriaReceta.CENA, porciones_estandar=1.0),
            Receta("r_beb", "Copa vino", [IngredienteReceta("p02", 0.15)], CategoriaReceta.BEBIDAS, porciones_estandar=1.0),
            Receta("r_des", "Tostada", [IngredienteReceta("p03", 0.1)], CategoriaReceta.DESAYUNO, porciones_estandar=1.0),
            Receta("r_com", "Paella", [IngredienteReceta("p01", 0.3)], CategoriaReceta.COMIDA, porciones_estandar=1.0),
        ],
        usuarios=[Usuario("u01", "Ana", RolUsuario.ADMIN, True)],
        usuario_actual_id="u01",
    )


class _FakeSession(dict):
    def __init__(self, store: dict):
        super().__init__()
        self._store = store

    def __setitem__(self, key, value):
        self._store[key] = value
        super().__setitem__(key, value)

    def get(self, key, default=None):
        return self._store.get(key, default)

    def __contains__(self, key):
        return key in self._store


class TestCenaRegistro(unittest.TestCase):
    def setUp(self) -> None:
        self.data = _datos()
        self.state: dict = {}
        fake_st = MagicMock()
        fake_st.session_state = _FakeSession(self.state)

        self.patches = [
            patch("app.core.services.servicio_registro_service.get_data", return_value=self.data),
            patch("app.core.services.servicio_registro_service.persist_data"),
            patch("app.core.services.cesta_service.get_data", return_value=self.data),
            patch("app.core.services.alert_service.sincronizar_alertas"),
            patch.dict("sys.modules", {"streamlit": fake_st}),
        ]
        for p in self.patches:
            p.start()
        import streamlit as st
        st.session_state = fake_st.session_state

        cena_service.servicio = crear_servicio(
            "cena", "cena",
            [CategoriaReceta.CENA, CategoriaReceta.BEBIDAS],
            titulo_documento="Registro de Cena",
        )
        for nombre in (
            "anadir_a_cesta", "anadir_receta_a_cesta", "registrar",
            "historial_ordenado", "fecha_mas_antigua", "registros_exportables",
            "configuracion_exportacion", "limpiar_cesta",
        ):
            setattr(cena_service, nombre, getattr(cena_service.servicio, nombre))

    def tearDown(self) -> None:
        for p in self.patches:
            p.stop()

    def test_acepta_receta_cena_y_bebidas(self) -> None:
        self.assertTrue(cena_service.anadir_receta_a_cesta("r_cen", 1.0).ok)
        self.assertTrue(cena_service.anadir_receta_a_cesta("r_beb", 1.0).ok)

    def test_rechaza_desayuno_y_comida(self) -> None:
        self.assertFalse(cena_service.anadir_receta_a_cesta("r_des", 1.0).ok)
        self.assertFalse(cena_service.anadir_receta_a_cesta("r_com", 1.0).ok)

    def test_registrar_descuento_e_origen(self) -> None:
        self.assertTrue(cena_service.anadir_receta_a_cesta("r_cen", 1.0).ok)
        self.assertTrue(cena_service.anadir_a_cesta("p03", 0.1).ok)
        resultado = cena_service.registrar(date(2026, 7, 21))
        self.assertTrue(resultado.ok, resultado.mensaje)

        reg = self.data.registros_servicio[0]
        self.assertEqual(reg.tipo_servicio, "cena")

        pescado = next(l for l in self.data.lotes if l.id == "l01")
        pan = next(l for l in self.data.lotes if l.id == "l03")
        self.assertAlmostEqual(pescado.cantidad_restante, 1.75, places=4)
        self.assertAlmostEqual(pan.cantidad_restante, 0.9, places=4)

        origenes = {d.origen for d in reg.lineas_detalle}
        self.assertIn(OrigenConsumo.INGREDIENTE_RECETA.value, origenes)
        self.assertIn(OrigenConsumo.PRODUCTO_DIRECTO.value, origenes)
        self.assertTrue(all(d.tipo_servicio == "cena" for d in reg.lineas_detalle))
        self.assertTrue(any("cena" in a.accion.lower() for a in self.data.actividades))

    def test_historial_independiente_de_comida(self) -> None:
        self.assertTrue(cena_service.anadir_receta_a_cesta("r_cen", 1.0).ok)
        self.assertTrue(cena_service.registrar(date(2026, 7, 21)).ok)
        # Simula un registro de otro tipo ya presente
        from app.core.models import LineaServicio, RegistroServicio
        self.data.registros_servicio.append(RegistroServicio(
            "co99", "comida", date(2026, 7, 21),
            [LineaServicio("p01", 0.1, 1.0)],
            1.0, "Ana",
        ))
        historial = cena_service.historial_ordenado()
        self.assertEqual(len(historial), 1)
        self.assertEqual(historial[0].tipo_servicio, "cena")

    def test_exportacion_titulo_cena(self) -> None:
        self.assertTrue(cena_service.anadir_receta_a_cesta("r_cen", 1.0).ok)
        self.assertTrue(cena_service.registrar(date(2026, 7, 22)).ok)

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            with isolated_persist(*EXPORT_SESSION_MODULES, data=self.data):
                resultado = exportar_periodo(
                    cena_service.configuracion_exportacion(),
                    date(2026, 7, 20),
                    datetime(2026, 7, 26, 23, 59, 59),
                    automatica=True,
                    carpeta_exports=carpeta,
                    archivo_meta=carpeta / "_meta.json",
                )
            self.assertTrue(resultado.ok, resultado.mensaje)
            self.assertTrue(resultado.nombre_archivo.startswith("Registro_de_Cena_"))
            from openpyxl import load_workbook
            libro = load_workbook(resultado.ruta)
            self.assertIn(nombre_hoja_dia(date(2026, 7, 22)), libro.sheetnames)


if __name__ == "__main__":
    unittest.main()
