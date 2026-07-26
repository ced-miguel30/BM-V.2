"""Pruebas Fase 5 — registro de bebidas.

Ejecutar:

    py -m unittest tests.test_multi_fase5_bebidas -v
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.models import (
    AppData,
    CategoriaReceta,
    IngredienteReceta,
    LoteStock,
    OrigenConsumo,
    Producto,
    Receta,
    UnidadProducto,
    Usuario,
)
from app.core.models.enums import RolUsuario
from app.core.services import bebida_service
from app.core.services.excel_bloques import nombre_hoja_dia
from app.core.services.exportacion_semanal_service import exportar_periodo
from app.core.services.servicio_registro_service import crear_servicio


def _datos() -> AppData:
    return AppData(
        productos=[
            Producto("p01", "Pan", UnidadProducto.KG, es_bebida=False),
            Producto(
                "p02", "Café", UnidadProducto.L, es_bebida=True,
                servicios_disponibles=["bebidas"],
            ),
            Producto(
                "p03", "Zumo", UnidadProducto.L, es_bebida=True,
                servicios_disponibles=["bebidas"],
            ),
        ],
        lotes=[
            LoteStock("l01", "p01", 5.0, 1.0, 1.0, fecha_compra=date(2026, 7, 1)),
            LoteStock("l02", "p02", 8.0, 2.0, 2.0, fecha_compra=date(2026, 7, 1)),
            LoteStock("l03", "p03", 4.0, 2.0, 2.0, fecha_compra=date(2026, 7, 1)),
        ],
        recetas=[
            Receta(
                "r_beb", "Café latte", [IngredienteReceta("p02", 0.2)],
                CategoriaReceta.BEBIDAS, servicios_disponibles=["bebidas"],
                porciones_estandar=1.0,
            ),
            Receta("r_des", "Tostada", [IngredienteReceta("p01", 0.1)], CategoriaReceta.DESAYUNO, porciones_estandar=1.0),
            Receta("r_com", "Paella", [IngredienteReceta("p01", 0.3)], CategoriaReceta.COMIDA, porciones_estandar=1.0),
            Receta("r_cen", "Sopa", [IngredienteReceta("p01", 0.2)], CategoriaReceta.CENA, porciones_estandar=1.0),
        ],
        usuarios=[Usuario("u01", "Ana", RolUsuario.ADMIN, True)],
        usuario_actual_id="u01",
    )


class _FakeSession(dict):
    def __init__(self, store: dict):
        super().__init__()
        self._store = store

    def __setitem__(self, key, value):
        self._store[key] = value
        super().__setitem__(key, value)

    def get(self, key, default=None):
        return self._store.get(key, default)

    def __contains__(self, key):
        return key in self._store


class TestBebidasRegistro(unittest.TestCase):
    def setUp(self) -> None:
        self.data = _datos()
        self.state: dict = {}
        fake_st = MagicMock()
        fake_st.session_state = _FakeSession(self.state)

        self.patches = [
            patch("app.core.services.servicio_registro_service.get_data", return_value=self.data),
            patch("app.core.services.servicio_registro_service.persist_data"),
            patch("app.core.services.cesta_service.get_data", return_value=self.data),
            patch("app.core.services.alert_service.sincronizar_alertas"),
            patch.dict("sys.modules", {"streamlit": fake_st}),
        ]
        for p in self.patches:
            p.start()
        import streamlit as st
        st.session_state = fake_st.session_state

        bebida_service.servicio = crear_servicio(
            "bebidas", "bebidas",
            [CategoriaReceta.BEBIDAS],
            solo_bebidas_sueltas=True,
            titulo_documento="Registro de Bebidas",
            export_tipo="registro_bebidas",
        )
        for nombre in (
            "anadir_a_cesta", "anadir_receta_a_cesta", "registrar",
            "historial_ordenado", "fecha_mas_antigua", "registros_exportables",
            "configuracion_exportacion", "limpiar_cesta", "productos_catalogo",
        ):
            setattr(bebida_service, nombre, getattr(bebida_service.servicio, nombre))

    def tearDown(self) -> None:
        for p in self.patches:
            p.stop()

    def test_acepta_solo_receta_bebidas(self) -> None:
        self.assertTrue(bebida_service.anadir_receta_a_cesta("r_beb", 1.0).ok)
        self.assertFalse(bebida_service.anadir_receta_a_cesta("r_des", 1.0).ok)
        self.assertFalse(bebida_service.anadir_receta_a_cesta("r_com", 1.0).ok)
        self.assertFalse(bebida_service.anadir_receta_a_cesta("r_cen", 1.0).ok)

    def test_rechaza_producto_no_bebida(self) -> None:
        resultado = bebida_service.anadir_a_cesta("p01", 0.5)
        self.assertFalse(resultado.ok)
        self.assertIn("bebida", resultado.mensaje.lower())

    def test_acepta_producto_bebida(self) -> None:
        self.assertTrue(bebida_service.anadir_a_cesta("p03", 0.3).ok)

    def test_catalogo_solo_bebidas(self) -> None:
        ids = {p["id"] for p in bebida_service.productos_catalogo()}
        self.assertEqual(ids, {"p02", "p03"})

    def test_registrar_con_origen_y_tipo_servicio(self) -> None:
        self.assertTrue(bebida_service.anadir_receta_a_cesta("r_beb", 1.0).ok)
        self.assertTrue(bebida_service.anadir_a_cesta("p03", 0.25).ok)
        resultado = bebida_service.registrar(date(2026, 7, 21))
        self.assertTrue(resultado.ok, resultado.mensaje)

        reg = self.data.registros_servicio[0]
        self.assertEqual(reg.tipo_servicio, "bebidas")

        cafe = next(l for l in self.data.lotes if l.id == "l02")
        zumo = next(l for l in self.data.lotes if l.id == "l03")
        self.assertAlmostEqual(cafe.cantidad_restante, 1.8, places=4)
        self.assertAlmostEqual(zumo.cantidad_restante, 1.75, places=4)

        origenes = {d.origen for d in reg.lineas_detalle}
        self.assertIn(OrigenConsumo.INGREDIENTE_RECETA.value, origenes)
        self.assertIn(OrigenConsumo.PRODUCTO_DIRECTO.value, origenes)
        self.assertTrue(all(d.tipo_servicio == "bebidas" for d in reg.lineas_detalle))
        self.assertTrue(any("bebida" in a.accion.lower() for a in self.data.actividades))

    def test_exportacion_titulo_bebidas(self) -> None:
        self.assertTrue(bebida_service.anadir_a_cesta("p02", 0.1).ok)
        self.assertTrue(bebida_service.registrar(date(2026, 7, 22)).ok)

        config = bebida_service.configuracion_exportacion()
        self.assertEqual(config.tipo, "registro_bebidas")
        self.assertEqual(config.titulo_documento, "Registro de Bebidas")

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            resultado = exportar_periodo(
                config,
                date(2026, 7, 20),
                datetime(2026, 7, 26, 23, 59, 59),
                automatica=True,
                carpeta_exports=carpeta,
                archivo_meta=carpeta / "_meta.json",
            )
            self.assertTrue(resultado.ok, resultado.mensaje)
            self.assertTrue(resultado.nombre_archivo.startswith("Registro_de_Bebidas_"))
            from openpyxl import load_workbook
            libro = load_workbook(resultado.ruta)
            self.assertIn(nombre_hoja_dia(date(2026, 7, 22)), libro.sheetnames)


if __name__ == "__main__":
    unittest.main()
