"""Pruebas de la Fase 2: construcción de `RegistroExportable` para desayuno y
merma, y su integración con el motor central de exportación semanal.

Ejecutar desde la raíz del proyecto con:

    python -m unittest discover -s tests -v

No dependen de una sesión de Streamlit real: se inyecta un `AppData`
sintético parcheando `get_data` en cada servicio.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime, time
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.models import (
    AppData,
    ExtraRecetaDesayuno,
    LineaDesayuno,
    LineaMerma,
    MotivoMerma,
    OmisionRecetaDesayuno,
    Producto,
    RegistroDesayuno,
    RegistroMerma,
    RegistroRecetaDesayuno,
    UnidadProducto,
)
from app.core.services import desayuno_service, merma_service
from app.core.services.exportacion_semanal_service import exportar_periodo


def _productos() -> list[Producto]:
    return [
        Producto("p01", "Café", UnidadProducto.L),
        Producto("p02", "Leche", UnidadProducto.L),
        Producto("p03", "Pan", UnidadProducto.UD),
        Producto("p04", "Mantequilla", UnidadProducto.KG),
    ]


class TestRegistrosExportablesDesayuno(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData(productos=_productos())
        self._patcher = patch("app.core.services.desayuno_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()

    def test_desayuno_largo_con_receta_extra_omision_y_suelto(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno(
                "d01", date(2026, 7, 21),
                lineas=[
                    LineaDesayuno("p01", 2.0, 1.5, False),
                    LineaDesayuno("p02", 0.5, 0.4, True),
                    LineaDesayuno("p03", 10.0, 3.0, False),
                ],
                coste_total=4.9,
                registrado_por="Ana",
                num_huespedes=25,
                registros_recetas=[
                    RegistroRecetaDesayuno(
                        "r01", "Tostada con mantequilla", 10.0,
                        extras=[ExtraRecetaDesayuno("p02", 0.5)],
                        omisiones=[OmisionRecetaDesayuno("p04")],
                    ),
                ],
                hora=time(8, 15),
            ),
        ]

        registros = desayuno_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 1)
        registro = registros[0]

        self.assertEqual(registro.fecha, date(2026, 7, 21))
        self.assertEqual(registro.hora, time(8, 15))
        self.assertEqual(registro.tipo, "Desayuno")
        self.assertEqual(registro.identificador, "d01")
        self.assertEqual(registro.usuario, "Ana")
        self.assertEqual(registro.columnas, ["Tipo", "Producto / Receta", "Detalle", "Cantidad", "Unidad", "Coste"])

        tipos = [fila[0] for fila in registro.filas]
        self.assertIn("Receta", tipos)
        self.assertIn("Extra/Omisión", tipos)
        self.assertIn("Omisión", tipos)
        self.assertIn("Producto", tipos)

        nombres = [fila[1] for fila in registro.filas]
        self.assertIn("Tostada con mantequilla", nombres)
        self.assertIn("Leche", nombres)
        self.assertIn("Mantequilla", nombres)
        self.assertIn("Café", nombres)
        self.assertIn("Pan", nombres)

        resumen = dict(registro.resumen)
        self.assertEqual(resumen["Huéspedes"], "25")

    def test_varios_registros_de_desayuno_en_un_dia_no_se_fusionan(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 21), lineas=[LineaDesayuno("p01", 1.0, 1.0)],
                              registrado_por="Ana", hora=time(8, 0)),
            RegistroDesayuno("d02", date(2026, 7, 21), lineas=[LineaDesayuno("p02", 1.0, 1.0)],
                              registrado_por="Luis", hora=time(9, 30)),
        ]
        registros = desayuno_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 2)
        self.assertEqual({r.identificador for r in registros}, {"d01", "d02"})

    def test_dias_sin_desayunos_no_generan_registros(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 21), lineas=[LineaDesayuno("p01", 1.0, 1.0)]),
        ]
        registros = desayuno_service.registros_exportables(date(2026, 7, 27), datetime(2026, 8, 2, 23, 59, 59))
        self.assertEqual(registros, [])

    def test_exportacion_parcial_hasta_el_momento_actual(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 20), lineas=[LineaDesayuno("p01", 1.0, 1.0)]),
            RegistroDesayuno("d02", date(2026, 7, 23), lineas=[LineaDesayuno("p01", 1.0, 1.0)]),
            RegistroDesayuno("d03", date(2026, 7, 25), lineas=[LineaDesayuno("p01", 1.0, 1.0)]),
        ]
        # "Ahora" es miércoles 22/07 a mediodía: solo debe incluir lo hasta esa fecha.
        registros = desayuno_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 22, 12, 0, 0))
        self.assertEqual({r.identificador for r in registros}, {"d01"})

    def test_fecha_mas_antigua(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 6), lineas=[LineaDesayuno("p01", 1.0, 1.0)]),
            RegistroDesayuno("d02", date(2026, 7, 21), lineas=[LineaDesayuno("p01", 1.0, 1.0)]),
        ]
        self.assertEqual(desayuno_service.fecha_mas_antigua(), date(2026, 7, 6))

    def test_fecha_mas_antigua_sin_datos_es_none(self) -> None:
        self.assertIsNone(desayuno_service.fecha_mas_antigua())


class TestRegistrosExportablesMerma(unittest.TestCase):
    def setUp(self) -> None:
        self.data = AppData(productos=_productos())
        self._patcher = patch("app.core.services.merma_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()

    def test_varias_mermas_del_mismo_producto_no_se_fusionan(self) -> None:
        self.data.mermas = [
            RegistroMerma(
                "m01", date(2026, 7, 21),
                lineas=[
                    LineaMerma("p01", 1.0, 0.8, MotivoMerma.EXPIRACION, lote_id="lote_a"),
                ],
                registrado_por="Ana", hora=time(10, 0),
            ),
            RegistroMerma(
                "m02", date(2026, 7, 21),
                lineas=[
                    LineaMerma("p01", 2.0, 1.6, MotivoMerma.MERMA, lote_id="lote_b"),
                ],
                registrado_por="Luis", hora=time(16, 0),
            ),
        ]
        registros = merma_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 2)

        filas_totales = [fila for r in registros for fila in r.filas]
        lotes = [fila[1] for fila in filas_totales]
        self.assertEqual(sorted(lotes), ["lote_a", "lote_b"])

    def test_columnas_y_desglose_de_una_merma(self) -> None:
        self.data.mermas = [
            RegistroMerma(
                "m01", date(2026, 7, 21),
                lineas=[
                    LineaMerma("p03", 5.0, 1.5, MotivoMerma.PRODUCTO_MALO, "Se cayó", "lote_x"),
                ],
                coste_total=1.5,
                registrado_por="Ana",
                hora=time(11, 30),
            ),
        ]
        registros = merma_service.registros_exportables(date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59))
        self.assertEqual(len(registros), 1)
        registro = registros[0]
        self.assertEqual(registro.tipo, "Merma")
        self.assertEqual(
            registro.columnas,
            ["Producto", "Lote", "Cantidad", "Unidad", "Motivo", "Servicio", "Coste", "Comentario"],
        )
        fila = registro.filas[0]
        self.assertEqual(
            fila,
            ["Pan", "lote_x", 5.0, "Ud", "Producto malo", "Sin desglose histórico", 1.5, "Se cayó"],
        )

    def test_dias_sin_mermas_no_generan_registros(self) -> None:
        self.data.mermas = [RegistroMerma("m01", date(2026, 7, 21))]
        registros = merma_service.registros_exportables(date(2026, 7, 27), datetime(2026, 8, 2, 23, 59, 59))
        self.assertEqual(registros, [])


class TestIntegracionExportacionSemanalCompleta(unittest.TestCase):
    """De extremo a extremo: registros de dominio -> Excel real generado por
    el motor central (Fase 1), usando los builders de Fase 2."""

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.carpeta = Path(self._tmp.name)
        self.archivo_meta = self.carpeta / "_meta.json"
        self.data = AppData(productos=_productos())
        self._patcher = patch("app.core.services.desayuno_service.get_data", return_value=self.data)
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()
        self._tmp.cleanup()

    def test_exportacion_semanal_automatica_genera_una_hoja_por_dia_con_registros(self) -> None:
        self.data.desayunos = [
            RegistroDesayuno("d01", date(2026, 7, 20), lineas=[LineaDesayuno("p01", 1.0, 1.0)], hora=time(8, 0)),
            RegistroDesayuno("d02", date(2026, 7, 20), lineas=[LineaDesayuno("p02", 1.0, 1.0)], hora=time(9, 0)),
            RegistroDesayuno("d03", date(2026, 7, 23), lineas=[LineaDesayuno("p03", 3.0, 2.0)], hora=time(8, 30)),
        ]
        config = desayuno_service.configuracion_exportacion()
        resultado = exportar_periodo(
            config, date(2026, 7, 20), datetime(2026, 7, 26, 23, 59, 59),
            automatica=True, fecha_exportacion=date(2026, 7, 27),
            carpeta_exports=self.carpeta, archivo_meta=self.archivo_meta,
        )
        self.assertTrue(resultado.ok)
        self.assertEqual(
            resultado.nombre_archivo,
            "Registro_de_Desayuno_2026-07-27_2026-07-20_a_2026-07-26.xlsx",
        )

        libro = load_workbook(resultado.ruta)
        self.assertEqual(set(libro.sheetnames), {"Info", "Lun 20-07", "Jue 23-07"})

        hoja_lunes = libro["Lun 20-07"]
        valores = " | ".join(
            str(c.value) for row in hoja_lunes.iter_rows() for c in row if c.value is not None
        )
        self.assertIn("d01", valores)
        self.assertIn("d02", valores)


if __name__ == "__main__":
    unittest.main()
