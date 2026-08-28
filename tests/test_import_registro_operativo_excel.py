"""Tests import_registro_operativo_excel (dry-run, fixture mínimo)."""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from app.core.models import AppData, Producto, UnidadProducto
from app.core.models.buffet import LineaConfigBuffet, TIPO_LINEA_SIMPLE
from tests.demo_isolation import EXPORT_SESSION_MODULES, isolated_persist


def _fixture_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"
    ws.append(["Fecha", "Huespedes", "Tipo", "Nombre", "Cantidad ↑↓", "Importado"])
    ws.append([date(2026, 8, 10), 1, "Extra", "Kiwi", 1, ""])

    wb.create_sheet("RegistroBebidasDesayuno")
    wb_d = wb["RegistroBebidasDesayuno"]
    wb_d.append(["Fecha", "Tipo", "Nombre", "Cantidad ↑↓", "Notas", "Importado"])

    wc = wb.create_sheet("RegistroComida")
    wc.append(["Fecha", "Tipo", "Nombre", "Cantidad ↑↓", "Notas", "Importado"])

    wn = wb.create_sheet("RegistroCena")
    wn.append(["Fecha", "Tipo", "Nombre", "Cantidad ↑↓", "Notas", "Importado"])

    wb.create_sheet("ConfigBuffet")
    cfg = wb["ConfigBuffet"]
    cfg.append(["Seccion", "Orden", "Concepto", "ProductoId", "Unidad", "CantDefecto", "Tipo", "ProductoBote", "RecetaId", "Activo"])
    cfg.append(["Frutas", 1, "Kiwi", "p71", "gr", 0.08, "simple", "", "", "Si"])

    wb.create_sheet("ConsumoBuffet")
    buf = wb["ConsumoBuffet"]
    buf.append(["Fecha", "Seccion", "Concepto", "Cantidad", "Motivo", "ZumoBote", "Coste", "Notas", "Importado"])
    buf.append([date(2026, 8, 10), "Frutas", "Kiwi", 1, "Consumo", "", "", "", "", ""])

    wb.save(path)
    wb.close()


class TestImportRegistroOperativoExcel(unittest.TestCase):
    def setUp(self) -> None:
        self._session: dict = {}
        self._st = mock.patch("streamlit.session_state", self._session)
        self._st.start()
        from tests.streamlit_store_harness import cleanup_container, use_patched_streamlit_stores

        use_patched_streamlit_stores()
        self.addCleanup(cleanup_container)

        self.data = AppData(
            productos=[Producto("p71", "Kiwi", UnidadProducto.KG)],
            config_buffet=[
                LineaConfigBuffet("cb1", "Frutas", 1, "Kiwi", "p71", "gr", 0.08, tipo_linea=TIPO_LINEA_SIMPLE),
            ],
        )
        self._iso = isolated_persist(*EXPORT_SESSION_MODULES, data=self.data)
        self._iso.__enter__()
        self.addCleanup(self._iso.__exit__, None, None, None)

        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.xlsx = Path(self.tmp.name) / "test.xlsx"
        self.json = Path(self.tmp.name) / "datos.json"
        _fixture_xlsx(self.xlsx)
        from app.data.serializers import save_json, appdata_to_dict

        save_json(self.json, appdata_to_dict(self.data))

    def tearDown(self) -> None:
        self._st.stop()

    def test_dry_run_multihoja(self) -> None:
        import scripts.import_registro_operativo_excel as imp

        with mock.patch.object(
            sys, "argv",
            ["import", str(self.xlsx), "--path", str(self.json), "--dry-run"],
        ):
            code = imp.main()
        self.assertEqual(code, 0)


if __name__ == "__main__":
    unittest.main()
