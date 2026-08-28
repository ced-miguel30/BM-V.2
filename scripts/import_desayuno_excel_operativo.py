"""Importa docs/plantillas/registro_desayuno_operativo_LISTA_ACTUALIZADA_ACTUALIZADA.xlsx a BM-DATOS.

Agrupa por Fecha → cesta desayuno → registrar_desayuno (idempotente).

Uso:
  .\\.venv\\Scripts\\python.exe scripts\\import_desayuno_excel_operativo.py RUTA.xlsx
  .\\.venv\\Scripts\\python.exe scripts\\import_desayuno_excel_operativo.py RUTA.xlsx --dry-run
  .\\.venv\\Scripts\\python.exe scripts\\import_desayuno_excel_operativo.py RUTA.xlsx --path RUTA\\datos_hotel.json
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.bootstrap import configure_for_flet, get_container, reset_container
from app.core.auth.roles import ROL_DIRECCION
from app.core.auth.session import ACTOR_TYPE_USUARIO, AuthSession, save_auth_session
from app.core.services import desayuno_service as des
from app.core.services.pack_unidades import piezas_a_ud_paquete, ud_paquete_a_piezas
from app.core.services.receta_service import (
    ETIQUETA_TOSTADA_DEL_DIA,
    receta_tostada_del_dia,
)
from app.core.services.text_search import normalizar_texto

HOTEL_DEFAULT = Path(os.environ["LOCALAPPDATA"]) / "BM-V2-local" / "data" / "datos_hotel.json"
CLAVE_PREFIX = "desayuno-xlsx"
CLAVE_VER = "v3"

# Labels de Extra/Omitir que son «tipo de huevo» (sustituyen el frito de Desayuno inglés).
_TIPOS_HUEVO = frozenset(
    {
        "huevo frito",
        "huevo pochado",
        "huevo cocido",
        "huevos revueltos",
        "huevo revuelto",
        "huevo",  # omitir el huevo de la ficha
        "huevo cascara",
        "sin huevo",
    }
)
_HUEVO_FRITO_PID = "p48"

# Pan: blanco (p09) ↔ integral (p11) ↔ sin gluten (p04). Sustituye el de la ficha.
_PAN_BLANCO = "p09"
_PAN_INTEGRAL = "p11"
_PAN_SIN_GLUTEN = "p04"
_PANS_SUSTITUIBLES = frozenset({_PAN_BLANCO, _PAN_INTEGRAL, _PAN_SIN_GLUTEN})
_TIPOS_PAN = frozenset(
    {
        "tostada",
        "tostada integral",
        "tostada sin gluten",
        "pan blanco",
        "pan integral",
        "pan sin gluten",
        "molde comun",
        "molde integral",
        "sin tostada",
        "sin pan",
    }
)


def _norm(s: str) -> str:
    return normalizar_texto(s or "")


def _es_tipo_huevo(label: str) -> bool:
    return _norm(label) in _TIPOS_HUEVO


def _es_tipo_pan(label: str) -> bool:
    return _norm(label) in _TIPOS_PAN


def _es_desayuno_ingles(nombre: str) -> bool:
    return _norm(nombre) == "desayuno ingles"


def _pan_destino(label: str) -> str | None:
    """Producto de pan elegido, o None si es solo omitir."""
    key = _norm(label)
    if key in ("tostada integral", "pan integral", "molde integral"):
        return _PAN_INTEGRAL
    if key in ("tostada sin gluten", "pan sin gluten"):
        return _PAN_SIN_GLUTEN
    if key in ("tostada", "pan blanco", "molde comun"):
        return _PAN_BLANCO
    return None


def _qty_pan_en_receta(receta) -> tuple[str | None, float]:
    """Devuelve (producto_id_pan, cantidad nativa) del pan en la ficha."""
    por_pid: dict[str, float] = {}
    for ing in receta.ingredientes or []:
        if ing.producto_id in _PANS_SUSTITUIBLES:
            por_pid[ing.producto_id] = por_pid.get(ing.producto_id, 0.0) + float(
                ing.cantidad or 0
            )
    if not por_pid:
        return None, 0.0
    pid = max(por_pid, key=por_pid.get)
    return pid, por_pid[pid]


def _qty_pan_sustitucion(
    from_pid: str, from_qty_nativa: float, to_pid: str, porciones: float
) -> float:
    """Misma nº de piezas/rebanadas, convertida a Ud nativa del destino."""
    piezas = ud_paquete_a_piezas(from_pid, from_qty_nativa)
    if piezas <= 0:
        piezas = 1.0
    return round(piezas_a_ud_paquete(to_pid, piezas) * float(porciones), 6)


def _omitir_pan_si_hay(pid: str) -> None:
    r = des.anadir_mod_a_receta_en_cesta(pid, -1e9)
    if not r.ok and "no está" not in (r.mensaje or "").lower():
        raise ValueError(r.mensaje)


def _omitir_todos_los_panes() -> None:
    for pid in _PANS_SUSTITUIBLES:
        _omitir_pan_si_hay(pid)


def _auth() -> None:
    save_auth_session(
        AuthSession(
            authenticated=True,
            actor_type=ACTOR_TYPE_USUARIO,
            actor_id="import-xlsx-des",
            actor_label="Import Excel desayuno",
            role=ROL_DIRECCION,
            session_id="import-xlsx-des-session",
            login_at=datetime.now(timezone.utc).isoformat(),
            terminal_id=None,
            login="import",
        )
    )


def _boot(path: Path) -> None:
    reset_container()
    configure_for_flet(data_path=str(path))
    _auth()


def _parse_fecha(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val, default: float | None = None) -> float | None:
    if val is None or val == "":
        return default
    try:
        return float(str(val).replace(",", "."))
    except ValueError:
        return default


@dataclass
class LineaExcel:
    row: int
    fecha: date
    huespedes: int | None
    tipo: str
    nombre: str
    cantidad: float
    extras: list[tuple[str, float | None]]  # hasta 4 (label, cant)
    omitir1: str
    omitir2: str
    notas: str


@dataclass
class DiaPlan:
    fecha: date
    huespedes: int | None = None
    lineas: list[LineaExcel] = field(default_factory=list)
    row_indices: list[int] = field(default_factory=list)


def _header_map(ws) -> dict[str, int]:
    raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[str, int] = {}
    for i, h in enumerate(raw):
        if h is None:
            continue
        key = str(h).strip()
        # Normalizar alias de cabecera Cantidad / CantN
        if key.startswith("Cantidad"):
            key = "Cantidad ↑↓"
        elif key.startswith("Cant") and len(key) >= 5 and key[4].isdigit():
            n = key[4]
            key = f"Cant{n} ↑↓"
        out[key] = i
    missing = [h for h in ("Fecha", "Tipo", "Nombre", "Cantidad ↑↓") if h not in out]
    if missing:
        raise SystemExit(f"Faltan columnas en Registro: {missing}. Cabeceras: {raw}")
    return out


def _leer_registro(path: Path) -> list[LineaExcel]:
    wb = load_workbook(path, data_only=True)
    if "Registro" not in wb.sheetnames:
        raise SystemExit("El Excel debe tener hoja «Registro».")
    ws = wb["Registro"]
    cols = _header_map(ws)
    lineas: list[LineaExcel] = []
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def cell(name: str, default=None):
            idx = cols.get(name)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        fecha = _parse_fecha(cell("Fecha"))
        tipo = str(cell("Tipo") or "").strip()
        nombre = str(cell("Nombre") or "").strip()
        if not fecha and not tipo and not nombre:
            continue
        if not fecha:
            raise SystemExit(f"Fila {r_i}: Fecha obligatoria.")
        if not tipo or not nombre:
            raise SystemExit(f"Fila {r_i}: Tipo y Nombre obligatorios.")
        cant = _parse_float(cell("Cantidad ↑↓"), 1.0)
        if cant is None or cant <= 0:
            raise SystemExit(f"Fila {r_i}: Cantidad debe ser > 0.")
        # Huéspedes: 1 = nuevo comensal; 0 o vacío = mismo comensal / no suma.
        hues_raw = cell("Huespedes")
        if hues_raw is None or hues_raw == "":
            hues_val: int | None = None
        else:
            hues_f = _parse_float(hues_raw, None)
            hues_val = int(hues_f) if hues_f is not None else None
        extras: list[tuple[str, float | None]] = []
        for n in (1, 2, 3, 4):
            lab = str(cell(f"Extra{n}") or "").strip()
            if not lab:
                continue
            extras.append((lab, _parse_float(cell(f"Cant{n} ↑↓"), None)))
        lineas.append(
            LineaExcel(
                row=r_i,
                fecha=fecha,
                huespedes=hues_val,
                tipo=tipo,
                nombre=nombre,
                cantidad=float(cant),
                extras=extras,
                omitir1=str(cell("Omitir1") or "").strip(),
                omitir2=str(cell("Omitir2") or "").strip(),
                notas=str(cell("Notas") or "").strip(),
            )
        )
    wb.close()
    return lineas


def _agrupar(lineas: list[LineaExcel]) -> list[DiaPlan]:
    """Misma Fecha = un solo desayuno. Huéspedes = suma de 1s (0 no suma)."""
    por: dict[date, DiaPlan] = {}
    for ln in lineas:
        dia = por.get(ln.fecha)
        if dia is None:
            dia = DiaPlan(fecha=ln.fecha)
            por[ln.fecha] = dia
        dia.lineas.append(ln)
        dia.row_indices.append(ln.row)
    for dia in por.values():
        suma = 0
        hay_marca = False
        for ln in dia.lineas:
            if ln.huespedes is None:
                continue
            hay_marca = True
            if ln.huespedes > 0:
                suma += int(ln.huespedes)
        if hay_marca and suma >= 1:
            dia.huespedes = suma
        else:
            # Sin marcas 1/0: fallback = nº de líneas Receta
            n_rec = sum(1 for ln in dia.lineas if _norm(ln.tipo) == "receta")
            dia.huespedes = max(n_rec, 1)
    return [por[k] for k in sorted(por)]


def _mapa_recetas(data) -> dict[str, object]:
    m: dict[str, object] = {}
    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        m[_norm(r.nombre)] = r
    return m


def _mapa_extras_label() -> dict[str, dict]:
    m: dict[str, dict] = {}
    for e in des.extras_rapidos_desayuno() + des.leches_rapidas_desayuno():
        m[_norm(e["label"])] = e
        m[_norm(e["nombre"])] = e
    return m


def _mapa_productos(data) -> dict[str, object]:
    m: dict[str, object] = {}
    for p in data.productos:
        if not getattr(p, "activo", True):
            continue
        m[_norm(p.nombre)] = p
        codigo = getattr(p, "codigo", None) or ""
        if codigo:
            m[_norm(codigo)] = p
        m[_norm(p.id)] = p
    return m


def _resolver_receta(data, nombre: str, fecha: date, rec_map: dict):
    if _norm(nombre) == _norm(ETIQUETA_TOSTADA_DEL_DIA):
        return receta_tostada_del_dia(fecha)
    return rec_map.get(_norm(nombre))


def _qty_extra(extra_map: dict, label: str, cant_filas: float | None) -> tuple[str, float]:
    e = extra_map.get(_norm(label))
    if e is None:
        raise ValueError(f"Extra no encontrado: «{label}»")
    mult = float(cant_filas) if cant_filas and cant_filas > 0 else 1.0
    return e["producto_id"], round(float(e["cantidad"]) * mult, 6)


def _resolver_omit_pid(data, receta, label: str, prod_map: dict, extra_map: dict) -> str:
    key = _norm(label)
    if key in ("sin huevo", "huevo", "huevo cascara"):
        return _HUEVO_FRITO_PID
    if key in ("sin tostada", "sin pan"):
        # Preferir el pan que lleve la ficha
        pid_pan, _ = _qty_pan_en_receta(receta)
        return pid_pan or _PAN_BLANCO
    # 1) label de extra rápido
    e = extra_map.get(key)
    if e:
        return e["producto_id"]
    # 2) nombre producto
    p = prod_map.get(key)
    if p:
        return p.id
    # 3) match contra ingredientes de la receta
    for ing in receta.ingredientes or []:
        prod = next((x for x in data.productos if x.id == ing.producto_id), None)
        if prod and _norm(prod.nombre) == key:
            return prod.id
        if _norm(ing.producto_id) == key:
            return ing.producto_id
    raise ValueError(f"Omitir: no se resolvió «{label}»")


def _clave(fecha: date, batch: int | None = None) -> str:
    base = f"{CLAVE_PREFIX}-{fecha.isoformat()}-{CLAVE_VER}"
    if batch is None:
        return base
    return f"{base}-b{batch}"


def _ya_importado(data, fecha: date) -> object | None:
    prefix = f"{CLAVE_PREFIX}-{fecha.isoformat()}-{CLAVE_VER}"
    return next(
        (
            d
            for d in data.desayunos
            if str(getattr(d, "clave_idempotencia", None) or "").startswith(prefix)
            and not getattr(d, "anulado", False)
        ),
        None,
    )


def _anadir_linea_a_cesta(
    ln: LineaExcel,
    *,
    fecha: date,
    rec_map: dict,
    extra_map: dict,
    prod_map: dict,
) -> None:
    data = get_container().app_data_store.get()
    tipo = _norm(ln.tipo)
    if tipo == "receta":
        rec = _resolver_receta(data, ln.nombre, fecha, rec_map)
        if rec is None:
            raise ValueError(f"Receta no encontrada: «{ln.nombre}»")
        ingles = _es_desayuno_ingles(ln.nombre)
        egg_extras: list[tuple[str, float | None]] = []
        pan_destino: str | None = None
        pan_solo_omitir = False
        other_extras: list[tuple[str, float | None]] = []
        pid_pan_ficha, qty_pan_ficha = _qty_pan_en_receta(rec)

        for lab, canti in ln.extras:
            if ingles and _es_tipo_huevo(lab) and _norm(lab) not in (
                "huevo",
                "huevo cascara",
                "sin huevo",
            ):
                egg_extras.append((lab, canti))
            elif _es_tipo_pan(lab):
                dest = _pan_destino(lab)
                if dest is None:
                    pan_solo_omitir = True
                elif pid_pan_ficha:
                    pan_destino = dest
                else:
                    other_extras.append((lab, canti))
            else:
                other_extras.append((lab, canti))

        omit_labels = [x for x in (ln.omitir1, ln.omitir2) if x]
        for lab in omit_labels:
            if _es_tipo_pan(lab) and _pan_destino(lab) is None:
                pan_solo_omitir = True

        for lab, canti in other_extras:
            pid, qty = _qty_extra(extra_map, lab, canti)
            r = des.anadir_mod_pendiente_receta(pid, qty)
            if not r.ok:
                raise ValueError(r.mensaje)
        for lab, canti in egg_extras:
            pid, qty = _qty_extra(extra_map, lab, canti)
            r = des.anadir_mod_pendiente_receta(pid, qty)
            if not r.ok:
                raise ValueError(r.mensaje)

        r = des.anadir_receta_a_cesta(rec.id, float(ln.cantidad))
        if not r.ok:
            raise ValueError(r.mensaje)

        for lab in omit_labels:
            if _es_tipo_pan(lab):
                continue
            pid = _resolver_omit_pid(data, rec, lab, prod_map, extra_map)
            r = des.anadir_mod_a_receta_en_cesta(pid, -1e9)
            if not r.ok and "no está" not in (r.mensaje or "").lower():
                raise ValueError(r.mensaje)

        if ingles and (
            egg_extras
            or any(_es_tipo_huevo(x) for x in omit_labels)
        ):
            if any(
                i.producto_id == _HUEVO_FRITO_PID
                for i in (rec.ingredientes or [])
            ):
                r = des.anadir_mod_a_receta_en_cesta(_HUEVO_FRITO_PID, -1e9)
                if (
                    not r.ok
                    and "no está" not in (r.mensaje or "").lower()
                ):
                    raise ValueError(r.mensaje)

        if pan_solo_omitir and not pan_destino:
            _omitir_todos_los_panes()
        elif pan_destino and pid_pan_ficha and pan_destino != pid_pan_ficha:
            qty_nativa = _qty_pan_sustitucion(
                pid_pan_ficha, qty_pan_ficha, pan_destino, float(ln.cantidad)
            )
            _omitir_pan_si_hay(pid_pan_ficha)
            r = des.anadir_mod_a_receta_en_cesta(pan_destino, qty_nativa)
            if not r.ok:
                raise ValueError(r.mensaje)
        for lab in omit_labels:
            if not _es_tipo_pan(lab):
                continue
            dest = _pan_destino(lab)
            if dest is None:
                continue
            _omitir_pan_si_hay(dest)
    elif tipo in ("extra", "producto"):
        if _norm(ln.nombre) in extra_map:
            pid, qty = _qty_extra(extra_map, ln.nombre, ln.cantidad)
        else:
            p = prod_map.get(_norm(ln.nombre))
            if p is None:
                raise ValueError(f"Producto/extra no encontrado: «{ln.nombre}»")
            pid = p.id
            qty = float(ln.cantidad)
        r = des.anadir_a_cesta(pid, qty)
        if not r.ok:
            raise ValueError(r.mensaje)
    else:
        raise ValueError(f"Tipo desconocido: «{ln.tipo}» (use Receta/Extra/Producto)")


def _importar_dia(
    dia: DiaPlan,
    *,
    dry_run: bool,
    rec_map: dict,
    extra_map: dict,
    prod_map: dict,
) -> dict:
    data = get_container().app_data_store.get()
    existente = _ya_importado(data, dia.fecha)
    if existente is not None:
        return {
            "ok": True,
            "skipped": True,
            "mensaje": f"Ya importado ({existente.id})",
            "status": f"SKIP {existente.id}",
        }

    huespedes = int(dia.huespedes or 1)

    if dry_run:
        preview = []
        errs: list[str] = []
        for ln in dia.lineas:
            tipo = _norm(ln.tipo)
            try:
                if tipo == "receta":
                    rec = _resolver_receta(data, ln.nombre, dia.fecha, rec_map)
                    if rec is None:
                        raise ValueError(f"Receta no encontrada: «{ln.nombre}»")
                    for lab, canti in ln.extras:
                        _qty_extra(extra_map, lab, canti)
                    for lab in (ln.omitir1, ln.omitir2):
                        if lab:
                            _resolver_omit_pid(data, rec, lab, prod_map, extra_map)
                elif tipo in ("extra", "producto"):
                    if _norm(ln.nombre) in extra_map:
                        _qty_extra(extra_map, ln.nombre, ln.cantidad)
                    elif prod_map.get(_norm(ln.nombre)) is None:
                        raise ValueError(f"Producto/extra no encontrado: «{ln.nombre}»")
                else:
                    raise ValueError(f"Tipo desconocido: «{ln.tipo}»")
                preview.append(f"{ln.tipo}:{ln.nombre}x{ln.cantidad:g}")
            except ValueError as exc:
                errs.append(f"fila {ln.row}: {exc}")
        if errs:
            return {
                "ok": False,
                "dry_run": True,
                "mensaje": " | ".join(errs),
                "status": "ERROR",
            }
        return {
            "ok": True,
            "dry_run": True,
            "mensaje": (
                f"dry-run {dia.fecha} huespedes={huespedes} "
                f"lineas={len(dia.lineas)} -> " + "; ".join(preview)
            ),
            "status": "DRY-RUN OK",
        }

    # Un día = un solo registro de desayuno (toda la cesta junta).
    des.limpiar_cesta()
    errores: list[str] = []
    for ln in dia.lineas:
        try:
            _anadir_linea_a_cesta(
                ln,
                fecha=dia.fecha,
                rec_map=rec_map,
                extra_map=extra_map,
                prod_map=prod_map,
            )
        except ValueError as exc:
            errores.append(f"fila {ln.row}: {exc}")
    if errores:
        des.limpiar_cesta()
        return {
            "ok": False,
            "mensaje": " | ".join(errores),
            "status": "ERROR",
        }
    if des.cesta_vacia():
        return {"ok": False, "mensaje": "Cesta vacía", "status": "ERROR"}

    notas = " | ".join(ln.notas for ln in dia.lineas if ln.notas)[:200]
    res = des.registrar_desayuno(
        dia.fecha,
        huespedes,
        clave_idempotencia=_clave(dia.fecha),
        observaciones=f"Import Excel desayuno operativo. {notas}".strip(),
    )
    if not res.ok:
        des.limpiar_cesta()
        return {
            "ok": False,
            "mensaje": res.mensaje,
            "status": f"ERROR {res.mensaje}",
            "codigo": getattr(res, "codigo", None),
        }
    data = get_container().app_data_store.get()
    reg = _ya_importado(data, dia.fecha)
    rid = getattr(reg, "id", "?") if reg is not None else "?"
    return {
        "ok": True,
        "mensaje": f"OK {rid} (huespedes={huespedes}, lineas={len(dia.lineas)})",
        "status": f"OK {rid}",
    }


def _escribir_estados(xlsx: Path, por_fila: dict[int, str]) -> None:
    wb = load_workbook(xlsx)
    ws = wb["Registro"]
    cols = _header_map(ws)
    idx = cols.get("Importado")
    if idx is None:
        # añadir cabecera
        idx = 12
        ws.cell(1, idx + 1, "Importado")
    for row, status in por_fila.items():
        ws.cell(row, idx + 1, status)
    wb.save(xlsx)
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, nargs="?", default=None, help="Plantilla / archivo rellenado")
    parser.add_argument("--path", type=Path, default=HOTEL_DEFAULT)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--check",
        action="store_true",
        help="Solo verifica conexion a BM (ruta datos + catalogo), sin importar",
    )
    args = parser.parse_args()

    if not args.path.exists():
        print("No existe datos:", args.path)
        return 1

    if args.check or args.xlsx is None:
        print("=== Conexion BM ===")
        print("Datos:", args.path.resolve())
        _boot(args.path)
        data = get_container().app_data_store.get()
        hotel = ""
        if getattr(data, "configuracion", None):
            hotel = getattr(data.configuracion, "nombre_establecimiento", "") or ""
        n_rec = sum(
            1
            for r in data.recetas
            if getattr(r, "activo", True)
            and (
                (hasattr(r.categoria, "value") and r.categoria.value == "desayuno")
                or str(r.categoria) == "desayuno"
            )
        )
        n_ext = len(des.extras_rapidos_desayuno()) + len(des.leches_rapidas_desayuno())
        print("Hotel:", hotel or "(sin nombre)")
        print(f"Recetas desayuno activas: {n_rec}")
        print(f"Extras rapidos: {n_ext}")
        print("Estado: CONECTADO")
        print("===================")
        if args.check or args.xlsx is None:
            return 0

    if not args.xlsx.exists():
        print("No existe", args.xlsx)
        return 1

    lineas = _leer_registro(args.xlsx)
    if not lineas:
        print("Sin lineas en Registro (rellene el Excel y vuelva a ejecutar).")
        print("Conexion a datos OK:", args.path.resolve())
        return 0
    dias = _agrupar(lineas)

    print("=== Conexion BM ===")
    print("Excel:", args.xlsx.resolve())
    print("Datos:", args.path.resolve())
    print(f"Lineas={len(lineas)} dias={len(dias)} dry_run={args.dry_run}")

    _boot(args.path)
    data = get_container().app_data_store.get()
    hotel = ""
    if getattr(data, "configuracion", None):
        hotel = getattr(data.configuracion, "nombre_establecimiento", "") or ""
    print("Hotel:", hotel or "(sin nombre)")
    print("===================")

    rec_map = _mapa_recetas(data)
    extra_map = _mapa_extras_label()
    prod_map = _mapa_productos(data)

    por_fila: dict[int, str] = {}
    n_ok = n_skip = n_err = 0
    for dia in dias:
        result = _importar_dia(
            dia,
            dry_run=args.dry_run,
            rec_map=rec_map,
            extra_map=extra_map,
            prod_map=prod_map,
        )
        print(f"  {dia.fecha}: {result.get('mensaje')}")
        status = result.get("status") or ("OK" if result.get("ok") else "ERROR")
        for r in dia.row_indices:
            por_fila[r] = status
        if result.get("skipped"):
            n_skip += 1
        elif result.get("ok"):
            n_ok += 1
        else:
            n_err += 1

    if not args.dry_run:
        _escribir_estados(args.xlsx, por_fila)
        print("Estados escritos en columna Importado.")
    else:
        print("Dry-run: no se persistio ni se modifico el Excel.")

    print(f"Resumen: ok={n_ok} skip={n_skip} error={n_err}")
    return 0 if n_err == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
