"""Importa el Excel operativo (desayuno, comida, cena, consumo buffet) a BM-DATOS.

Uso:
  .\\.venv\\Scripts\\python.exe scripts\\import_registro_operativo_excel.py RUTA.xlsx
  .\\.venv\\Scripts\\python.exe scripts\\import_registro_operativo_excel.py RUTA.xlsx --dry-run
  .\\.venv\\Scripts\\python.exe scripts\\import_registro_operativo_excel.py RUTA.xlsx --solo Registro
  .\\.venv\\Scripts\\python.exe scripts\\import_registro_operativo_excel.py RUTA.xlsx --path RUTA\\datos_hotel.json
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.bootstrap import configure_for_flet, get_container, reset_container
from app.core.auth.roles import ROL_DIRECCION
from app.core.auth.session import ACTOR_TYPE_USUARIO, AuthSession, save_auth_session
from app.core.services import bebida_service, cena_service, comida_service, desayuno_service as des
from app.core.services.buffet_config_service import sincronizar_desde_excel
from app.core.services.buffet_consumo_service import LineaBuffetEntrada, importar_lineas_buffet
from app.core.services.pack_unidades import piezas_a_ud_paquete, ud_paquete_a_piezas
from app.core.services.receta_service import (
    ETIQUETA_TOSTADA_DEL_DIA,
    receta_tostada_del_dia,
)
from app.core.services.text_search import normalizar_texto

HOTEL_DEFAULT = Path(os.environ["LOCALAPPDATA"]) / "BM-V2-local" / "data" / "datos_hotel.json"
CLAVE_PREFIX = "desayuno-xlsx"
CLAVE_VER = "v3"
CLAVE_PREFIX_COMIDA = "comida-xlsx"
CLAVE_PREFIX_CENA = "cena-xlsx"
CLAVE_VER_SERVICIO = "v1"

CLAVE_PREFIX_BEBIDAS_DESAYUNO = "bebidas-desayuno-xlsx"

HOJAS_IMPORT = (
    "Registro", "RegistroBebidasDesayuno", "RegistroComida", "RegistroCena",
    "ConsumoBuffet", "ConfigBuffet",
)

# Labels de Extra/Omitir que son «tipo de huevo» (sustituyen el frito de Desayuno inglés).
_TIPOS_HUEVO = frozenset(
    {
        "huevo frito",
        "huevo pochado",
        "huevo cocido",
        "huevos revueltos",
        "huevo revuelto",
        "huevo",  # omitir el huevo de la ficha
        "huevo cascara",
        "sin huevo",
    }
)
_HUEVO_FRITO_PID = "p48"

# Pan: blanco (p09) ↔ integral (p11) ↔ sin gluten (p04). Sustituye el de la ficha.
_PAN_BLANCO = "p09"
_PAN_INTEGRAL = "p11"
_PAN_SIN_GLUTEN = "p04"
_PANS_SUSTITUIBLES = frozenset({_PAN_BLANCO, _PAN_INTEGRAL, _PAN_SIN_GLUTEN})
_TIPOS_PAN = frozenset(
    {
        "tostada",
        "tostada integral",
        "tostada sin gluten",
        "pan blanco",
        "pan integral",
        "pan sin gluten",
        "molde comun",
        "molde integral",
        "sin tostada",
        "sin pan",
    }
)


def _norm(s: str) -> str:
    return normalizar_texto(s or "")


def _es_tipo_huevo(label: str) -> bool:
    return _norm(label) in _TIPOS_HUEVO


def _es_tipo_pan(label: str) -> bool:
    return _norm(label) in _TIPOS_PAN


def _es_desayuno_ingles(nombre: str) -> bool:
    return _norm(nombre) == "desayuno ingles"


def _pan_destino(label: str) -> str | None:
    """Producto de pan elegido, o None si es solo omitir."""
    key = _norm(label)
    if key in ("tostada integral", "pan integral", "molde integral"):
        return _PAN_INTEGRAL
    if key in ("tostada sin gluten", "pan sin gluten"):
        return _PAN_SIN_GLUTEN
    if key in ("tostada", "pan blanco", "molde comun"):
        return _PAN_BLANCO
    return None


def _qty_pan_en_receta(receta) -> tuple[str | None, float]:
    """Devuelve (producto_id_pan, cantidad nativa) del pan en la ficha."""
    por_pid: dict[str, float] = {}
    for ing in receta.ingredientes or []:
        if ing.producto_id in _PANS_SUSTITUIBLES:
            por_pid[ing.producto_id] = por_pid.get(ing.producto_id, 0.0) + float(
                ing.cantidad or 0
            )
    if not por_pid:
        return None, 0.0
    pid = max(por_pid, key=por_pid.get)
    return pid, por_pid[pid]


def _qty_pan_sustitucion(
    from_pid: str, from_qty_nativa: float, to_pid: str, porciones: float
) -> float:
    """Misma nº de piezas/rebanadas, convertida a Ud nativa del destino."""
    piezas = ud_paquete_a_piezas(from_pid, from_qty_nativa)
    if piezas <= 0:
        piezas = 1.0
    return round(piezas_a_ud_paquete(to_pid, piezas) * float(porciones), 6)


def _omitir_pan_si_hay(pid: str) -> None:
    r = des.anadir_mod_a_receta_en_cesta(pid, -1e9)
    if not r.ok and "no está" not in (r.mensaje or "").lower():
        raise ValueError(r.mensaje)


def _omitir_todos_los_panes() -> None:
    for pid in _PANS_SUSTITUIBLES:
        _omitir_pan_si_hay(pid)


def _auth() -> None:
    save_auth_session(
        AuthSession(
            authenticated=True,
            actor_type=ACTOR_TYPE_USUARIO,
            actor_id="import-xlsx-des",
            actor_label="Import Excel desayuno",
            role=ROL_DIRECCION,
            session_id="import-xlsx-des-session",
            login_at=datetime.now(timezone.utc).isoformat(),
            terminal_id=None,
            login="import",
        )
    )


def _boot(path: Path) -> None:
    reset_container()
    configure_for_flet(data_path=str(path))
    _auth()


def _parse_fecha(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val, default: float | None = None) -> float | None:
    if val is None or val == "":
        return default
    try:
        return float(str(val).replace(",", "."))
    except ValueError:
        return default


def _cell_str(val, default: str = "") -> str:
    """Texto de celda; no pierde el 0 numérico (``0 or ''`` → vacío)."""
    if val is None:
        return default
    return str(val).strip()


def _tipo_y_huespedes_registro(
    tipo_raw,
    hues_raw,
    *,
    nombre: str = "",
) -> tuple[str, int | None]:
    """Normaliza Tipo/Huéspedes según cómo se rellena el Excel operativo.

    Casos admitidos:
    - Tipo = Receta/Extra/Producto (correcto).
    - Tipo = 0/1 y Nombre = plato → 0/1 es flag de huésped; Tipo = Receta.
    - Tipo vacío y Nombre = plato → Tipo = Receta.
    """
    tipo = _cell_str(tipo_raw)
    hues: int | None
    if hues_raw is None or hues_raw == "":
        hues = None
    else:
        hues_f = _parse_float(hues_raw, None)
        hues = int(hues_f) if hues_f is not None else None

    if tipo in ("0", "1"):
        if hues is None:
            hues = int(tipo)
        tipo = "Receta"
    elif not tipo and nombre:
        tipo = "Receta"
    return tipo, hues


@dataclass
class LineaExcel:
    row: int
    fecha: date
    huespedes: int | None
    tipo: str
    nombre: str
    cantidad: float
    extras: list[tuple[str, float | None]]  # hasta 4 (label, cant)
    omitir1: str
    omitir2: str
    notas: str


@dataclass
class DiaPlan:
    fecha: date
    huespedes: int | None = None
    lineas: list[LineaExcel] = field(default_factory=list)
    row_indices: list[int] = field(default_factory=list)


def _header_map(ws) -> dict[str, int]:
    raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[str, int] = {}
    for i, h in enumerate(raw):
        if h is None:
            continue
        key = str(h).strip()
        # Normalizar alias de cabecera Cantidad / CantN
        if key.startswith("Cantidad"):
            key = "Cantidad ↑↓"
        elif key.startswith("Cant") and len(key) >= 5 and key[4].isdigit():
            n = key[4]
            key = f"Cant{n} ↑↓"
        out[key] = i
    missing = [h for h in ("Fecha", "Tipo", "Nombre", "Cantidad ↑↓") if h not in out]
    if missing:
        raise SystemExit(f"Faltan columnas en Registro: {missing}. Cabeceras: {raw}")
    return out


def _leer_registro(path: Path) -> list[LineaExcel]:
    wb = load_workbook(path, data_only=True)
    if "Registro" not in wb.sheetnames:
        raise SystemExit("El Excel debe tener hoja «Registro».")
    ws = wb["Registro"]
    cols = _header_map(ws)
    lineas: list[LineaExcel] = []
    ultima_fecha: date | None = None
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def cell(name: str, default=None):
            idx = cols.get(name)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        fecha_celda = _parse_fecha(cell("Fecha"))
        fecha = fecha_celda or ultima_fecha
        nombre = _cell_str(cell("Nombre"))
        tipo, hues_val = _tipo_y_huespedes_registro(
            cell("Tipo"), cell("Huespedes"), nombre=nombre,
        )
        if not fecha_celda and not tipo and not nombre and not cell("Huespedes"):
            continue
        if not fecha:
            raise SystemExit(
                f"Fila {r_i}: Fecha obligatoria "
                "(o déjela en la primera fila del mismo día)."
            )
        if not nombre:
            raise SystemExit(f"Fila {r_i}: Nombre obligatorio.")
        if not tipo:
            raise SystemExit(f"Fila {r_i}: Tipo obligatorio (Receta/Extra/Producto).")
        cant = _parse_float(cell("Cantidad ↑↓"), 1.0)
        if cant is None or cant <= 0:
            raise SystemExit(f"Fila {r_i}: Cantidad debe ser > 0.")
        extras: list[tuple[str, float | None]] = []
        for n in (1, 2, 3, 4):
            lab = _cell_str(cell(f"Extra{n}"))
            if not lab:
                continue
            extras.append((lab, _parse_float(cell(f"Cant{n} ↑↓"), None)))
        lineas.append(
            LineaExcel(
                row=r_i,
                fecha=fecha,
                huespedes=hues_val,
                tipo=tipo,
                nombre=nombre,
                cantidad=float(cant),
                extras=extras,
                omitir1=_cell_str(cell("Omitir1")),
                omitir2=_cell_str(cell("Omitir2")),
                notas=_cell_str(cell("Notas")),
            )
        )
        ultima_fecha = fecha
    wb.close()
    return lineas


def _agrupar(lineas: list[LineaExcel]) -> list[DiaPlan]:
    """Misma Fecha = un solo desayuno. Huéspedes = suma de 1s (0 no suma)."""
    por: dict[date, DiaPlan] = {}
    for ln in lineas:
        dia = por.get(ln.fecha)
        if dia is None:
            dia = DiaPlan(fecha=ln.fecha)
            por[ln.fecha] = dia
        dia.lineas.append(ln)
        dia.row_indices.append(ln.row)
    for dia in por.values():
        suma = 0
        hay_marca = False
        for ln in dia.lineas:
            if ln.huespedes is None:
                continue
            hay_marca = True
            if ln.huespedes > 0:
                suma += int(ln.huespedes)
        if hay_marca and suma >= 1:
            dia.huespedes = suma
        else:
            # Sin marcas 1/0: fallback = nº de líneas Receta
            n_rec = sum(1 for ln in dia.lineas if _norm(ln.tipo) == "receta")
            dia.huespedes = max(n_rec, 1)
    return [por[k] for k in sorted(por)]


def _mapa_recetas(data) -> dict[str, object]:
    m: dict[str, object] = {}
    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        m[_norm(r.nombre)] = r
    return m


def _mapa_extras_label() -> dict[str, dict]:
    m: dict[str, dict] = {}
    for e in (
        des.extras_rapidos_desayuno()
        + des.leches_rapidas_desayuno()
        + des.bebidas_frias_rapidas_desayuno()
    ):
        m[_norm(e["label"])] = e
        m[_norm(e["nombre"])] = e
    return m


def _mapa_productos(data) -> dict[str, object]:
    m: dict[str, object] = {}
    for p in data.productos:
        if not getattr(p, "activo", True):
            continue
        m[_norm(p.nombre)] = p
        codigo = getattr(p, "codigo", None) or ""
        if codigo:
            m[_norm(codigo)] = p
        m[_norm(p.id)] = p
    return m


def _resolver_receta(data, nombre: str, fecha: date, rec_map: dict):
    if _norm(nombre) == _norm(ETIQUETA_TOSTADA_DEL_DIA):
        return receta_tostada_del_dia(fecha)
    return rec_map.get(_norm(nombre))


def _qty_extra(extra_map: dict, label: str, cant_filas: float | None) -> tuple[str, float]:
    e = extra_map.get(_norm(label))
    if e is None:
        raise ValueError(f"Extra no encontrado: «{label}»")
    mult = float(cant_filas) if cant_filas and cant_filas > 0 else 1.0
    return e["producto_id"], round(float(e["cantidad"]) * mult, 6)


def _resolver_omit_pid(data, receta, label: str, prod_map: dict, extra_map: dict) -> str:
    key = _norm(label)
    if key in ("sin huevo", "huevo", "huevo cascara"):
        return _HUEVO_FRITO_PID
    if key in ("sin tostada", "sin pan"):
        # Preferir el pan que lleve la ficha
        pid_pan, _ = _qty_pan_en_receta(receta)
        return pid_pan or _PAN_BLANCO
    # 1) label de extra rápido
    e = extra_map.get(key)
    if e:
        return e["producto_id"]
    # 2) nombre producto
    p = prod_map.get(key)
    if p:
        return p.id
    # 3) match contra ingredientes de la receta
    for ing in receta.ingredientes or []:
        prod = next((x for x in data.productos if x.id == ing.producto_id), None)
        if prod and _norm(prod.nombre) == key:
            return prod.id
        if _norm(ing.producto_id) == key:
            return ing.producto_id
    raise ValueError(f"Omitir: no se resolvió «{label}»")


def _clave(fecha: date, batch: int | None = None) -> str:
    base = f"{CLAVE_PREFIX}-{fecha.isoformat()}-{CLAVE_VER}"
    if batch is None:
        return base
    return f"{base}-b{batch}"


def _ya_importado(data, fecha: date) -> object | None:
    prefix = f"{CLAVE_PREFIX}-{fecha.isoformat()}-{CLAVE_VER}"
    return next(
        (
            d
            for d in data.desayunos
            if str(getattr(d, "clave_idempotencia", None) or "").startswith(prefix)
            and not getattr(d, "anulado", False)
        ),
        None,
    )


def _anadir_linea_a_cesta(
    ln: LineaExcel,
    *,
    fecha: date,
    rec_map: dict,
    extra_map: dict,
    prod_map: dict,
) -> None:
    data = get_container().app_data_store.get()
    tipo = _norm(ln.tipo)
    if tipo == "receta":
        rec = _resolver_receta(data, ln.nombre, fecha, rec_map)
        if rec is None:
            raise ValueError(f"Receta no encontrada: «{ln.nombre}»")
        ingles = _es_desayuno_ingles(ln.nombre)
        egg_extras: list[tuple[str, float | None]] = []
        pan_destino: str | None = None
        pan_solo_omitir = False
        other_extras: list[tuple[str, float | None]] = []
        pid_pan_ficha, qty_pan_ficha = _qty_pan_en_receta(rec)

        for lab, canti in ln.extras:
            if ingles and _es_tipo_huevo(lab) and _norm(lab) not in (
                "huevo",
                "huevo cascara",
                "sin huevo",
            ):
                egg_extras.append((lab, canti))
            elif _es_tipo_pan(lab):
                dest = _pan_destino(lab)
                if dest is None:
                    pan_solo_omitir = True
                elif pid_pan_ficha:
                    pan_destino = dest
                else:
                    other_extras.append((lab, canti))
            else:
                other_extras.append((lab, canti))

        omit_labels = [x for x in (ln.omitir1, ln.omitir2) if x]
        for lab in omit_labels:
            if _es_tipo_pan(lab) and _pan_destino(lab) is None:
                pan_solo_omitir = True

        for lab, canti in other_extras:
            pid, qty = _qty_extra(extra_map, lab, canti)
            r = des.anadir_mod_pendiente_receta(pid, qty)
            if not r.ok:
                raise ValueError(r.mensaje)
        for lab, canti in egg_extras:
            pid, qty = _qty_extra(extra_map, lab, canti)
            r = des.anadir_mod_pendiente_receta(pid, qty)
            if not r.ok:
                raise ValueError(r.mensaje)

        r = des.anadir_receta_a_cesta(rec.id, float(ln.cantidad))
        if not r.ok:
            raise ValueError(r.mensaje)

        for lab in omit_labels:
            if _es_tipo_pan(lab):
                continue
            pid = _resolver_omit_pid(data, rec, lab, prod_map, extra_map)
            r = des.anadir_mod_a_receta_en_cesta(pid, -1e9)
            if not r.ok and "no está" not in (r.mensaje or "").lower():
                raise ValueError(r.mensaje)

        if ingles and (
            egg_extras
            or any(_es_tipo_huevo(x) for x in omit_labels)
        ):
            if any(
                i.producto_id == _HUEVO_FRITO_PID
                for i in (rec.ingredientes or [])
            ):
                r = des.anadir_mod_a_receta_en_cesta(_HUEVO_FRITO_PID, -1e9)
                if (
                    not r.ok
                    and "no está" not in (r.mensaje or "").lower()
                ):
                    raise ValueError(r.mensaje)

        if pan_solo_omitir and not pan_destino:
            _omitir_todos_los_panes()
        elif pan_destino and pid_pan_ficha and pan_destino != pid_pan_ficha:
            qty_nativa = _qty_pan_sustitucion(
                pid_pan_ficha, qty_pan_ficha, pan_destino, float(ln.cantidad)
            )
            _omitir_pan_si_hay(pid_pan_ficha)
            r = des.anadir_mod_a_receta_en_cesta(pan_destino, qty_nativa)
            if not r.ok:
                raise ValueError(r.mensaje)
        for lab in omit_labels:
            if not _es_tipo_pan(lab):
                continue
            dest = _pan_destino(lab)
            if dest is None:
                continue
            _omitir_pan_si_hay(dest)
    elif tipo in ("extra", "producto"):
        if _norm(ln.nombre) in extra_map:
            pid, qty = _qty_extra(extra_map, ln.nombre, ln.cantidad)
        else:
            p = prod_map.get(_norm(ln.nombre))
            if p is None:
                raise ValueError(f"Producto/extra no encontrado: «{ln.nombre}»")
            pid = p.id
            qty = float(ln.cantidad)
        r = des.anadir_a_cesta(pid, qty)
        if not r.ok:
            raise ValueError(r.mensaje)
    else:
        raise ValueError(f"Tipo desconocido: «{ln.tipo}» (use Receta/Extra/Producto)")


def _importar_dia(
    dia: DiaPlan,
    *,
    dry_run: bool,
    rec_map: dict,
    extra_map: dict,
    prod_map: dict,
) -> dict:
    data = get_container().app_data_store.get()
    existente = _ya_importado(data, dia.fecha)
    if existente is not None:
        return {
            "ok": True,
            "skipped": True,
            "mensaje": f"Ya importado ({existente.id})",
            "status": f"SKIP {existente.id}",
        }

    huespedes = int(dia.huespedes or 1)

    if dry_run:
        preview = []
        errs: list[str] = []
        for ln in dia.lineas:
            tipo = _norm(ln.tipo)
            try:
                if tipo == "receta":
                    rec = _resolver_receta(data, ln.nombre, dia.fecha, rec_map)
                    if rec is None:
                        raise ValueError(f"Receta no encontrada: «{ln.nombre}»")
                    for lab, canti in ln.extras:
                        _qty_extra(extra_map, lab, canti)
                    for lab in (ln.omitir1, ln.omitir2):
                        if lab:
                            _resolver_omit_pid(data, rec, lab, prod_map, extra_map)
                elif tipo in ("extra", "producto"):
                    if _norm(ln.nombre) in extra_map:
                        _qty_extra(extra_map, ln.nombre, ln.cantidad)
                    elif prod_map.get(_norm(ln.nombre)) is None:
                        raise ValueError(f"Producto/extra no encontrado: «{ln.nombre}»")
                else:
                    raise ValueError(f"Tipo desconocido: «{ln.tipo}»")
                preview.append(f"{ln.tipo}:{ln.nombre}x{ln.cantidad:g}")
            except ValueError as exc:
                errs.append(f"fila {ln.row}: {exc}")
        if errs:
            return {
                "ok": False,
                "dry_run": True,
                "mensaje": " | ".join(errs),
                "status": "ERROR",
            }
        return {
            "ok": True,
            "dry_run": True,
            "mensaje": (
                f"dry-run {dia.fecha} huespedes={huespedes} "
                f"lineas={len(dia.lineas)} -> " + "; ".join(preview)
            ),
            "status": "DRY-RUN OK",
        }

    # Un día = un solo registro de desayuno (toda la cesta junta).
    des.limpiar_cesta()
    errores: list[str] = []
    for ln in dia.lineas:
        try:
            _anadir_linea_a_cesta(
                ln,
                fecha=dia.fecha,
                rec_map=rec_map,
                extra_map=extra_map,
                prod_map=prod_map,
            )
        except ValueError as exc:
            errores.append(f"fila {ln.row}: {exc}")
    if errores:
        des.limpiar_cesta()
        return {
            "ok": False,
            "mensaje": " | ".join(errores),
            "status": "ERROR",
        }
    if des.cesta_vacia():
        return {"ok": False, "mensaje": "Cesta vacía", "status": "ERROR"}

    notas = " | ".join(ln.notas for ln in dia.lineas if ln.notas)[:200]
    res = des.registrar_desayuno(
        dia.fecha,
        huespedes,
        clave_idempotencia=_clave(dia.fecha),
        observaciones=f"Import Excel desayuno operativo. {notas}".strip(),
    )
    if not res.ok:
        des.limpiar_cesta()
        return {
            "ok": False,
            "mensaje": res.mensaje,
            "status": f"ERROR {res.mensaje}",
            "codigo": getattr(res, "codigo", None),
        }
    data = get_container().app_data_store.get()
    reg = _ya_importado(data, dia.fecha)
    rid = getattr(reg, "id", "?") if reg is not None else "?"
    return {
        "ok": True,
        "mensaje": f"OK {rid} (huespedes={huespedes}, lineas={len(dia.lineas)})",
        "status": f"OK {rid}",
    }


def _escribir_estados_hoja(xlsx: Path, hoja: str, por_fila: dict[int, str]) -> None:
    wb = load_workbook(xlsx)
    if hoja not in wb.sheetnames:
        wb.close()
        return
    ws = wb[hoja]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = next((i for i, h in enumerate(headers) if h and str(h).strip() == "Importado"), None)
    if idx is None:
        idx = len(headers)
        ws.cell(1, idx + 1, "Importado")
    for row, status in por_fila.items():
        ws.cell(row, idx + 1, status)
    wb.save(xlsx)
    wb.close()


def _escribir_estados(xlsx: Path, por_fila: dict[int, str]) -> None:
    _escribir_estados_hoja(xlsx, "Registro", por_fila)


@dataclass
class LineaServicioExcel:
    row: int
    fecha: date
    tipo: str
    nombre: str
    cantidad: float
    notas: str


def _header_map_simple(ws, requeridas: tuple[str, ...]) -> dict[str, int]:
    raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[str, int] = {}
    for i, h in enumerate(raw):
        if h is None:
            continue
        key = str(h).strip()
        if key.startswith("Cantidad"):
            key = "Cantidad ↑↓"
        out[key] = i
    missing = [h for h in requeridas if h not in out]
    if missing:
        raise SystemExit(f"Faltan columnas: {missing}. Cabeceras: {raw}")
    return out


def _leer_hoja_servicio(path: Path, hoja: str) -> list[LineaServicioExcel]:
    wb = load_workbook(path, data_only=True)
    if hoja not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[hoja]
    cols = _header_map_simple(ws, ("Fecha", "Tipo", "Nombre", "Cantidad ↑↓"))
    lineas: list[LineaServicioExcel] = []
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def cell(name: str, default=None):
            idx = cols.get(name)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        fecha = _parse_fecha(cell("Fecha"))
        tipo = str(cell("Tipo") or "").strip()
        nombre = str(cell("Nombre") or "").strip()
        if not fecha and not tipo and not nombre:
            continue
        if not fecha:
            raise SystemExit(f"{hoja} fila {r_i}: Fecha obligatoria.")
        if not tipo or not nombre:
            raise SystemExit(f"{hoja} fila {r_i}: Tipo y Nombre obligatorios.")
        cant = _parse_float(cell("Cantidad ↑↓"), 1.0)
        if cant is None or cant <= 0:
            raise SystemExit(f"{hoja} fila {r_i}: Cantidad debe ser > 0.")
        lineas.append(
            LineaServicioExcel(
                row=r_i,
                fecha=fecha,
                tipo=tipo,
                nombre=nombre,
                cantidad=float(cant),
                notas=str(cell("Notas") or "").strip(),
            )
        )
    wb.close()
    return lineas


@dataclass
class DiaPlanBuffet:
    fecha: date
    lineas: list[LineaBuffetEntrada] = field(default_factory=list)
    row_indices: list[int] = field(default_factory=list)


@dataclass
class DiaPlanServicio:
    fecha: date
    lineas: list[LineaServicioExcel] = field(default_factory=list)
    row_indices: list[int] = field(default_factory=list)


def _mapa_recetas_bebidas_desayuno(data) -> dict[str, object]:
    m: dict[str, object] = {}
    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        cat = r.categoria.value if hasattr(r.categoria, "value") else str(r.categoria)
        if cat == "bebidas" and des.es_receta_bebida_desayuno(r.nombre):
            m[_norm(r.nombre)] = r
        elif "roger de flor" in _norm(r.nombre):
            m[_norm(r.nombre)] = r
    # Alias «Cava Roger de Flor» → receta Botella Roger de Flor si existe.
    roger = m.get(_norm("Botella Roger de Flor")) or m.get(_norm("Cava Roger de Flor"))
    if roger is not None:
        for alias in ("cava roger de flor", "roger de flor", "botella roger de flor"):
            m.setdefault(_norm(alias), roger)
    return m


def _mapa_productos_bebidas_desayuno(data) -> dict[str, object]:
    m: dict[str, object] = {}
    for p in data.productos:
        if not getattr(p, "activo", True) or not getattr(p, "es_bebida", False):
            continue
        servicios = [s for s in (getattr(p, "servicios_disponibles", None) or []) if isinstance(s, str)]
        if servicios and "desayuno" not in servicios and "bebidas" not in servicios:
            continue
        m[_norm(p.nombre)] = p
        codigo = getattr(p, "codigo", None) or ""
        if codigo:
            m[_norm(codigo)] = p
        m[_norm(p.id)] = p
    # Alias amigables de la terminal (Agua sin gas PET, Coca-Cola lata…).
    by_id = {p.id: p for p in data.productos}
    for e in des.bebidas_frias_rapidas_desayuno():
        prod = by_id.get(e["producto_id"])
        if prod is None:
            continue
        m[_norm(e["label"])] = prod
    return m


def _mapa_recetas_tipo(data, tipo: str) -> dict[str, object]:
    from app.core.models import CategoriaReceta

    permitidas = {CategoriaReceta.COMIDA.value, CategoriaReceta.BEBIDAS.value} if tipo == "comida" else {
        CategoriaReceta.CENA.value, CategoriaReceta.BEBIDAS.value,
    }
    m: dict[str, object] = {}
    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        cat = r.categoria.value if hasattr(r.categoria, "value") else str(r.categoria)
        if cat not in permitidas:
            continue
        m[_norm(r.nombre)] = r
    return m


def _agrupar_servicio(lineas: list[LineaServicioExcel]) -> list[DiaPlanServicio]:
    por: dict[date, DiaPlanServicio] = {}
    for ln in lineas:
        dia = por.get(ln.fecha)
        if dia is None:
            dia = DiaPlanServicio(fecha=ln.fecha)
            por[ln.fecha] = dia
        dia.lineas.append(ln)
        dia.row_indices.append(ln.row)
    return [por[k] for k in sorted(por)]


def _clave_servicio(prefix: str, fecha: date) -> str:
    return f"{prefix}-{fecha.isoformat()}-{CLAVE_VER_SERVICIO}"


def _ya_importado_servicio(data, tipo: str, fecha: date, prefix: str):
    clave = _clave_servicio(prefix, fecha)
    return next(
        (
            r
            for r in data.registros_servicio
            if r.tipo_servicio == tipo
            and getattr(r, "clave_idempotencia", None) == clave
            and not getattr(r, "anulado", False)
        ),
        None,
    )


def _anadir_linea_servicio(
    svc,
    ln: LineaServicioExcel,
    *,
    fecha: date,
    rec_map: dict,
    extra_map: dict,
    prod_map: dict,
) -> None:
    data = get_container().app_data_store.get()
    tipo = _norm(ln.tipo)
    if tipo == "receta":
        rec = rec_map.get(_norm(ln.nombre))
        if rec is None:
            raise ValueError(f"Receta no encontrada: «{ln.nombre}»")
        r = svc.anadir_receta_a_cesta(rec.id, float(ln.cantidad))
        if not r.ok:
            raise ValueError(r.mensaje)
    elif tipo in ("extra", "producto"):
        if _norm(ln.nombre) in extra_map:
            e = extra_map[_norm(ln.nombre)]
            pid, qty = e["producto_id"], float(e["cantidad"]) * float(ln.cantidad)
        else:
            p = prod_map.get(_norm(ln.nombre))
            if p is None:
                raise ValueError(f"Producto/extra no encontrado: «{ln.nombre}»")
            pid, qty = p.id, float(ln.cantidad)
        r = svc.anadir_a_cesta(pid, qty)
        if not r.ok:
            raise ValueError(r.mensaje)
    else:
        raise ValueError(f"Tipo desconocido: «{ln.tipo}»")


def _importar_dia_servicio(
    dia: DiaPlanServicio,
    *,
    svc,
    tipo: str,
    prefix: str,
    dry_run: bool,
    rec_map: dict,
    extra_map: dict,
    prod_map: dict,
) -> dict:
    data = get_container().app_data_store.get()
    existente = _ya_importado_servicio(data, tipo, dia.fecha, prefix)
    if existente is not None:
        return {
            "ok": True,
            "skipped": True,
            "mensaje": f"Ya importado ({existente.id})",
            "status": f"SKIP {existente.id}",
        }
    if dry_run:
        preview = []
        errs = []
        for ln in dia.lineas:
            try:
                _anadir_linea_servicio(
                    svc, ln, fecha=dia.fecha, rec_map=rec_map, extra_map=extra_map, prod_map=prod_map,
                )
                preview.append(f"{ln.tipo}:{ln.nombre}x{ln.cantidad:g}")
            except ValueError as exc:
                errs.append(f"fila {ln.row}: {exc}")
        svc.limpiar_cesta()
        if errs:
            return {"ok": False, "dry_run": True, "mensaje": " | ".join(errs), "status": "ERROR"}
        return {
            "ok": True,
            "dry_run": True,
            "mensaje": f"dry-run {dia.fecha} -> " + "; ".join(preview),
            "status": "DRY-RUN OK",
        }
    svc.limpiar_cesta()
    errores = []
    for ln in dia.lineas:
        try:
            _anadir_linea_servicio(
                svc, ln, fecha=dia.fecha, rec_map=rec_map, extra_map=extra_map, prod_map=prod_map,
            )
        except ValueError as exc:
            errores.append(f"fila {ln.row}: {exc}")
    if errores:
        svc.limpiar_cesta()
        return {"ok": False, "mensaje": " | ".join(errores), "status": "ERROR"}
    if svc.cesta_vacia():
        return {"ok": False, "mensaje": "Cesta vacía", "status": "ERROR"}
    notas = " | ".join(getattr(ln, "notas", "") for ln in dia.lineas if getattr(ln, "notas", ""))[:200]
    res = svc.registrar(
        dia.fecha,
        0,
        clave_idempotencia=_clave_servicio(prefix, dia.fecha),
        observaciones=f"Import Excel {tipo}. {notas}".strip(),
    )
    if not res.ok:
        svc.limpiar_cesta()
        return {"ok": False, "mensaje": res.mensaje, "status": f"ERROR {res.mensaje}"}
    reg = _ya_importado_servicio(get_container().app_data_store.get(), tipo, dia.fecha, prefix)
    rid = getattr(reg, "id", "?") if reg else "?"
    return {"ok": True, "mensaje": f"OK {rid}", "status": f"OK {rid}"}


def _leer_config_buffet(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    if "ConfigBuffet" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["ConfigBuffet"]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[idx.get("Concepto", 2)] if idx.get("Concepto") is not None else None
        if not label:
            continue
        activo_raw = row[idx.get("Activo", 9)] if idx.get("Activo") is not None else "Si"
        filas.append({
            "seccion": row[idx.get("Seccion", 0)],
            "orden": row[idx.get("Orden", 1)],
            "label": str(label).strip(),
            "producto_id": row[idx.get("ProductoId", 3)],
            "unidad": row[idx.get("Unidad", 4)],
            "cantidad_defecto": row[idx.get("CantDefecto", 5)],
            "tipo_linea": row[idx.get("Tipo", 6)],
            "producto_bote_id": row[idx.get("ProductoBote", 7)],
            "receta_id": row[idx.get("RecetaId", 8)],
            "activo": str(activo_raw or "Si").strip().lower() in ("si", "sí", "1", "true", "x"),
        })
    wb.close()
    return filas


def _leer_consumo_buffet(path: Path) -> list[DiaPlanBuffet]:
    wb = load_workbook(path, data_only=True)
    if "ConsumoBuffet" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["ConsumoBuffet"]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    por: dict[date, DiaPlanBuffet] = {}
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def cell(name: str, default=None):
            i = idx.get(name)
            if i is None or i >= len(row):
                return default
            return row[i]

        fecha = _parse_fecha(cell("Fecha"))
        concepto = str(cell("Concepto") or "").strip()
        if not fecha and not concepto:
            continue
        if not fecha or not concepto:
            raise SystemExit(f"ConsumoBuffet fila {r_i}: Fecha y Concepto obligatorios.")
        cant = _parse_float(cell("Cantidad"), None)
        if cant is None or cant <= 0:
            continue
        motivo = str(cell("Motivo") or "Consumo").strip()
        dia = por.get(fecha)
        if dia is None:
            dia = DiaPlanBuffet(fecha=fecha)
            por[fecha] = dia
        dia.row_indices.append(r_i)
        dia.lineas.append(
            LineaBuffetEntrada(
                config_id=None,
                label=concepto,
                seccion=str(cell("Seccion") or "").strip(),
                cantidad=float(cant),
                motivo=motivo,
                notas=str(cell("Notas") or "").strip(),
            )
        )
    wb.close()
    return [por[k] for k in sorted(por)]


def _procesar_buffet(
    dias: list[DiaPlanBuffet],
    *,
    dry_run: bool,
) -> tuple[dict[int, str], int, int, int]:
    por_fila: dict[int, str] = {}
    n_ok = n_skip = n_err = 0
    for dia in dias:
        result = importar_lineas_buffet(
            dia.fecha,
            dia.lineas,
            dry_run=dry_run,
        )
        print(f"  Buffet {dia.fecha}: {result.mensaje}")
        if result.skipped:
            status = f"SKIP {result.registro_id or ''}".strip()
        elif result.dry_run and result.ok:
            status = "DRY-RUN OK"
        elif result.ok:
            status = f"OK {result.registro_id or ''}".strip()
        else:
            status = "ERROR"
        for r in dia.row_indices:
            por_fila[r] = status
        if result.skipped:
            n_skip += 1
        elif result.ok:
            n_ok += 1
        else:
            n_err += 1
    return por_fila, n_ok, n_skip, n_err


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, nargs="?", default=None, help="Plantilla / archivo rellenado")
    parser.add_argument("--path", type=Path, default=HOTEL_DEFAULT)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--check",
        action="store_true",
        help="Solo verifica conexion a BM (ruta datos + catalogo), sin importar",
    )
    parser.add_argument(
        "--solo",
        choices=[
            "Registro", "RegistroBebidasDesayuno", "RegistroComida", "RegistroCena",
            "ConsumoBuffet", "ConfigBuffet",
        ],
        action="append",
        help="Importar solo las hojas indicadas (repetible)",
    )
    args = parser.parse_args()
    hojas = args.solo or list(HOJAS_IMPORT)

    if not args.path.exists():
        print("No existe datos:", args.path)
        return 1

    if args.check or args.xlsx is None:
        print("=== Conexion BM ===")
        print("Datos:", args.path.resolve())
        _boot(args.path)
        data = get_container().app_data_store.get()
        hotel = ""
        if getattr(data, "configuracion", None):
            hotel = getattr(data.configuracion, "nombre_establecimiento", "") or ""
        n_rec = sum(
            1
            for r in data.recetas
            if getattr(r, "activo", True)
            and (
                (hasattr(r.categoria, "value") and r.categoria.value == "desayuno")
                or str(r.categoria) == "desayuno"
            )
        )
        n_ext = (
            len(des.extras_rapidos_desayuno())
            + len(des.leches_rapidas_desayuno())
            + len(des.bebidas_frias_rapidas_desayuno())
        )
        print("Hotel:", hotel or "(sin nombre)")
        print(f"Recetas desayuno activas: {n_rec}")
        print(f"Extras rapidos: {n_ext}")
        print("Estado: CONECTADO")
        print("===================")
        if args.check or args.xlsx is None:
            return 0

    if not args.xlsx.exists():
        print("No existe", args.xlsx)
        return 1

    print("=== Conexion BM ===")
    print("Excel:", args.xlsx.resolve())
    print("Datos:", args.path.resolve())
    print(f"Hojas={hojas} dry_run={args.dry_run}")

    _boot(args.path)
    data = get_container().app_data_store.get()
    hotel = ""
    if getattr(data, "configuracion", None):
        hotel = getattr(data.configuracion, "nombre_establecimiento", "") or ""
    print("Hotel:", hotel or "(sin nombre)")
    print("===================")

    n_ok = n_skip = n_err = 0

    if "ConfigBuffet" in hojas:
        cfg_filas = _leer_config_buffet(args.xlsx)
        if cfg_filas:
            if args.dry_run:
                print(f"ConfigBuffet: {len(cfg_filas)} linea(s) (dry-run, no persistido)")
            else:
                n, _ = sincronizar_desde_excel(data, cfg_filas)
                get_container().app_data_store.persist(data)
                print(f"ConfigBuffet: {n} linea(s) sincronizadas")
                n_ok += 1

    extra_map = _mapa_extras_label()
    prod_map = _mapa_productos(data)

    if "Registro" in hojas:
        lineas = _leer_registro(args.xlsx)
        dias = _agrupar(lineas)
        print(f"Desayuno: lineas={len(lineas)} dias={len(dias)}")
        rec_map = _mapa_recetas(data)
        por_fila: dict[int, str] = {}
        for dia in dias:
            result = _importar_dia(
                dia,
                dry_run=args.dry_run,
                rec_map=rec_map,
                extra_map=extra_map,
                prod_map=prod_map,
            )
            print(f"  Desayuno {dia.fecha}: {result.get('mensaje')}")
            status = result.get("status") or ("OK" if result.get("ok") else "ERROR")
            for r in dia.row_indices:
                por_fila[r] = status
            if result.get("skipped"):
                n_skip += 1
            elif result.get("ok"):
                n_ok += 1
            else:
                n_err += 1
        if not args.dry_run and por_fila:
            _escribir_estados_hoja(args.xlsx, "Registro", por_fila)

    for hoja, svc, tipo, prefix, rec_map_fn, prod_map_fn in (
        (
            "RegistroBebidasDesayuno",
            bebida_service,
            "bebidas",
            CLAVE_PREFIX_BEBIDAS_DESAYUNO,
            _mapa_recetas_bebidas_desayuno,
            _mapa_productos_bebidas_desayuno,
        ),
        (
            "RegistroComida",
            comida_service,
            "comida",
            CLAVE_PREFIX_COMIDA,
            lambda d: _mapa_recetas_tipo(d, "comida"),
            _mapa_productos,
        ),
        (
            "RegistroCena",
            cena_service,
            "cena",
            CLAVE_PREFIX_CENA,
            lambda d: _mapa_recetas_tipo(d, "cena"),
            _mapa_productos,
        ),
    ):
        if hoja not in hojas:
            continue
        lineas_s = _leer_hoja_servicio(args.xlsx, hoja)
        dias_s = _agrupar_servicio(lineas_s)
        print(f"{hoja}: lineas={len(lineas_s)} dias={len(dias_s)}")
        rec_map = rec_map_fn(data)
        prod_map_local = prod_map_fn(data)
        por_fila_s: dict[int, str] = {}
        for dia in dias_s:
            result = _importar_dia_servicio(
                dia,
                svc=svc,
                tipo=tipo,
                prefix=prefix,
                dry_run=args.dry_run,
                rec_map=rec_map,
                extra_map=extra_map,
                prod_map=prod_map_local,
            )
            print(f"  {tipo.capitalize()} {dia.fecha}: {result.get('mensaje')}")
            status = result.get("status") or ("OK" if result.get("ok") else "ERROR")
            for r in dia.row_indices:
                por_fila_s[r] = status
            if result.get("skipped"):
                n_skip += 1
            elif result.get("ok"):
                n_ok += 1
            else:
                n_err += 1
        if not args.dry_run and por_fila_s:
            _escribir_estados_hoja(args.xlsx, hoja, por_fila_s)

    if "ConsumoBuffet" in hojas:
        dias_b = _leer_consumo_buffet(args.xlsx)
        print(f"ConsumoBuffet: dias={len(dias_b)}")
        por_fila_b, ok_b, skip_b, err_b = _procesar_buffet(dias_b, dry_run=args.dry_run)
        n_ok += ok_b
        n_skip += skip_b
        n_err += err_b
        if not args.dry_run and por_fila_b:
            _escribir_estados_hoja(args.xlsx, "ConsumoBuffet", por_fila_b)

    if args.dry_run:
        print("Dry-run: no se persistio ni se modifico el Excel (salvo ConfigBuffet si aplica).")
    else:
        print("Estados escritos en columna Importado.")

    print(f"Resumen: ok={n_ok} skip={n_skip} error={n_err}")
    return 0 if n_err == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
