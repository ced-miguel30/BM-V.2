"""Importa registros desayuno (1-11 ago) + TPV comida/bebidas (1-15 ago).

Fuentes:
  docs/añadidos manual/registro desayuno.xlsx
  docs/añadidos manual/_tpv_parsed.json

Idempotente por clave_idempotencia. Repone stock mínimo antes de confirmar.
"""
from __future__ import annotations

import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from statistics import median

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.bootstrap import configure_for_flet, get_container, reset_container
from app.core.auth.roles import ROL_DIRECCION
from app.core.auth.session import ACTOR_TYPE_USUARIO, AuthSession, save_auth_session
from app.core.models import CategoriaReceta, IngredienteReceta, MotivoMerma, OrigenServicioMerma, TurnoMerma
from app.core.services import bebida_service as beb
from app.core.services import comida_service as com
from app.core.services import desayuno_service as des
from app.core.services import merma_service as merma_svc
from app.core.services import receta_service as rec_svc
from app.core.services import stock_service as stock_svc
from app.core.services.inventory_batch_service import stock_disponible
from app.core.services.receta_service import receta_tostada_del_dia
from app.core.services.unidad_service import convertir_a_unidad_producto
from app.core.storage.session_store import persist_data

HOTEL = Path(os.environ["LOCALAPPDATA"]) / "BM-V2-local" / "data" / "datos_hotel.json"
TPV_JSON = ROOT / "docs" / "añadidos manual" / "_tpv_parsed.json"
XLSX = ROOT / "docs" / "añadidos manual" / "registro desayuno.xlsx"
REPORT = ROOT / "docs" / "añadidos manual" / "_import_agosto_report.json"

# Filtros opcionales (None = sin filtro). Útil para imports delta.
TPV_FECHA_MIN: date | None = None  # inclusive dd as date
TPV_FECHA_MAX: date | None = None
DESAYUNO_DIAS: set[int] | None = None  # p.ej. {14,15,16,17,18}

CLAVE_PREFIX = "import-ago26"
BARRIL_L = 20.0
PICADO_POR_COCTEL = 0.05
RESP_MERMA = "rm02"

# Peso neto (kg) por Ud de envase — conversión gr/Kg → Ud
PACK_KG: dict[str, float] = {
    "p117": 3.0,   # champiñón laminado 3KG
    "p185": 0.2,   # espinacas ~200g
    "p286": 3.0,   # alubia tomate lata 3KG
    "p122": 0.85,  # melocotón en su jugo
    "p405": 2.5,
    "p367": 2.5,   # frutas del bosque
}

PACK_L: dict[str, float] = {
    "p236": 0.75,
    "p237": 0.75,
    "p240": 0.75,
    "p106": 0.75,
    "p399": BARRIL_L,
    "p241": 0.70,
    "p143": 0.70,
    "b78": 1.0,
    "b18": 1.0,
    "p197": 1.0,
}

CATALOG = {
    "huevo": "p48",
    "huevo_liq": "p46",
    "bacon": "p14",
    "salchicha": "p20",
    "hashbrown": "p29",
    "cafe_illy": "p266",
    "leche_entera": "p127",
    "pan_tostada": "p09",
    "pan_integral": "p11",
    "salmon": "p32",
    "aguacate": "b05",
    "tomate": "p51",
    "cherry": "p52",
    "espinaca": "p185",
    "jamon_cocido": "p102",
    "queso_loncha": "p44",
    "queso_fresco": "p38",
    "champi": "p117",
    "judias": "p286",
    "mantequilla": "p152",
    "frutos_rojos": "p367",
    "cebolla": "p59",
    "pimiento": "p56",
}

SMOOTHIES = ["p354", "p07", "p333", "p334", "p336", "b02"]
HELADOS = ["b07", "p86"]  # chocolate / vainilla

COCTEL_RECIPES = {
    "Piña Colada",
    "Mojito",
    "Aperol Spritz",
    "Daiquiri",
    "Caipirinha",
    "Sex on the Beach",
    "Margarita",
    "Royal Marina",
    "Copa de sangria",
    "Sangria 1L",
    "Sangria de cava 1L",
    "Blue Hawaii",
    "Espresso Martini",
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _auth() -> None:
    save_auth_session(
        AuthSession(
            authenticated=True,
            actor_type=ACTOR_TYPE_USUARIO,
            actor_id="seed-dir",
            actor_label="Seed Dirección",
            role=ROL_DIRECCION,
            session_id="seed-session",
            login_at=datetime.now(timezone.utc).isoformat(),
            terminal_id=None,
            login="seed",
        )
    )


def _boot():
    reset_container()
    configure_for_flet(data_path=str(HOTEL))
    _auth()
    return get_container()


def _to_nativa(data, producto_id: str, cantidad: float, unidad: str | None) -> float:
    prod = next(p for p in data.productos if p.id == producto_id)
    qty = float(cantidad)
    if unidad is None or unidad == prod.unidad.value:
        return qty
    u = "gr" if unidad == "g" else unidad
    # 1 rebanada / 1 pieza de molde o bollería empaquetada
    if u in ("reb", "rebanada", "rebanadas", "pieza", "piezas", "ud_pieza"):
        from app.core.services.pack_unidades import UNIDADES_POR_PAQUETE, piezas_a_ud_paquete

        if producto_id in UNIDADES_POR_PAQUETE and prod.unidad.value == "Ud":
            return piezas_a_ud_paquete(producto_id, qty)
    if u in ("ml", "cl", "L") and prod.unidad.value == "Ud":
        litros = qty * {"ml": 0.001, "cl": 0.01, "L": 1.0}[u]
        pack = PACK_L.get(producto_id, 1.0)
        return round(litros / pack, 6) if pack > 0 else round(litros, 6)
    if u in ("ml", "cl", "L") and prod.unidad.value == "L":
        return round(qty * {"ml": 0.001, "cl": 0.01, "L": 1.0}[u], 6)
    if u in ("mg", "gr", "Kg") and prod.unidad.value == "Ud":
        gramos = qty * {"mg": 0.001, "gr": 1.0, "Kg": 1000.0}[u]
        pack = PACK_KG.get(producto_id)
        if pack and pack > 0:
            # 4 decimales: evita drift FIFO (0.063534 vs 0.0636) al acumular extras
            return round(gramos / (pack * 1000.0), 4)
        # Sin pack conocido: no tratar gramos como Ud enteras
        raise ValueError(
            f"Producto {producto_id} ({prod.nombre}) está en Ud pero no hay PACK_KG "
            f"para convertir {qty} {u}. Añada el peso neto del envase."
        )
    if u in ("mg", "gr", "Kg") and prod.unidad.value == "Kg":
        return round(qty * {"mg": 0.000001, "gr": 0.001, "Kg": 1.0}[u], 6)
    nativa = convertir_a_unidad_producto(qty, u, prod.unidad)
    return nativa if nativa > 0 else qty


def _ing(data, pid: str, qty: float, unit: str | None = None) -> IngredienteReceta:
    prod = next(p for p in data.productos if p.id == pid)
    nativa = _to_nativa(data, pid, qty, unit)
    u = unit or prod.unidad.value
    if u == "g":
        u = "gr"
    return IngredienteReceta(pid, nativa, float(qty), u)


def _find_receta(data, nombre: str):
    n = _norm(nombre)
    return next((r for r in data.recetas if r.activo and _norm(r.nombre) == n), None)


def _ensure_receta(data, nombre: str, ings: list[IngredienteReceta], servicios: list[str] | None = None) -> str:
    svcs = servicios or ["bebidas", "comida", "cena"]
    rec = _find_receta(data, nombre)
    if rec:
        rr = rec_svc.editar_receta(
            rec.id,
            nombre,
            ings,
            CategoriaReceta.BEBIDAS,
            servicios_disponibles=svcs,
            porciones_estandar=1.0,
        )
        if not rr.ok:
            raise SystemExit(f"Editar «{nombre}»: {rr.mensaje}")
        return rec.id
    rr = rec_svc.crear_receta(
        nombre,
        ings,
        CategoriaReceta.BEBIDAS,
        servicios_disponibles=svcs,
        porciones_estandar=1.0,
    )
    if not rr.ok:
        raise SystemExit(f"Crear «{nombre}»: {rr.mensaje}")
    data = get_container().app_data_store.get()
    rec = _find_receta(data, nombre)
    assert rec is not None
    return rec.id


def _ensure_bebida_flag(data, pid: str) -> None:
    for p in data.productos:
        if p.id == pid and not p.es_bebida:
            p.es_bebida = True


def ensure_recipes_and_flags(report: dict) -> None:
    data = get_container().app_data_store.get()
    for pid in SMOOTHIES + HELADOS + ["p17", "p148", "p399", "p366", "b32", "p261"]:
        _ensure_bebida_flag(data, pid)
    # pack override barril
    packs = dict(getattr(data, "pack_l_overrides", None) or {})
    packs["p399"] = BARRIL_L
    data.pack_l_overrides = packs
    persist_data(data)

    data = get_container().app_data_store.get()
    created = []
    specs = [
        ("Copa Montecillo Blanco", [_ing(data, "p236", 15, "cl")]),
        ("Copa Montecillo Rosado", [_ing(data, "p237", 15, "cl")]),
        ("Botella Montecillo Blanco", [_ing(data, "p236", 1, "Ud")]),
        ("Botella Montecillo Rosado", [_ing(data, "p237", 1, "Ud")]),
        ("Botella Roger de Flor", [_ing(data, "p240", 1, "Ud")], ["bebidas", "comida", "cena", "desayuno"]),
        ("Jarra NAO Marinera", [_ing(data, "p399", 350, "ml")]),
        ("Caña NAO Marinera", [_ing(data, "p399", 200, "ml")]),
        ("Cubitazo Estrella Galicia", [_ing(data, "b32", 6, "Ud")]),
    ]
    for item in specs:
        nombre, ings = item[0], item[1]
        svcs = item[2] if len(item) > 2 else None
        rid = _ensure_receta(data, nombre, ings, svcs)
        created.append(f"{nombre} ({rid})")
        data = get_container().app_data_store.get()
    report["recipes_ensured"] = created


# Precio real bolsa Cubiton (resto de reposiciones siguen con placeholder 2,50 €/Ud).
_PRECIOS_UNITARIOS_CONOCIDOS = {
    "p17": 0.80,   # CUBITON
    "p148": 0.90,  # CUBITON PICADO
}


def ensure_stock(producto_id: str, minimo: float, fecha: date) -> str | None:
    data = get_container().app_data_store.get()
    have = stock_disponible(data, producto_id)
    if have + 1e-9 >= minimo:
        return None
    falta = max(minimo - have, 0.01)
    prod = next((p for p in data.productos if p.id == producto_id), None)
    if prod and prod.unidad.value == "Ud" and falta < 1:
        falta = max(1.0, round(falta + 0.49))
    unit = _PRECIOS_UNITARIOS_CONOCIDOS.get(producto_id)
    if unit is not None:
        precio = max(0.01, round(falta * unit, 2))
    else:
        precio = max(1.0, round(falta * 2.5, 2))
    r = stock_svc.registrar_lote(
        producto_id,
        precio_total=precio,
        cantidad=round(falta, 4),
        fecha_compra=fecha - timedelta(days=1),
        marca_proveedor="IMPORT-AGO26",
    )
    if r.ok:
        return f"repo {producto_id} +{falta:.4g} (había {have:.4g})"
    return f"FAIL repo {producto_id}: {r.mensaje}"


def needs_from_baskets(service) -> dict[str, float]:
    need: dict[str, float] = defaultdict(float)
    for lin in service.get_cesta():
        need[lin.producto_id] += float(lin.cantidad)
    for g in service.get_cesta_recetas():
        for ing in g.ingredientes:
            if getattr(ing, "es_omision", False):
                continue
            need[ing.producto_id] += float(ing.cantidad)
    return {k: v for k, v in need.items() if v > 0}


def topup_for_service(service, fecha: date, report: dict) -> None:
    need = needs_from_baskets(service)
    for pid, qty in need.items():
        msg = ensure_stock(pid, qty, fecha)
        if msg:
            report.setdefault("reposiciones", []).append(msg)


# --- TPV ----------------------------------------------------------------------

def build_name_map(data) -> dict[str, tuple[str, str]]:
    """nombre_norm -> ('recipe'|'product'|'special', id_or_token)"""
    M: dict[str, tuple[str, str]] = {}

    def R(alias: str, receta: str):
        r = _find_receta(data, receta)
        if r:
            M[_norm(alias)] = ("recipe", r.id)

    def P(alias: str, pid: str):
        M[_norm(alias)] = ("product", pid)

    def S(alias: str, token: str):
        M[_norm(alias)] = ("special", token)

    # Comida
    R("ENSALADA CESAR", "Ensalada cesar")
    R("SANDWICH CLUB", "Sandwich club")
    R("PAPAS FRITAS", "Patatas fritas")
    R("PATATAS FRITAS", "Patatas fritas")
    R("PAN BAO CON TERIYAKI", "Pan bao")
    R("PAN BAO", "Pan bao")
    R("TARTAR DE SALMON CON AGUACATE", "Tartar salmon con aguacate")
    R("TARTAR SALMON", "Tartar salmon con aguacate")
    R("LANGOSTINOS CRUJIENT", "Langostinos crujientes")
    R("LANGOSTINOSCRUJIENT", "Langostinos crujientes")
    R("COCKTAIL DE GAMBAS", "Cocktail de gambas")
    R("ENSALADADEQUESOFRESCOYFRESAS", "Ensalada queso de cabra y fresa")
    R("ENSALADA QUESO", "Ensalada queso de cabra y fresa")
    R("ENSALADA DE QUESO DE CABRA Y FRESAS", "Ensalada queso de cabra y fresa")
    R("ENSALADA DE QUESO DE CABRAS Y FRESASY FRESAS", "Ensalada queso de cabra y fresa")
    R("ENSALADA QUESO DE CABRA Y FRESAS", "Ensalada queso de cabra y fresa")
    R("ENSALADADEQUESODECABRAYFRESAS", "Ensalada queso de cabra y fresa")
    R("POKE BOWL", "Poke bowl")
    R("POKEBOWL", "Poke bowl")
    R("GAMBAS AL AJILLO", "Gambas al ajillo")
    R("CALAMARES FRITOS", "Calamares a la romana")
    R("CALAMARES A LA ROMANA", "Calamares a la romana")
    R("BACALAO TEMPURA", "Bacalao en tempura")
    R("NACHOS CON GUACAMOLE", "Nachos con guacamole")
    R("NACHOS", "Nachos con guacamole")
    R("CROQUETAS VARIADAS", "Croquetas jamon serrano")
    R("CROQUETAS DE JAMONIBERICO", "Croquetas jamon serrano")
    R("CROQUETAS DE JAMON", "Croquetas jamon serrano")
    R("HAMBURGUESA BREEZE", "Hamburguesa")
    R("HAMBURGUESA", "Hamburguesa")
    R("COPA DE HELADO 3 BOLAS", "Bola de helado")
    R("BOLAS HELADOS ARTESANALES", "Bola de helado")

    # Café
    R("EXPRESSO", "Espresso")
    R("ESPRESSO", "Espresso")
    R("CAFE LATTE", "Cafe con leche")
    R("CAFE CAPPUCCINO", "Capuchino")
    R("CAPPUCCINO", "Capuchino")
    R("AMERICANO", "Americano")
    R("CORTADO", "Cortado")

    # Cócteles
    R("PINA COLADA", "Piña Colada")
    R("MOJITO", "Mojito")
    R("APEROL SPLITZ", "Aperol Spritz")
    R("APEROL SPRITZ", "Aperol Spritz")
    R("DAIKIRI", "Daiquiri")
    R("DAIQUIRI", "Daiquiri")
    R("CAIPIRINA", "Caipirinha")
    R("CAIPIROSKA", "Caipirinha")
    R("SEX ON THE BEACH", "Sex on the Beach")
    R("MARGARITA", "Margarita")
    R("COCKTAIL ROYAL MARINA", "Royal Marina")
    R("ROYAL MARINA", "Royal Marina")
    R("COPA SANGRIA", "Copa de sangria")
    R("SANGRIA 1L", "Sangria 1L")
    R("SANGRIA DE CAVA 1L", "Sangria de cava 1L")
    R("APEROL", "Aperol")

    # Vinos / cava
    R("COPA DEVINO MONTECILLO BLANCO", "Copa Montecillo Blanco")
    R("COPA DE VINO MONTECILLO BLANCO", "Copa Montecillo Blanco")
    R("COPADEVINO MONTECILLO BLANCO", "Copa Montecillo Blanco")
    R("COPADEVINOMONTECILLO BLANCO", "Copa Montecillo Blanco")
    R("COPA DEVINOMONTECILLO BLANCO", "Copa Montecillo Blanco")
    R("COPA VINO MONTECILLO BLANCO", "Copa Montecillo Blanco")
    R("COPA DE VINO MONTECILLO ROSADO", "Copa Montecillo Rosado")
    R("COPADEVINO MONTECILLO ROSADO", "Copa Montecillo Rosado")
    R("COPADEVINOMONTECILLOROSADO", "Copa Montecillo Rosado")
    R("COPA VINO MONTECILLO ROSADO", "Copa Montecillo Rosado")
    R("COPA VINO MONTECILLO CRIANZA", "Copa Montecillo Crianza")
    R("VINO BLANCO MONTECILLO", "Botella Montecillo Blanco")
    R("ROYAR DE FLOR", "Botella Roger de Flor")
    R("ROGER DE FLOR", "Botella Roger de Flor")
    R("EL GRIFO BLANCO", "Botella El Grifo")
    R("ANNACODORNIUPRIMAVIDE", "Botella Codorniu Prima Vides")
    R("ANNA CODORNIU MINI", "Botella Anna Codorniu Mini")
    R("ANNA CODORNIU", "Botella Anna Codorniu Mini")

    # Cervezas
    R("JARRA NAO GRANDE", "Jarra NAO Marinera")
    R("JARRA NAOGRANDE", "Jarra NAO Marinera")
    R("JARRA NAO", "Jarra NAO Marinera")
    R("CANA NAO MARINERA", "Caña NAO Marinera")
    R("CAÑA NAO MARINERA", "Caña NAO Marinera")
    R("CUBITAZO", "Cubitazo Estrella Galicia")
    P("NAO MARINERA", "p366")
    P("ESTRELLA GALICIA MINI", "b32")
    P("ESTRELLA DAM", "b37")
    P("ESTRELLA DAMM", "b37")
    P("ETRELLA GALICIA 0 0", "b36")
    P("HEINEKEN", "p150")
    P("CORONITA", "b85")
    P("DORADA ESPECIAL", "b35")

    # Soft
    P("AGUA SIN GAS", "p264")  # Aquabona PET 50cl
    P("AGUA CON GAS", "p265")  # Aquabona Singular PET 50cl
    P("COCA COLA ZERO", "b60")
    P("COCA-COLA ZERO", "b60")
    P("COCA COLA", "b59")
    P("COCA-COLA", "b59")
    P("FANTA LIMON", "b83")
    P("FANTA NARANJA", "b61")
    P("SPRITE", "b84")
    P("TONICA", "p261")
    P("NESTEAMANGO PINA", "p258")
    P("NESTE MANGO PINA", "p258")
    P("ZUMO NARANJA NATURAL", "b28")

    # Spirits 4cl recipes
    R("TANQUERAY", "Tanqueray")
    R("GORDON DRY", "Gordon's")
    R("GORDONS", "Gordon's")
    R("BEEFEATERS", "Beefeater")
    R("BEEFEATER", "Beefeater")
    R("MALIBU", "Malibu")  # may miss — fallback product
    R("MARTINI ROJO", "Martini Rosso")

    # Cubetas / especiales
    S("CUBETAPEQUENA", "ice_small")
    S("CUBETA PEQUENA", "ice_small")
    S("CUBETA HIELO GRANDE", "ice_large")
    S("CUBETA HIELO", "ice_small")
    S("SMOOTHIE", "smoothie")
    S("HELADO", "helado")
    S("COCTEL DEL DIA", "pending_cocktail")
    S("COCTEL DEL DÍA", "pending_cocktail")

    return M


def match_map(M: dict, nombre: str) -> tuple[str, str] | None:
    n = _norm(nombre)
    if n in M:
        return M[n]
    # fuzzy contains longest key
    best = None
    best_len = 0
    for k, v in M.items():
        if k and (k in n or n in k) and len(k) > best_len:
            best = v
            best_len = len(k)
    if best:
        return best
    # keyword fallbacks
    keywords = [
        ("COCTEL DEL", ("special", "pending_cocktail")),
        ("CUBETA", ("special", "ice_small")),
        ("CUBITAZO", M.get(_norm("CUBITAZO"))),
        ("SMOOTHIE", ("special", "smoothie")),
        ("HELADO", ("special", "helado")),
        ("JARRA", M.get(_norm("JARRA NAO GRANDE"))),
        ("CANA NAO", M.get(_norm("CANA NAO MARINERA"))),
        ("CAÑA NAO", M.get(_norm("CANA NAO MARINERA"))),
        ("MONTECILLO BLANCO", M.get(_norm("COPA DEVINO MONTECILLO BLANCO"))),
        ("MONTECILLO ROSADO", M.get(_norm("COPADEVINO MONTECILLO ROSADO"))),
        ("MONTECILLO CRIANZA", M.get(_norm("COPA VINO MONTECILLO CRIANZA"))),
        ("ROYAR", M.get(_norm("ROYAR DE FLOR"))),
        ("ROGER", M.get(_norm("ROYAR DE FLOR"))),
        ("CODORNI", M.get(_norm("ANNA CODORNIU MINI"))),
        ("GRIFO", M.get(_norm("EL GRIFO BLANCO"))),
        ("SPLITZ", M.get(_norm("APEROL SPLITZ"))),
        ("SPRITZ", M.get(_norm("APEROL SPLITZ"))),
        ("PINA COLADA", M.get(_norm("PINA COLADA"))),
        ("MOJITO", M.get(_norm("MOJITO"))),
        ("DAIKIRI", M.get(_norm("DAIKIRI"))),
        ("CESAR", M.get(_norm("ENSALADA CESAR"))),
        ("CLUB", M.get(_norm("SANDWICH CLUB"))),
        ("FRITAS", M.get(_norm("PAPAS FRITAS"))),
        ("CROQUETA", M.get(_norm("CROQUETAS VARIADAS"))),
        ("HAMBURG", M.get(_norm("HAMBURGUESA BREEZE"))),
        ("LATTE", M.get(_norm("CAFE LATTE"))),
        ("EXPRESSO", M.get(_norm("EXPRESSO"))),
        ("CAPPUCCINO", M.get(_norm("CAFE CAPPUCCINO"))),
        ("COCA", M.get(_norm("COCA-COLA"))),
        ("AGUA SIN", M.get(_norm("AGUA SIN GAS"))),
        ("AGUA CON", M.get(_norm("AGUA CON GAS"))),
        ("HEINEKEN", M.get(_norm("HEINEKEN"))),
        ("NAO MARINERA", M.get(_norm("NAO MARINERA"))),
        ("FANTA LIMON", M.get(_norm("FANTA LIMON"))),
        ("FANTA NARANJA", M.get(_norm("FANTA NARANJA"))),
        ("SPRITE", M.get(_norm("SPRITE"))),
        ("TONICA", M.get(_norm("TONICA"))),
        ("NESTE", ("product", "p258")),
        ("MALIBU", ("product", "b18")),
        ("TANQUER", M.get(_norm("TANQUERAY"))),
        ("GORDON", M.get(_norm("GORDON DRY"))),
        ("BEEFEAT", M.get(_norm("BEEFEATERS"))),
        ("MARTINI", M.get(_norm("MARTINI ROJO"))),
        ("SANGRIA DE CAVA", M.get(_norm("SANGRIA DE CAVA 1L"))),
        ("SANGRIA", M.get(_norm("SANGRIA 1L"))),
        ("ZUMO", M.get(_norm("ZUMO NARANJA NATURAL"))),
    ]
    for needle, val in keywords:
        if needle in n and val:
            return val
    return None


def _parse_fecha(val) -> date:
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no parseable: {val!r}")


def infer_qty_for_group(lines: list[dict]) -> list[tuple[date, float, float, str]]:
    imports = [float(x["importe"]) for x in lines if float(x.get("importe") or 0) > 0]
    unit = median(sorted(imports)[: max(1, len(imports) // 2)]) if imports else 1.0
    if unit <= 0:
        unit = 1.0
    out = []
    for x in lines:
        f = _parse_fecha(x["fecha"])
        imp = float(x.get("importe") or 0)
        if imp <= 0:
            q = 1.0
        else:
            q = max(1, int(round(imp / unit)))
            if q > 1 and abs(imp / q - unit) > unit * 0.45:
                q2 = max(1, q - 1)
                if abs(imp / q2 - unit) < abs(imp / q - unit):
                    q = q2
        out.append((f, float(q), imp, x.get("nombre", "")))
    return out


def is_comida_recipe(data, rid: str) -> bool:
    r = next((x for x in data.recetas if x.id == rid), None)
    return bool(r and r.categoria == CategoriaReceta.COMIDA)


def import_tpv(
    report: dict,
    rows: list[dict] | None = None,
    *,
    observaciones: str | None = None,
) -> None:
    data = get_container().app_data_store.get()
    if rows is None:
        rows = json.loads(TPV_JSON.read_text(encoding="utf-8"))
    obs_base = observaciones or "Import TPV agosto 2026"
    M = build_name_map(data)

    # Fix Malibu if no recipe
    if _norm("MALIBU") not in M and _find_receta(data, "Malibu") is None:
        M[_norm("MALIBU")] = ("product", "b18")

    by_name: dict[str, list] = defaultdict(list)
    for r in rows:
        by_name[_norm(r["nombre"])].append(r)

    comida: dict[date, dict] = defaultdict(lambda: {"recipes": Counter(), "products": Counter()})
    bebida: dict[date, dict] = defaultdict(lambda: {"recipes": Counter(), "products": Counter()})
    pending = []
    unmapped = []
    cocktails: dict[date, int] = defaultdict(int)
    smoothie_i = 0
    helado_i = 0

    for name, lines in by_name.items():
        hit = match_map(M, name)
        if not hit:
            for ln in lines:
                unmapped.append(ln)
            continue
        kind, token = hit
        for f, qty, imp, raw in infer_qty_for_group(lines):
            if TPV_FECHA_MIN and f < TPV_FECHA_MIN:
                continue
            if TPV_FECHA_MAX and f > TPV_FECHA_MAX:
                continue
            if kind == "special" and token == "pending_cocktail":
                from app.core.services.receta_service import receta_coctel_del_dia

                rec_dia = receta_coctel_del_dia(f)
                if rec_dia is not None:
                    bebida[f]["recipes"][rec_dia.id] += qty
                    if rec_dia.nombre in COCTEL_RECIPES:
                        cocktails[f] += int(qty)
                else:
                    pending.append(
                        {
                            "fecha": f.isoformat(),
                            "qty": qty,
                            "importe": imp,
                            "nombre": raw or name,
                        }
                    )
                continue
            if kind == "special" and token == "smoothie":
                for _ in range(int(qty)):
                    pid = SMOOTHIES[smoothie_i % len(SMOOTHIES)]
                    smoothie_i += 1
                    bebida[f]["products"][pid] += 1
                continue
            if kind == "special" and token == "helado":
                for _ in range(int(qty)):
                    pid = HELADOS[helado_i % len(HELADOS)]
                    helado_i += 1
                    comida[f]["products"][pid] += 0.1  # ~bola desde cubeta 5L
                continue
            if kind == "special" and token.startswith("ice_"):
                n = 12 if token == "ice_large" else 6
                bebida[f]["products"]["p17"] += n * qty
                continue
            if kind == "recipe":
                bucket = comida if is_comida_recipe(data, token) else bebida
                bucket[f]["recipes"][token] += qty
                r = next(x for x in data.recetas if x.id == token)
                if r.nombre in COCTEL_RECIPES:
                    cocktails[f] += int(qty)
                continue
            if kind == "product":
                # helados sueltos → comida; resto bebidas
                if token in HELADOS:
                    comida[f]["products"][token] += qty
                else:
                    bebida[f]["products"][token] += qty
                continue
            unmapped.append({"fecha": f.isoformat(), "nombre": name, "qty": qty})

    report["tpv_pending_coctel_dia"] = pending
    report["tpv_unmapped"] = unmapped
    report["tpv_cocktail_counts"] = {k.isoformat(): v for k, v in cocktails.items()}

    def flush(day: date, bucket: dict, service, tipo: str):
        if not bucket["recipes"] and not bucket["products"]:
            return
        clave = f"{CLAVE_PREFIX}-{tipo}-{day.isoformat()}"
        hist = service.historial_ordenado()
        if any(getattr(x, "clave_idempotencia", None) == clave and not getattr(x, "anulado", False) for x in hist):
            report.setdefault("skipped_idempotent", []).append(clave)
            return
        service.limpiar_cesta()
        for rid, porc in bucket["recipes"].items():
            r = service.anadir_receta_a_cesta(rid, float(porc))
            if not r.ok:
                report.setdefault("errors", []).append(f"{clave} recipe {rid}: {r.mensaje}")
        for pid, qty in bucket["products"].items():
            # marcar bebida si hace falta
            data_now = get_container().app_data_store.get()
            _ensure_bebida_flag(data_now, pid)
            if tipo == "bebidas":
                persist_data(data_now)
            r = service.anadir_a_cesta(pid, float(qty))
            if not r.ok:
                report.setdefault("errors", []).append(f"{clave} product {pid}: {r.mensaje}")
        if service.cesta_vacia():
            report.setdefault("errors", []).append(f"{clave}: cesta vacía tras errores")
            return
        topup_for_service(service, day, report)
        # re-check stock once more
        topup_for_service(service, day, report)
        res = service.registrar(
            day,
            observaciones=f"{obs_base} ({tipo})",
            clave_idempotencia=clave,
        )
        if res.ok:
            data_now = get_container().app_data_store.get()
            reg = next(
                (x for x in data_now.registros_servicio if getattr(x, "clave_idempotencia", None) == clave),
                None,
            )
            coste = sum(getattr(l, "coste", 0) or 0 for l in (getattr(reg, "lineas", None) or [])) if reg else None
            report.setdefault("ok", []).append(
                {
                    "clave": clave,
                    "ref": getattr(reg, "id", None),
                    "coste": coste,
                    "tipo": tipo,
                    "fecha": day.isoformat(),
                    "msg": res.mensaje,
                }
            )
        else:
            detail = getattr(res, "detalle_stock", None) or []
            report.setdefault("errors", []).append(f"{clave}: {res.mensaje} | {detail[:8]}")
        service.limpiar_cesta()

    days = sorted(set(comida) | set(bebida) | set(cocktails))
    for day in days:
        if day in comida:
            flush(day, comida[day], com, "comida")
        if day in bebida or day in cocktails:
            if day in cocktails and cocktails[day] > 0:
                bebida[day]["products"]["p148"] += PICADO_POR_COCTEL * cocktails[day]
            flush(day, bebida[day], beb, "bebidas")
        # merma hielo 10%
        n_c = cocktails.get(day, 0)
        if n_c > 0:
            register_ice_merma(day, n_c, report)


def register_ice_merma(day: date, n_cocktails: int, report: dict) -> None:
    merma_qty = PICADO_POR_COCTEL * n_cocktails * 0.10
    note_key = f"{CLAVE_PREFIX}-merma-hielo-{day.isoformat()}"
    comentario = f"{note_key} | Descongelación hielo picado (~10% de {n_cocktails} cócteles)"
    data = get_container().app_data_store.get()
    existing = getattr(data, "mermas", None) or []
    for m in existing:
        if getattr(m, "anulado", False):
            continue
        for lin in getattr(m, "lineas", None) or []:
            if note_key in (getattr(lin, "comentario", None) or ""):
                report.setdefault("skipped_idempotent", []).append(note_key)
                return
    msg = ensure_stock("p148", merma_qty, day)
    if msg:
        report.setdefault("reposiciones", []).append(msg)
    data = get_container().app_data_store.get()
    lote = next(
        (
            l
            for l in data.lotes
            if l.producto_id == "p148" and float(l.cantidad_restante) >= merma_qty - 1e-9
        ),
        None,
    )
    if not lote:
        report.setdefault("errors", []).append(f"merma {day}: sin lote p148")
        return
    merma_svc.limpiar_cesta_merma()
    resp = next(r for r in data.responsables_merma if r.id == RESP_MERMA)
    r = merma_svc.anadir_a_cesta_merma(
        lote.id,
        merma_qty,
        MotivoMerma.OTRO.value,
        OrigenServicioMerma.BEBIDAS.value,
        comentario=comentario,
        turno_snapshot=TurnoMerma.TARDE.value,
        responsable_id=resp.id,
        responsable_nombre=resp.nombre,
    )
    if not r.ok:
        report.setdefault("errors", []).append(f"merma cesta {day}: {r.mensaje}")
        return
    res = merma_svc.registrar_merma(day)
    if res.ok:
        report.setdefault("mermas_hielo", []).append(
            {"fecha": day.isoformat(), "qty": merma_qty, "cocktails": n_cocktails}
        )
    else:
        report.setdefault("errors", []).append(f"merma {day}: {res.mensaje}")
    merma_svc.limpiar_cesta_merma()


# --- Desayunos ----------------------------------------------------------------

def _qty_prefix(text: str) -> tuple[int, str]:
    m = re.match(r"^(\d+)\s+(.*)$", text.strip(), re.I)
    if m:
        return int(m.group(1)), m.group(2)
    return 1, text


def parse_desayuno_line(data, text: str, fecha: date) -> dict:
    raw = text
    t = _norm(text)
    # typos frecuentes (evitar substrings que rompan palabras correctas)
    t = re.sub(r"\bDESAYUO\b", "DESAYUNO", t)
    t = re.sub(r"\bDEAYUNO\b", "DESAYUNO", t)
    t = re.sub(r"\bDESYAUNO\b", "DESAYUNO", t)
    t = re.sub(r"\bDESAYUNI\b", "DESAYUNO", t)
    t = re.sub(r"\bNGLES\b", "INGLES", t)
    t = re.sub(r"\bTROTILLA\b", "TORTILLA", t)
    t = re.sub(r"\bTORTILA\b", "TORTILLA", t)
    t = re.sub(r"\bRANCESA\b", "FRANCESA", t)
    t = re.sub(r"\bPOCHE\b", "POCHADO", t)
    t = re.sub(r"\bEXTAR\b", "EXTRA", t)
    t = re.sub(r"\bSALCHCIHA\b", "SALCHICHA", t)
    t = re.sub(r"\bSALCHICA\b", "SALCHICHA", t)
    t = re.sub(r"\bINMGLES\b", "INGLES", t)
    t = re.sub(r"\bTOSTDA\b", "TOSTADA", t)
    t = re.sub(r"\bQUYESO\b", "QUESO", t)
    t = re.sub(r"HUEVOSHUEVO", "HUEVO", t)
    t = re.sub(r"HUEVOS(?=HUEVO)", "HUEVOS ", t)
    t = t.replace("FRANBCESA", "FRANCESA")
    t = t.replace("DESAYUNO NGLES", "DESAYUNO INGLES")
    t = t.replace("TOSTADA RANCESA", "TOSTADA FRANCESA")
    if t.startswith("TOSTADA DIA") or t == "TOSTADA DIA":
        t = "TOSTADA DEL DIA"
    if "TOSTADAS CHAMPI" in t:
        t = t.replace("TOSTADAS CHAMPI", "TOSTADA CHAMPI")
    # X2 / 1 prefijos
    t = re.sub(r"^X\s*(\d+)\s+", r"\1 ", t)
    # "2 TOSTADAS DEL DIA" → cantidad en texto
    t = re.sub(r"\bTOSTADAS DEL DIA\b", "TOSTADA DEL DIA", t)
    recipes: list[tuple[str, float, list[tuple[str, float, str]]]] = []
    products: list[tuple[str, float, str]] = []
    notes: list[str] = []

    def add_rec(nombre: str, porc: float, extras: list | None = None):
        if nombre == "Tostada del dia":
            r = receta_tostada_del_dia(fecha)
            if not r:
                notes.append("sin tostada weekday")
                return
            recipes.append((r.nombre, porc, extras or []))
            return
        recipes.append((nombre, porc, extras or []))

    # Desayuno inglés
    if "DESAYUNO INGLES" in t or "DESAYUNO INGLÉS" in _norm(raw):
        extras = []
        if "EXTRA PAN" in t or ("PAN" in t and "EXTRA" in t):
            extras.append(("pan_tostada", 1, "reb"))
        if "EXTRA SALCHICHA" in t or "EXTAR SALCHICHA" in t:
            extras.append(("salchicha", 50, "gr"))
        add_rec("Desayuno ingles", 1, extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "SANDWICH DE LA CASA" in t or "SANDWICH CASA" in t:
        add_rec("Sandwich de la casa", 1)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}
    if "SANDWICH MIXTO" in t or "SANDWICH DE QUESO" in t or "SANDWICH SOLO QUESO" in t:
        add_rec("Sandwich mixto", 1)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}
    if "SANDWICH" in t and ("JAMON" in t or "QUESO" in t):
        extras = []
        if "CHAMPI" in t:
            extras.append(("champi", 20, "gr"))
        add_rec("Sandwich mixto", 1, extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}
    if "SANDWICH VEGETAL" in t:
        add_rec("Sandwich vegetal", 1)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "TOSTADA FRANCES" in t:
        n_tost = 1
        m = re.search(r"(\d+)\s*TOSTADA", t)
        if m:
            n_tost = int(m.group(1))
        add_rec("Tostada francesa", float(n_tost))
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "TOSTADA CHAMPI" in t:
        extras = []
        if "BACON" in t:
            extras.append(("bacon", 15, "gr"))
        add_rec("Tostada champinones", 1, extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "TOSTADA DEL DIA" in t or "TOSTADA DEL DÍA" in _norm(raw) or t == "TOSTADA DIA":
        extras = []
        if "POCHAD" in t:
            extras.append(("huevo", 1, "Ud"))
        if "AGUACATE" in t:
            extras.append(("aguacate", 60, "gr"))
        if "SALMON" in t:
            extras.append(("salmon", 40, "gr"))
        n_tost = 1
        m = re.search(r"(\d+)\s*TOSTADA", t)
        if m:
            n_tost = int(m.group(1))
        add_rec("Tostada del dia", float(n_tost), extras)
        if "HUEVO POCHAD" in t and "Y" in t:
            add_rec("Huevo pochado", 1)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "TOSTADA" in t and "SALMON" in t and "DEL DIA" not in t:
        add_rec("Tostada del dia", 1, [("salmon", 40, "gr")])
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    # Solo jamón/queso sueltos (sin sandwich/tortilla)
    if ("JAMON SERRANO" in t or "JAMON" in t) and "TORTILLA" not in t and "SANDWICH" not in t and "HUEVO" not in t:
        n = 1
        m = re.search(r"X\s*(\d+)", t) or re.search(r"(\d+)\s*$", t)
        if "X2" in t or "X 2" in t:
            n = 2
        products.append(("jamon_cocido", 20.0 * n, "gr"))  # sin serrano en catálogo
        if "QUESO" in t:
            products.append(("queso_loncha", 20.0 * n, "gr"))
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "TOSTADA" in t and ("AGUACATE" in t or "SALMON" in t or "POCHAD" in t):
        extras = []
        if "AGUACATE" in t:
            extras.append(("aguacate", 60, "gr"))
        if "SALMON" in t:
            extras.append(("salmon", 40, "gr"))
        if "POCHAD" in t:
            extras.append(("huevo", 1, "Ud"))
        add_rec("Tostada del dia", 1, extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    if "TOSTADA" in t and ("ALUBIA" in t or "JUDIA" in t or "MARRON" in t or "INTEGRAL" in t):
        extras = []
        if "ALUBIA" in t or "JUDIA" in t:
            extras.append(("judias", 20, "gr"))
        if "INTEGRAL" in t:
            extras.append(("pan_integral", 1, "reb"))
            # tostadas integrales sueltas
            m = re.search(r"(\d+)\s*TOSTADA", t)
            n = int(m.group(1)) if m else 1
            products.append(("pan_integral", float(n), "reb"))
            if "ALUBIA" in t or "JUDIA" in t:
                products.append(("judias", 20.0 * n, "gr"))
            return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}
        add_rec("Tostada del dia", 1, extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    # Tortilla
    if "TORTILLA" in t:
        extras = []
        for key, needles, qty, unit in [
            ("queso_loncha", ["QUESO"], 20, "gr"),
            ("jamon_cocido", ["JAMON"], 20, "gr"),
            ("champi", ["CHAMPI"], 20, "gr"),
            ("cebolla", ["CEBOLLA"], 20, "gr"),
            ("tomate", ["TOMATE"], 30, "gr"),
            ("pimiento", ["PIMIENTO"], 20, "gr"),
            ("hashbrown", ["HASH", "HASBROWN"], 70, "gr"),
            ("espinaca", ["ESPINAC"], 15, "gr"),
            ("pan_tostada", [" PAN", "Y PAN"], 1, "reb"),
        ]:
            if any(n in t for n in needles):
                extras.append((key, qty, unit))
        add_rec("Tortilla", 1, extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    # Huevos
    n_pref, rest = _qty_prefix(raw)
    tr = _norm(rest)

    if "HUEVO" in t or "POCHAD" in t or "REVUELT" in t or "FRITO" in t or "COCID" in t or "PASADO" in t:
        m = re.search(r"(\d+)\s*HUEVO", t)
        n_huevos = int(m.group(1)) if m else (n_pref if "HUEVO" in t else 1)
        extras = []
        if "BACON" in t or "BEICON" in t:
            nb = re.search(r"(\d+)\s*BACON", t)
            extras.append(("bacon", 15 * (int(nb.group(1)) if nb else 1), "gr"))
        if "HASH" in t or "HASBROWN" in t:
            nh = re.search(r"(\d+)\s*HASH", t) or re.search(r"(\d+)\s*HASB", t)
            extras.append(("hashbrown", 70 * (int(nh.group(1)) if nh else 1), "gr"))
        if "AGUACATE" in t:
            extras.append(("aguacate", 60, "gr"))
        if "SALMON" in t:
            extras.append(("salmon", 40, "gr"))
        if "PAN INTEGRAL" in t:
            extras.append(("pan_integral", 1, "reb"))
        elif re.search(r"\bPAN\b", t) or "2 PAN" in t:
            np_ = re.search(r"(\d+)\s*PAN", t)
            extras.append(("pan_tostada", int(np_.group(1)) if np_ else 1, "reb"))
        if "JUDIA" in t or "ALUBIA" in t:
            extras.append(("judias", 20, "gr"))
        if "SALCHICHA" in t:
            extras.append(("salchicha", 50, "gr"))

        if "DURO" in t or "COCID" in t:
            add_rec("Huevo cocido", float(n_huevos), extras)
        elif "POCHAD" in t or "PASADO" in t:
            add_rec("Huevo pochado", float(n_huevos), extras)
        elif "FRITO" in t:
            add_rec("Huevo frito", float(n_huevos), extras)
        elif "REVUELT" in t:
            products.append(("huevo", float(n_huevos), "Ud"))
            products.extend(extras)
        else:
            add_rec("Huevo frito", float(n_huevos), extras)
        return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    # Solo extras / pan / etc.
    if "AGUACATE" in t or "SALMON" in t or "PAN" in t or "BACON" in t or "HASH" in t or "SALCHICHA" in t or "TOMATE" in t or "CHERRY" in t:
        if "AGUACATE" in t:
            products.append(("aguacate", 60, "gr"))
        if "SALMON" in t:
            products.append(("salmon", 40, "gr"))
        if "PAN INTEGRAL" in t:
            products.append(("pan_integral", 1, "reb"))
        elif "PAN" in t:
            products.append(("pan_tostada", 1, "reb"))
        if "BACON" in t:
            products.append(("bacon", 15, "gr"))
        if "HASH" in t or "HASB" in t:
            products.append(("hashbrown", 70, "gr"))
        if "SALCHICHA" in t:
            products.append(("salchicha", 50, "gr"))
        if "QUESO" in t:
            products.append(("queso_loncha", 20, "gr"))
        if "CHAMPI" in t:
            products.append(("champi", 20, "gr"))
        if "TOMATE" in t or "CHERRY" in t:
            n_t, _ = _qty_prefix(raw)
            products.append(("tomate" if "CHERRY" not in t else "cherry", 30.0 * n_t, "gr"))
        if products:
            return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}

    notes.append(f"sin_parse: {raw[:140]}")
    return {"recipes": recipes, "products": products, "notes": notes, "raw": raw}


def import_desayunos(report: dict) -> None:
    wb = load_workbook(XLSX, data_only=True)
    by_day: dict[date, list[dict]] = defaultdict(list)
    parse_notes: list[str] = []

    for sheet_name in wb.sheetnames:
        m = re.match(r"^(\d{1,2})", str(sheet_name).strip())
        if not m:
            continue
        day_n = int(m.group(1))
        if not 1 <= day_n <= 31:
            continue
        if DESAYUNO_DIAS is not None and day_n not in DESAYUNO_DIAS:
            continue
        fecha = date(2026, 8, day_n)
        ws = wb[sheet_name]
        data = get_container().app_data_store.get()
        for row in ws.iter_rows(min_row=1, values_only=True):
            cells = [c for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            text = " | ".join(str(c).strip() for c in cells)
            if _norm(text) in {"PEDIDO", "NOMBRE", "HABITACION", "OBS", "OBSERVACIONES"}:
                continue
            parsed = parse_desayuno_line(data, text, fecha)
            by_day[fecha].append(parsed)
            parse_notes.extend(parsed["notes"])

    report["desayuno_parse_notes"] = parse_notes[:300]
    report["desayuno_parse_notes_total"] = len(parse_notes)

    name_to_id = {}
    data = get_container().app_data_store.get()
    for r in data.recetas:
        if r.activo:
            name_to_id[_norm(r.nombre)] = r.id

    for fecha, items in sorted(by_day.items()):
        clave = f"{CLAVE_PREFIX}-desayuno-{fecha.isoformat()}"
        data = get_container().app_data_store.get()
        if any(
            getattr(x, "clave_idempotencia", None) == clave and not getattr(x, "anulado", False)
            for x in data.desayunos
        ):
            report.setdefault("skipped_idempotent", []).append(clave)
            continue

        des.limpiar_cesta()
        data = get_container().app_data_store.get()
        # refresh tostada ids
        for r in data.recetas:
            if r.activo:
                name_to_id[_norm(r.nombre)] = r.id

        huespedes = max(1, len(items))
        for it in items:
            for rec_nombre, porc, extras in it["recipes"]:
                rid = name_to_id.get(_norm(rec_nombre))
                if not rid and _norm(rec_nombre).startswith("TOSTADA"):
                    tr = receta_tostada_del_dia(fecha)
                    rid = tr.id if tr else None
                if not rid:
                    report.setdefault("errors", []).append(f"{clave}: falta receta «{rec_nombre}»")
                    continue
                for key, qty, unit in extras:
                    pid = CATALOG[key]
                    cant = _to_nativa(data, pid, qty, unit)
                    rr = des.anadir_mod_pendiente_receta(pid, float(cant))
                    if not rr.ok:
                        report.setdefault("errors", []).append(f"{clave} mod {key}: {rr.mensaje}")
                rr = des.anadir_receta_a_cesta(rid, float(porc))
                if not rr.ok:
                    report.setdefault("errors", []).append(f"{clave} receta {rec_nombre}: {rr.mensaje}")
            for key, qty, unit in it["products"]:
                pid = CATALOG.get(key, key if key.startswith(("p", "b")) else None)
                if not pid:
                    report.setdefault("errors", []).append(f"{clave}: producto desconocido {key}")
                    continue
                cant = _to_nativa(data, pid, qty, unit)
                rr = des.anadir_a_cesta(pid, float(cant))
                if not rr.ok:
                    report.setdefault("errors", []).append(f"{clave} prod {key}: {rr.mensaje}")

        if des.cesta_vacia():
            report.setdefault("errors", []).append(f"{clave}: cesta vacía")
            continue

        topup_for_service(des, fecha, report)
        topup_for_service(des, fecha, report)
        res = des.registrar_desayuno(
            fecha,
            huespedes,
            clave_idempotencia=clave,
            observaciones=f"Import Excel desayuno agosto 2026 ({len(items)} líneas)",
        )
        if res.ok:
            data = get_container().app_data_store.get()
            reg = next((d for d in data.desayunos if getattr(d, "clave_idempotencia", None) == clave), None)
            coste = sum(getattr(l, "coste", 0) or 0 for l in (getattr(reg, "lineas", None) or [])) if reg else None
            report.setdefault("ok", []).append(
                {
                    "clave": clave,
                    "ref": getattr(reg, "id", None),
                    "coste": coste,
                    "tipo": "desayuno",
                    "fecha": fecha.isoformat(),
                    "lineas_fuente": len(items),
                    "msg": res.mensaje,
                }
            )
        else:
            detail = getattr(res, "detalle_stock", None) or []
            report.setdefault("errors", []).append(f"{clave}: {res.mensaje} | {detail[:8]}")
        des.limpiar_cesta()


def main() -> int:
    if not HOTEL.exists():
        print("No hotel data:", HOTEL)
        return 1
    if not TPV_JSON.exists():
        print("Falta", TPV_JSON)
        return 1

    _boot()
    report: dict = {
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "ok": [],
        "errors": [],
        "reposiciones": [],
        "skipped_idempotent": [],
    }

    print("== Recetas / flags ==")
    ensure_recipes_and_flags(report)

    print("== TPV comida/bebidas ==")
    _boot()
    import_tpv(report)

    print("== Desayunos ==")
    _boot()
    import_desayunos(report)

    report["finished_at"] = datetime.now().isoformat(timespec="seconds")
    unmapped_names = sorted({_norm(x.get("nombre", "")) for x in report.get("tpv_unmapped", [])})
    report["summary"] = {
        "ok": len(report.get("ok", [])),
        "errors": len(report.get("errors", [])),
        "unmapped_lines": len(report.get("tpv_unmapped", [])),
        "unmapped_names": unmapped_names,
        "pending_coctel_dia": len(report.get("tpv_pending_coctel_dia", [])),
        "mermas_hielo": len(report.get("mermas_hielo", [])),
        "reposiciones": len(report.get("reposiciones", [])),
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    if report["errors"]:
        print("ERRORS:", *report["errors"][:20], sep="\n  ")
    print("Report:", REPORT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
