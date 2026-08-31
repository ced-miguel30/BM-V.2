"""Genera / actualiza docs/plantillas/registro_desayuno_operativo_ACTUALIZADA.xlsx.

Hojas: Instrucciones, Registro, RegistroBebidasDesayuno, RegistroComida, RegistroCena,
ConfigBuffet, ConsumoBuffet, Catalogo.

Si el archivo de salida ya existe y tiene filas en hojas de registro, se PRESERVAN.
(no se borran datos ya apuntados). Solo se refrescan Instrucciones, Catalogo
y validaciones.

Uso:
  .\\.venv\\Scripts\\python.exe scripts\\build_plantilla_desayuno_excel.py
  .\\.venv\\Scripts\\python.exe scripts\\build_plantilla_desayuno_excel.py --path RUTA\\datos_hotel.json
  .\\.venv\\Scripts\\python.exe scripts\\build_plantilla_desayuno_excel.py --out RUTA.xlsx
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.bootstrap import configure_for_flet, get_container, reset_container
from app.core.models import CategoriaReceta
from app.core.services import desayuno_service as des
from app.core.services.buffet_config_service import ensure_config_buffet
from app.core.services.receta_service import ETIQUETA_TOSTADA_DEL_DIA
from app.core.services.text_search import normalizar_texto

HOTEL_DEFAULT = Path(os.environ["LOCALAPPDATA"]) / "BM-V2-local" / "data" / "datos_hotel.json"
OUT_DEFAULT = ROOT / "docs" / "plantillas" / "registro_desayuno_operativo_ACTUALIZADA.xlsx"

HEADERS_SERVICIO = ["Fecha", "Tipo", "Nombre", "Cantidad ↑↓", "Notas", "Importado"]

HEADERS_CONFIG_BUFFET = [
    "Seccion", "Orden", "Concepto", "ProductoId", "Unidad", "CantDefecto",
    "Tipo", "ProductoBote", "RecetaId", "Activo",
]

HEADERS_CONSUMO_BUFFET = [
    "Fecha", "Seccion", "Concepto", "Cantidad", "Motivo",
    "Notas", "Importado",
]

MOTIVOS_BUFFET_EXCEL = "Consumo,Merma,Expiración,Limpieza"

FILL_BUFFET = PatternFill("solid", fgColor="FFF2CC")

HEADERS = [
    "Fecha",
    "Huespedes",
    "Tipo",
    "Nombre",
    "Cantidad ↑↓",
    "Extra1",
    "Cant1 ↑↓",
    "Extra2",
    "Cant2 ↑↓",
    "Extra3",
    "Cant3 ↑↓",
    "Extra4",
    "Cant4 ↑↓",
    "Omitir1",
    "Omitir2",
    "Notas",
    "Importado",
]

HUEVOS_OFRECIDOS = [
    "Huevo frito",
    "Huevo pochado",
    "Huevo cocido",
    "Huevos revueltos",
    "Sin huevo",
]

PANES_OFRECIDOS = [
    "Tostada",  # molde común / blanco
    "Tostada integral",
    "Tostada sin gluten",
    "Pan blanco",
    "Pan integral",
    "Pan sin gluten",
    "Croissant sin gluten",
    "Magdalena sin gluten",
    "Muesli sin gluten",
    "Cacao sin gluten",
    "Sin tostada",
]

# Siempre en catálogo bebidas desayuno (p. ej. Cava Roger de Flor).
BEBIDAS_DESAYUNO_SIEMPRE = (
    "Botella Roger de Flor",
    "Cava Roger de Flor",
)

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_CAT = PatternFill("solid", fgColor="E2EFDA")
FILL_HUEVO = PatternFill("solid", fgColor="FCE4D6")
FILL_PAN = PatternFill("solid", fgColor="DDEBF7")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _boot(path: Path) -> None:
    reset_container()
    configure_for_flet(data_path=str(path))


def _catalogo_recetas_servicio(data, categorias: list[CategoriaReceta]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    cats = {c.value for c in categorias}
    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        cat = r.categoria.value if hasattr(r.categoria, "value") else str(r.categoria)
        if cat == CategoriaReceta.BEBIDAS.value:
            if CategoriaReceta.BEBIDAS not in categorias and not des.es_receta_bebida_desayuno(r.nombre):
                continue
        elif cat not in cats:
            continue
        key = normalizar_texto(r.nombre)
        if key in seen:
            continue
        seen.add(key)
        names.append(r.nombre)
    return sorted(names, key=lambda s: normalizar_texto(s))


def _count_col_values(ws, col_letter: str) -> int:
    """Filas con valor en col (desde fila 2). Mínimo 1 para rangos DV."""
    n = 0
    for cell in ws[col_letter]:
        if cell.row == 1:
            continue
        if cell.value is None or str(cell.value).strip() == "":
            continue
        n += 1
    return max(n, 1)


def aplicar_validaciones_plantilla(wb) -> None:
    """Reaplica desplegables Nombre/Tipo/Extras/Buffet desde Catalogo/ConfigBuffet.

    openpyxl (y el marcado Importado del import) puede perder validaciones de
    lista con referencia a otra hoja; llamar siempre antes de ``wb.save``.
    """
    if "Catalogo" not in wb.sheetnames:
        return
    ws_c = wb["Catalogo"]
    n_all = _count_col_values(ws_c, "C")
    n_combo = _count_col_values(ws_c, "E")
    n_comida = _count_col_values(ws_c, "G")
    n_cena = _count_col_values(ws_c, "H")
    n_bebidas = _count_col_values(ws_c, "I")
    n_config = 1
    if "ConfigBuffet" in wb.sheetnames:
        n_config = _count_col_values(wb["ConfigBuffet"], "C")

    def _list_dv(formula: str) -> DataValidation:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.showErrorMessage = False  # permite escribir aunque no coincida exacto
        return dv

    if "Registro" in wb.sheetnames:
        ws = wb["Registro"]
        ws.data_validations.dataValidation = []
        dv_tipo = _list_dv('"Receta,Extra,Producto"')
        ws.add_data_validation(dv_tipo)
        dv_tipo.add("C2:C500")
        dv_nombre = _list_dv(f"Catalogo!$C$2:$C${n_all + 1}")
        ws.add_data_validation(dv_nombre)
        dv_nombre.add("D2:D500")
        dv_extra = _list_dv(f"Catalogo!$E$2:$E${n_combo + 1}")
        ws.add_data_validation(dv_extra)
        for col in ("F", "H", "J", "L", "N", "O"):
            dv_extra.add(f"{col}2:{col}500")
        dv_cant = DataValidation(
            type="whole", operator="between", formula1="1", formula2="30", allow_blank=True,
        )
        ws.add_data_validation(dv_cant)
        for col in ("E", "G", "I", "K", "M"):
            dv_cant.add(f"{col}2:{col}500")

    for hoja, formula in (
        ("RegistroComida", f"Catalogo!$G$2:$G${n_comida + 1}"),
        ("RegistroCena", f"Catalogo!$H$2:$H${n_cena + 1}"),
        ("RegistroBebidasDesayuno", f"Catalogo!$I$2:$I${n_bebidas + 1}"),
    ):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        ws.data_validations.dataValidation = []
        dv_tipo = _list_dv('"Receta,Extra,Producto"')
        ws.add_data_validation(dv_tipo)
        dv_tipo.add("B2:B500")
        dv_nombre = _list_dv(formula)
        ws.add_data_validation(dv_nombre)
        dv_nombre.add("C2:C500")
        dv_cant = DataValidation(
            type="whole", operator="between", formula1="1", formula2="30", allow_blank=True,
        )
        ws.add_data_validation(dv_cant)
        dv_cant.add("D2:D500")

    if "ConsumoBuffet" in wb.sheetnames:
        ws = wb["ConsumoBuffet"]
        ws.data_validations.dataValidation = []
        dv_motivo = _list_dv(f'"{MOTIVOS_BUFFET_EXCEL}"')
        ws.add_data_validation(dv_motivo)
        dv_motivo.add("E2:E500")
        dv_concepto = _list_dv(f"ConfigBuffet!$C$2:$C${n_config + 1}")
        ws.add_data_validation(dv_concepto)
        dv_concepto.add("C2:C500")


def _escribir_hoja_registro_simple(
    ws,
    headers: list[str],
    preservado: tuple[list, list[tuple]] | None,
    *,
    lista_nombre_col: str,
    lista_ref: str,
    filas_vacias: int = 120,
) -> int:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    filas_datos = 0
    start_empty = 2
    if preservado:
        old_headers, old_rows = preservado
        for r_i, old in enumerate(old_rows, start=2):
            mapped = _mapear_fila(old_headers, old, headers)
            for c_i, val in enumerate(mapped, start=1):
                ws.cell(r_i, c_i, val).border = THIN
            filas_datos += 1
        start_empty = 2 + filas_datos
    for r_i in range(start_empty, start_empty + filas_vacias):
        for c_i in range(1, len(headers) + 1):
            ws.cell(r_i, c_i).border = THIN
    ws.data_validations.dataValidation = []
    dv_tipo = DataValidation(type="list", formula1='"Receta,Extra,Producto"', allow_blank=True)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add("B2:B500")
    nombre_idx = headers.index("Nombre") + 1
    col = get_column_letter(nombre_idx)
    dv_nombre = DataValidation(type="list", formula1=lista_ref, allow_blank=True)
    ws.add_data_validation(dv_nombre)
    dv_nombre.add(f"{col}2:{col}500")
    cant_idx = next(i for i, h in enumerate(headers) if h.startswith("Cantidad"))
    cant_col = get_column_letter(cant_idx + 1)
    dv_cant = DataValidation(type="whole", operator="between", formula1="1", formula2="30", allow_blank=True)
    ws.add_data_validation(dv_cant)
    dv_cant.add(f"{cant_col}2:{cant_col}500")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return filas_datos


def _escribir_config_buffet(ws, data) -> int:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    for col, h in enumerate(HEADERS_CONFIG_BUFFET, start=1):
        cell = ws.cell(1, col, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
    ensure_config_buffet(data)
    configs = sorted(
        [c for c in data.config_buffet if c.activo],
        key=lambda c: (c.seccion, c.orden, c.label),
    )
    for i, cfg in enumerate(configs, start=2):
        ws.cell(i, 1, cfg.seccion)
        ws.cell(i, 2, cfg.orden)
        ws.cell(i, 3, cfg.label)
        ws.cell(i, 4, cfg.producto_id)
        ws.cell(i, 5, cfg.unidad)
        ws.cell(i, 6, cfg.cantidad_defecto)
        ws.cell(i, 7, cfg.tipo_linea)
        ws.cell(i, 8, cfg.producto_bote_id or "")
        ws.cell(i, 9, cfg.receta_id or "")
        ws.cell(i, 10, "Si" if cfg.activo else "No")
        for c in range(1, 11):
            ws.cell(i, c).fill = FILL_BUFFET
    for col, w in zip("ABCDEFGHIJ", [14, 8, 32, 12, 10, 12, 16, 12, 12, 8]):
        ws.column_dimensions[col].width = w
    return max(len(configs), 1)


def _escribir_consumo_buffet(ws, data, n_config: int, preservado) -> int:
    for col, h in enumerate(HEADERS_CONSUMO_BUFFET, start=1):
        cell = ws.cell(1, col, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
    filas_datos = 0
    start_empty = 2
    if preservado:
        old_headers, old_rows = preservado
        for r_i, old in enumerate(old_rows, start=2):
            mapped = _mapear_fila(old_headers, old, HEADERS_CONSUMO_BUFFET)
            for c_i, val in enumerate(mapped, start=1):
                ws.cell(r_i, c_i, val).border = THIN
            filas_datos += 1
        start_empty = 2 + filas_datos
    max_row = max(start_empty + 80, 150)
    for r_i in range(2, max_row):
        for c_i in range(1, len(HEADERS_CONSUMO_BUFFET) + 1):
            ws.cell(r_i, c_i).border = THIN
    ws.data_validations.dataValidation = []
    dv_motivo = DataValidation(type="list", formula1=f'"{MOTIVOS_BUFFET_EXCEL}"', allow_blank=True)
    ws.add_data_validation(dv_motivo)
    dv_motivo.add("E2:E500")
    dv_concepto = DataValidation(
        type="list",
        formula1=f"ConfigBuffet!$C$2:$C${n_config + 1}",
        allow_blank=True,
    )
    ws.add_data_validation(dv_concepto)
    dv_concepto.add("C2:C500")
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFG", [12, 14, 36, 10, 14, 22, 12]):
        ws.column_dimensions[col].width = w
    return filas_datos


def _catalogo_recetas(data) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    names.append(ETIQUETA_TOSTADA_DEL_DIA)
    seen.add(normalizar_texto(ETIQUETA_TOSTADA_DEL_DIA))
    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        cat = r.categoria.value if hasattr(r.categoria, "value") else str(r.categoria)
        if cat == CategoriaReceta.DESAYUNO.value:
            pass
        elif cat == CategoriaReceta.BEBIDAS.value:
            if not des.es_receta_bebida_desayuno(r.nombre):
                continue
        else:
            continue
        key = normalizar_texto(r.nombre)
        if key in seen:
            continue
        seen.add(key)
        names.append(r.nombre)
    return sorted(names, key=lambda s: normalizar_texto(s))


def _catalogo_bebidas_desayuno(data) -> list[str]:
    """Recetas bebida de desayuno + botellas sueltas (p. ej. Cava Roger de Flor)."""
    names: list[str] = []
    seen: set[str] = set()

    def add(name: str) -> None:
        key = normalizar_texto(name)
        if not key or key in seen:
            return
        seen.add(key)
        names.append(name)

    for r in data.recetas:
        if not getattr(r, "activo", True):
            continue
        cat = r.categoria.value if hasattr(r.categoria, "value") else str(r.categoria)
        if cat == CategoriaReceta.BEBIDAS.value and des.es_receta_bebida_desayuno(r.nombre):
            add(r.nombre)
        elif "roger de flor" in normalizar_texto(r.nombre):
            add(r.nombre)

    for p in data.productos:
        if not getattr(p, "activo", True) or not getattr(p, "es_bebida", False):
            continue
        servicios = [s for s in (getattr(p, "servicios_disponibles", None) or []) if isinstance(s, str)]
        if not servicios or "desayuno" in servicios:
            add(p.nombre)

    for lab in BEBIDAS_DESAYUNO_SIEMPRE:
        add(lab)

    for e in des.bebidas_frias_rapidas_desayuno():
        add(str(e.get("label") or ""))

    return sorted(names, key=lambda s: normalizar_texto(s))


def _catalogo_extras() -> list[str]:
    labels = [e["label"] for e in des.extras_rapidos_desayuno()]
    labels += [e["label"] for e in des.leches_rapidas_desayuno()]
    labels += [e["label"] for e in des.bebidas_frias_rapidas_desayuno()]
    out: list[str] = []
    seen: set[str] = set()
    for lab in labels:
        k = normalizar_texto(lab)
        if k in seen:
            continue
        seen.add(k)
        out.append(lab)
    return out


def _write_instrucciones(ws, hotel: Path, n_rec: int, n_ext: int) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "Registro operativo de desayuno (BM)"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    lines = [
        "",
        "CÓMO USAR",
        "1. Vaya a la hoja Registro.",
        "2. En cada línea ponga la Fecha (AAAA-MM-DD). Misma fecha = UN solo desayuno en BM.",
        "3. Huéspedes (por línea, se SUMAN en el día):",
        "   • 1 = nuevo comensal (cuenta 1 persona).",
        "   • 0 = mismo comensal que pide otro plato (NO suma persona).",
        "   • Vacío = igual que 0 (no suma).",
        "   Ejemplo: comensal pide inglés + café → fila inglés Huespedes=1, fila café=0.",
        "   Total del día = suma de la columna (solo los 1).",
        "4. Tipo: Receta (plato/bebida), Extra o Producto. Si lo dejas vacío y hay Nombre, se asume Receta.",
        "5. Nombre: elija del desplegable (lista en hoja Catalogo).",
        "6. Cantidad ↑↓: raciones o unidades de ESA línea (1–30).",
        "7. Extra1…Extra4: extras sobre una Receta (bacon, aguacate, huevo, pan…).",
        "8. Omitir1/Omitir2: quitar un ingrediente (Bacon, Sin huevo, Sin tostada…).",
        "",
        "IMPORTANTE",
        "• No ponga 1 en todas las filas: si lo hace, contará 1 huésped por plato.",
        "• Cantidad es raciones del plato, no el número de huéspedes.",
        "• Tras importar, puede borrar las filas del Excel; en BM ya queda el registro.",
        "",
        "HUEVOS (Desayuno inglés)",
        "• La ficha «Desayuno ingles» incluye Huevo frito por defecto.",
        "• Para otro tipo: Extra1 = Huevo pochado / Huevos revueltos / Huevo cocido",
        "  (sustituye al frito; no sume un segundo huevo).",
        "• Para sin huevo: Omitir1 = Sin huevo (o Huevo frito).",
        "• Huevo suelto: Tipo=Receta y Nombre=Huevo frito / pochado / …",
        "",
        "PAN (tostadas / sándwiches)",
        "• Por defecto: molde común (blanco) = Tostada / Pan blanco.",
        "• Integral: Extra1 = Tostada integral (o Pan integral).",
        "• Sin gluten: Extra1 = Tostada/Pan, Croissant, Magdalena, Muesli o Cacao sin gluten.",
        "• Sin pan: Omitir1 = Sin tostada.",
        "",
        "9. Guarde y ejecute 2_importar_a_bm.cmd (mejor --dry-run antes).",
        "10. La columna Importado la rellena el script (no la edite).",
        "",
        "COMIDA / CENA (hojas RegistroComida, RegistroCena)",
        "• Misma Fecha = un solo registro del servicio en BM.",
        "• Tipo: Receta/Extra/Producto (vacío = Receta). Misma fecha: basta en la 1ª fila del día.",
        "",
        "BEBIDAS DESAYUNO (hoja RegistroBebidasDesayuno)",
        "• Cafés, tés, Cola Cao, Cava, agua/soda/refrescos (catálogo columna I).",
        "• Agua/refresco: Tipo = Producto o Extra; Nombre = etiqueta (ej. Agua sin gas PET, Coca-Cola lata).",
        "  1 unidad = 1 botella/lata descontada del inventario.",
        "• Misma Fecha = un registro de bebidas en BM.",
        "• Tipo: Receta (café, cava…) o Producto/Extra (botella suelta). Cantidad: raciones/unidades.",
        "",
        "CONSUMO BUFFET (hoja ConsumoBuffet)",
        "• Concepto: elija de ConfigBuffet (editable). Motivo: Consumo / Merma / Expiración / Limpieza.",
        "• Zumo naranja natural 1L: Concepto = «Zumo naranja natural 1L», Cantidad = litros.",
        "  BM usa la receta (kg NARANJA ZUMO b06 por litro; editable en Recetas).",
        "• Zumo de bote/brik: Concepto = «Zumo naranja brik 1L» (producto ZUMO NARANJA 1L BRIK).",
        "• No hay columnas de naranjas ni zumo bote: elija el concepto correcto en el desplegable.",
        "",
        "Al regenerar la plantilla NO se borran filas ya rellenadas en hojas de registro.",
    ]
    for i, line in enumerate(lines, start=2):
        ws[f"A{i}"] = line
        if line.startswith("CÓMO") or line.startswith("HUEVOS") or line.startswith("PAN") or line.startswith("IMPORTANTE"):
            ws[f"A{i}"].font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 110
    ws["A40"] = f"Catalogo generado desde: {hotel}"
    ws["A41"] = f"Recetas desayuno: {n_rec} | Extras: {n_ext}"
    ws["A40"].font = Font(italic=True, size=10, color="666666")
    ws["A41"].font = Font(italic=True, size=10, color="666666")


def _leer_hoja_existente(path: Path, nombre: str) -> tuple[list, list[tuple]] | None:
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, data_only=False)
    except Exception:
        return None
    if nombre not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[nombre]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(x not in (None, "") for x in row):
            rows.append(tuple(row))
    wb.close()
    if not rows:
        return None
    return headers, rows


def _leer_registro_existente(path: Path) -> tuple[list, list[tuple]] | None:
    return _leer_hoja_existente(path, "Registro")


def _mapear_fila(old_headers: list, row: tuple, new_headers: list) -> list:
    """Copia valores por nombre de columna (tolerante a plantillas antiguas)."""
    idx = {str(h).strip(): i for i, h in enumerate(old_headers) if h is not None}
    out: list = []
    for h in new_headers:
        key = h
        if key.startswith("Cantidad"):
            # alias
            i = next((idx[k] for k in idx if k.startswith("Cantidad")), None)
        elif key.startswith("Cant") and "↑" in key:
            n = key[4]
            i = next(
                (idx[k] for k in idx if k.startswith(f"Cant{n}")),
                None,
            )
        else:
            i = idx.get(key)
        if i is None or i >= len(row):
            out.append(None)
        else:
            out.append(row[i])
    return out


def _escribir_catalogo(ws_c, recetas: list[str], extras: list[str], data) -> tuple[int, int, int, int, int, int]:
    # Limpiar
    if ws_c.max_row:
        ws_c.delete_rows(1, ws_c.max_row)
    ws_c["A1"] = "Recetas_desayuno"
    ws_c["B1"] = "Extras_rapidos"
    ws_c["C1"] = "Nombres_todos"
    ws_c["D1"] = "Huevos_ofrecidos"
    ws_c["F1"] = "Panes_ofrecidos"
    ws_c["G1"] = "Recetas_comida"
    ws_c["H1"] = "Recetas_cena"
    ws_c["I1"] = "Bebidas_desayuno"
    for col in ("A", "B", "C", "D", "F", "G", "H", "I"):
        ws_c[f"{col}1"].font = FONT_HEADER
        ws_c[f"{col}1"].fill = FILL_HEADER
    rec_comida = _catalogo_recetas_servicio(data, [CategoriaReceta.COMIDA, CategoriaReceta.BEBIDAS])
    rec_cena = _catalogo_recetas_servicio(data, [CategoriaReceta.CENA, CategoriaReceta.BEBIDAS])
    rec_bebidas = _catalogo_bebidas_desayuno(data)
    for i, name in enumerate(recetas, start=2):
        ws_c[f"A{i}"] = name
        ws_c[f"A{i}"].fill = FILL_CAT
    for i, lab in enumerate(extras, start=2):
        ws_c[f"B{i}"] = lab
        ws_c[f"B{i}"].fill = FILL_CAT
    todos = list(dict.fromkeys(recetas + extras))
    for i, name in enumerate(todos, start=2):
        ws_c[f"C{i}"] = name
    for i, lab in enumerate(HUEVOS_OFRECIDOS, start=2):
        ws_c[f"D{i}"] = lab
        ws_c[f"D{i}"].fill = FILL_HUEVO
    for i, lab in enumerate(PANES_OFRECIDOS, start=2):
        ws_c[f"F{i}"] = lab
        ws_c[f"F{i}"].fill = FILL_PAN
    for i, name in enumerate(rec_comida, start=2):
        ws_c[f"G{i}"] = name
        ws_c[f"G{i}"].fill = FILL_CAT
    for i, name in enumerate(rec_cena, start=2):
        ws_c[f"H{i}"] = name
        ws_c[f"H{i}"].fill = FILL_CAT
    for i, name in enumerate(rec_bebidas, start=2):
        ws_c[f"I{i}"] = name
        ws_c[f"I{i}"].fill = FILL_CAT
    ws_c.column_dimensions["A"].width = 42
    ws_c.column_dimensions["B"].width = 28
    ws_c.column_dimensions["C"].width = 42
    ws_c.column_dimensions["D"].width = 22
    ws_c.column_dimensions["F"].width = 22
    ws_c.column_dimensions["G"].width = 42
    ws_c.column_dimensions["H"].width = 42
    ws_c.column_dimensions["I"].width = 42
    return (
        max(len(recetas), 1),
        max(len(extras), 1),
        max(len(todos), 1),
        max(len(rec_comida), 1),
        max(len(rec_cena), 1),
        max(len(rec_bebidas), 1),
    )


def _rellenar_lista_extra_omitir(ws_c, extras: list[str]) -> int:
    """Columna E = extras ∪ huevos ∪ panes (desplegables Extra/Omitir)."""
    ws_c["E1"] = "Extra_u_Omitir"
    ws_c["E1"].font = FONT_HEADER
    ws_c["E1"].fill = FILL_HEADER
    combo = list(dict.fromkeys(extras + HUEVOS_OFRECIDOS + PANES_OFRECIDOS))
    for lab in HUEVOS_OFRECIDOS + PANES_OFRECIDOS:
        if lab not in combo:
            combo.append(lab)
    for i, lab in enumerate(combo, start=2):
        ws_c[f"E{i}"] = lab
        if lab in HUEVOS_OFRECIDOS:
            ws_c[f"E{i}"].fill = FILL_HUEVO
        elif lab in PANES_OFRECIDOS:
            ws_c[f"E{i}"].fill = FILL_PAN
    ws_c.column_dimensions["E"].width = 28
    return max(len(combo), 1)


def build(out: Path, hotel: Path) -> Path:
    assert hotel.exists(), f"No existe {hotel}"
    _boot(hotel)
    data = get_container().app_data_store.get()
    recetas = _catalogo_recetas(data)
    extras = _catalogo_extras()
    if ensure_config_buffet(data):
        # Persistir receta zumo natural + conceptos buffet en datos_hotel
        from app.core.services.desayuno_service import _ctx as _des_ctx

        _des_ctx().uow.commit(data)
        data = get_container().app_data_store.get()
        ensure_config_buffet(data)

    preservado = _leer_registro_existente(out)
    preservado_comida = _leer_hoja_existente(out, "RegistroComida")
    preservado_cena = _leer_hoja_existente(out, "RegistroCena")
    preservado_bebidas = _leer_hoja_existente(out, "RegistroBebidasDesayuno")
    preservado_buffet = _leer_hoja_existente(out, "ConsumoBuffet")

    wb = Workbook()
    ws_i = wb.active
    ws_i.title = "Instrucciones"
    _write_instrucciones(ws_i, hotel, len(recetas), len(extras))

    ws_c = wb.create_sheet("Catalogo")
    _n_rec, n_ext, n_all, n_comida, n_cena, n_bebidas = _escribir_catalogo(ws_c, recetas, extras, data)
    n_combo = _rellenar_lista_extra_omitir(ws_c, extras)

    ws_cfg = wb.create_sheet("ConfigBuffet")
    n_config = _escribir_config_buffet(ws_cfg, data)

    ws_buffet = wb.create_sheet("ConsumoBuffet")
    filas_buffet = _escribir_consumo_buffet(
        ws_buffet, data, n_config, preservado_buffet,
    )

    ws_comida = wb.create_sheet("RegistroComida")
    filas_comida = _escribir_hoja_registro_simple(
        ws_comida,
        HEADERS_SERVICIO,
        preservado_comida,
        lista_nombre_col="Nombre",
        lista_ref=f"Catalogo!$G$2:$G${n_comida + 1}",
    )

    ws_cena = wb.create_sheet("RegistroCena")
    filas_cena = _escribir_hoja_registro_simple(
        ws_cena,
        HEADERS_SERVICIO,
        preservado_cena,
        lista_nombre_col="Nombre",
        lista_ref=f"Catalogo!$H$2:$H${n_cena + 1}",
    )

    ws_bebidas = wb.create_sheet("RegistroBebidasDesayuno")
    filas_bebidas = _escribir_hoja_registro_simple(
        ws_bebidas,
        HEADERS_SERVICIO,
        preservado_bebidas,
        lista_nombre_col="Nombre",
        lista_ref=f"Catalogo!$I$2:$I${n_bebidas + 1}",
    )

    ws_r = wb.create_sheet("Registro", 1)
    for col, h in enumerate(HEADERS, start=1):
        cell = ws_r.cell(1, col, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws_r.row_dimensions[1].height = 32

    widths = [
        12, 11, 10, 28, 12,
        16, 9, 16, 9, 16, 9, 16, 9,
        16, 16, 22, 12,
    ]
    for i, w in enumerate(widths, start=1):
        ws_r.column_dimensions[get_column_letter(i)].width = w

    n_cols = len(HEADERS)
    filas_datos = 0
    if preservado:
        old_headers, old_rows = preservado
        for r_i, old in enumerate(old_rows, start=2):
            mapped = _mapear_fila(old_headers, old, HEADERS)
            for c_i, val in enumerate(mapped, start=1):
                cell = ws_r.cell(r_i, c_i, val)
                cell.border = THIN
            filas_datos += 1
        # Filas vacías extra
        start_empty = 2 + filas_datos
    else:
        start_empty = 2

    for r_i in range(start_empty, max(start_empty + 50, 205)):
        for c_i in range(1, n_cols + 1):
            ws_r.cell(r_i, c_i).border = THIN

    # Validaciones con lista Extra_u_Omitir (col E)
    ws_r.data_validations.dataValidation = []
    dv_tipo = DataValidation(
        type="list", formula1='"Receta,Extra,Producto"', allow_blank=True
    )
    ws_r.add_data_validation(dv_tipo)
    dv_tipo.add("C2:C500")

    dv_nombre = DataValidation(
        type="list",
        formula1=f"Catalogo!$C$2:$C${n_all + 1}",
        allow_blank=True,
    )
    ws_r.add_data_validation(dv_nombre)
    dv_nombre.add("D2:D500")

    dv_extra = DataValidation(
        type="list",
        formula1=f"Catalogo!$E$2:$E${n_combo + 1}",
        allow_blank=True,
    )
    ws_r.add_data_validation(dv_extra)
    for col in ("F", "H", "J", "L", "N", "O"):
        dv_extra.add(f"{col}2:{col}500")

    dv_cant = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="30",
        allow_blank=True,
    )
    ws_r.add_data_validation(dv_cant)
    for col in ("E", "G", "I", "K", "M"):
        dv_cant.add(f"{col}2:{col}500")

    ws_r.freeze_panes = "A2"
    ws_r.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    note = f"Registro preservado: {filas_datos} fila(s)." if filas_datos else "Registro vacío (plantilla nueva)."
    ws_i["A48"] = note
    ws_i["A48"].font = Font(italic=True, size=10, color="1F4E79")
    ws_i["A49"] = (
        f"Comida: {filas_comida} | Cena: {filas_cena} | "
        f"Bebidas: {filas_bebidas} | Buffet: {filas_buffet} fila(s) preservadas."
    )
    ws_i["A49"].font = Font(italic=True, size=10, color="666666")

    out.parent.mkdir(parents=True, exist_ok=True)
    aplicar_validaciones_plantilla(wb)
    try:
        wb.save(out)
        return out
    except PermissionError:
        for suffix in ("_ACTUALIZADA", "_NUEVA", "_PAN"):
            alt = out.with_name(out.stem + f"{suffix}.xlsx")
            try:
                wb.save(alt)
                print(f"AVISO: {out.name} está abierto; guardado en {alt.name}")
                return alt
            except PermissionError:
                continue
        raise


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", type=Path, default=HOTEL_DEFAULT, help="datos_hotel.json")
    parser.add_argument("--out", type=Path, default=OUT_DEFAULT, help="xlsx de salida")
    args = parser.parse_args()
    dest = build(args.out, args.path)
    print("Plantilla:", dest)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
