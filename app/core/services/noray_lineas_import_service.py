"""Importación de líneas de compra exportadas desde Noray / Business Central.

Formatos soportados (Excel «Líneas»):
- Variante corta: Nº / N.º referencia art., Descripción, Cód. almacén,
  Cantidad, Cód. unidad medida, Coste unit. directo, % IVA, fechas.
- Variante larga: Nº, Descripción, Cód. almacén, Cantidad, unidad,
  Coste unit. directo excl. IVA, Grupo contable IVA, fechas, totales.

No muta AppData: solo parsea y propone matching de producto/ubicación.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.models import AppData
from app.core.services.money import as_decimal
from app.core.services.text_search import coincide_busqueda, normalizar_texto


@dataclass(frozen=True)
class NorayLineaParseada:
    codigo_articulo: str
    descripcion: str
    almacen: str
    cantidad: Decimal
    unidad: str
    coste_unitario: Decimal
    igic_pct: Decimal
    fecha: date | None
    fila_excel: int


# Estados de emparejamiento (nombre primero; código verifica).
MATCH_OK = "ok"
MATCH_REVISAR = "revisar"  # emparejado pero nombre/código no totalmente alineados
MATCH_CONFLICTO = "conflicto"  # nombre y código apuntan a productos distintos
MATCH_SIN = "sin_match"  # sin producto; ofrecer alta
MATCH_AMBIGUO = "ambiguo"  # varios candidatos por nombre


@dataclass
class NorayLineaMatch:
    linea: NorayLineaParseada
    producto_id: str | None = None
    producto_nombre: str | None = None
    producto_codigo: str | None = None
    ubicacion_sugerida_id: str | None = None
    ubicacion_sugerida_etiqueta: str | None = None
    estado: str = MATCH_SIN
    avisos: list[str] = field(default_factory=list)

    @property
    def match_ok(self) -> bool:
        """Listo para confirmar sin intervención (nombre + código alineados)."""
        return self.estado == MATCH_OK and bool(self.producto_id)

    @property
    def requiere_accion(self) -> bool:
        return self.estado in (MATCH_SIN, MATCH_CONFLICTO, MATCH_AMBIGUO) or (
            self.estado == MATCH_REVISAR
        )


@dataclass(frozen=True)
class ResultadoParseNoray:
    ok: bool
    mensaje: str
    lineas: tuple[NorayLineaParseada, ...] = ()
    fecha_documento: date | None = None
    referencia_sugerida: str = ""


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _cell_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _cell_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _cell_dec(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return as_decimal(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return as_decimal(s)
    except Exception:
        return Decimal("0")


def _norm_header(h: Any) -> str:
    return normalizar_texto(_cell_str(h))


def _map_headers(row: tuple[Any, ...]) -> dict[str, int]:
    """Índices por sinónimos de cabecera Noray."""
    aliases = {
        "codigo": (
            "n.",
            "n",
            "no",
            "n.o referencia art.",
            "n referencia art",
            "n. referencia art.",
            "codigo",
            "cod. articulo",
        ),
        "descripcion": ("descripcion", "description"),
        "almacen": ("cod. almacen", "cod almacen", "almacen"),
        "cantidad": ("cantidad", "qty"),
        "unidad": (
            "cod. unidad medida",
            "cod unidad medida",
            "unidad",
            "ud",
        ),
        "coste": (
            "coste unit. directo",
            "coste unit directo",
            "coste unit. directo excl. iva",
            "coste unit directo excl. iva",
            "precio",
        ),
        "iva": ("% iva", "iva", "grupo contable iva prod."),
        "fecha": (
            "fecha recepcion esperada",
            "fecha recep. esperada",
            "fecha recep. planificada",
            "fecha pedido",
            "fecha recep. prometida",
        ),
    }
    idx: dict[str, int] = {}
    headers = [_norm_header(c) for c in row]
    for key, syns in aliases.items():
        for i, h in enumerate(headers):
            if not h:
                continue
            for s in syns:
                sn = normalizar_texto(s)
                if h == sn or h.startswith(sn) or sn in h:
                    if key == "codigo" and "variante" in h:
                        continue
                    if key not in idx:
                        idx[key] = i
                    break
    return idx


def _igic_from_grupo(raw: str) -> Decimal:
    s = normalizar_texto(raw)
    if "cero" in s or s.endswith("-0") or s == "0":
        return Decimal("0")
    if "red" in s or "3" in s:
        return Decimal("3")
    if "gen" in s or "7" in s:
        return Decimal("7")
    try:
        return as_decimal(raw.replace("%", "").strip() or "0")
    except Exception:
        return Decimal("0")


def parsear_excel_lineas_noray(ruta: str | Path) -> ResultadoParseNoray:
    path = Path(ruta)
    if not path.exists():
        return ResultadoParseNoray(False, f"Archivo no encontrado: {path}")
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return ResultadoParseNoray(False, f"No se pudo abrir Excel: {exc}")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return ResultadoParseNoray(False, "Excel vacío.")
    headers = rows[0]
    col = _map_headers(tuple(headers))
    if "descripcion" not in col and "codigo" not in col:
        return ResultadoParseNoray(
            False,
            "Cabeceras Noray no reconocidas (faltan Descripción / Nº artículo).",
        )
    out: list[NorayLineaParseada] = []
    fechas: list[date] = []
    for r_i, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        tipo = _cell_str(row[0]) if len(row) else ""
        if tipo and normalizar_texto(tipo) not in ("articulo", "artículo", ""):
            # filas de texto / totales
            if "total" in normalizar_texto(tipo):
                continue
        codigo = _cell_str(row[col["codigo"]]) if "codigo" in col else ""
        desc = _cell_str(row[col["descripcion"]]) if "descripcion" in col else ""
        if not codigo and not desc:
            continue
        almacen = _cell_str(row[col["almacen"]]) if "almacen" in col else ""
        cantidad = _cell_dec(row[col["cantidad"]]) if "cantidad" in col else Decimal("0")
        if cantidad <= 0:
            continue
        unidad = _cell_str(row[col["unidad"]]) if "unidad" in col else "Ud"
        coste = _cell_dec(row[col["coste"]]) if "coste" in col else Decimal("0")
        iva_raw = _cell_str(row[col["iva"]]) if "iva" in col else "0"
        if "igic" in normalizar_texto(iva_raw) or "grupo" in normalizar_texto(
            _cell_str(headers[col["iva"]]) if "iva" in col else ""
        ):
            igic = _igic_from_grupo(iva_raw)
        else:
            igic = _cell_dec(iva_raw)
        fecha = _cell_date(row[col["fecha"]]) if "fecha" in col else None
        if fecha:
            fechas.append(fecha)
        out.append(
            NorayLineaParseada(
                codigo_articulo=codigo,
                descripcion=desc,
                almacen=almacen,
                cantidad=cantidad,
                unidad=unidad or "Ud",
                coste_unitario=coste,
                igic_pct=igic,
                fecha=fecha,
                fila_excel=r_i,
            )
        )
    if not out:
        return ResultadoParseNoray(False, "No hay líneas de artículo con cantidad > 0.")
    fecha_doc = max(fechas) if fechas else date.today()
    ref = f"NORAY-{fecha_doc.isoformat()}-{path.stem[:20]}"
    return ResultadoParseNoray(
        True,
        f"{len(out)} línea(s) leídas de «{path.name}».",
        lineas=tuple(out),
        fecha_documento=fecha_doc,
        referencia_sugerida=ref,
    )


def _sugerir_ubicacion(data: AppData, almacen: str) -> tuple[str | None, str | None]:
    """Mapea Cód. almacén Noray → ubicación BM (por código/nombre)."""
    a = normalizar_texto(almacen)
    if not a:
        return None, None
    activos = [u for u in (data.ubicaciones or []) if getattr(u, "activo", True)]
    # Heurística: ECONOMATO→ECO/Economato, COCINA, DESAYUNO→Cocina, SNACK→Restaurante
    prefer: list[str] = []
    if "econom" in a:
        prefer = ["eco", "economato"]
    elif "desay" in a:
        prefer = ["coc", "cocina", "desay"]
    elif "snack" in a or "bebi" in a or "comi" in a:
        prefer = ["res", "restaurante", "coc", "cocina"]
    elif "bar" in a:
        prefer = ["bar"]
    elif "camara" in a or "cámara" in a:
        prefer = ["cam", "camara"]
    for key in prefer:
        for u in activos:
            cod = normalizar_texto(getattr(u, "codigo", "") or "")
            nom = normalizar_texto(u.nombre or "")
            if key == cod or key in nom or key in cod:
                etiq = f"{getattr(u, 'codigo', '') or '—'} · {u.nombre}"
                return u.id, etiq
    # coincidencia directa nombre/código almacén
    for u in activos:
        cod = normalizar_texto(getattr(u, "codigo", "") or "")
        nom = normalizar_texto(u.nombre or "")
        if a == cod or a == nom or a in nom or nom in a:
            etiq = f"{getattr(u, 'codigo', '') or '—'} · {u.nombre}"
            return u.id, etiq
    if len(activos) == 1:
        u = activos[0]
        return u.id, f"{getattr(u, 'codigo', '') or '—'} · {u.nombre}"
    return None, None


def _codigo_prod(p) -> str:
    return (getattr(p, "codigo", None) or "").strip()


def _hits_por_codigo(activos: list, code: str) -> list:
    if not code:
        return []
    key = code.casefold()
    return [
        p
        for p in activos
        if _codigo_prod(p).casefold() == key
    ]


def _hits_por_nombre(activos: list, descripcion: str) -> tuple[list, str]:
    """Devuelve (hits, modo) con modo exact|parcial|vacío."""
    desc = (descripcion or "").strip()
    if not desc:
        return [], ""
    exact = [
        p
        for p in activos
        if normalizar_texto(p.nombre) == normalizar_texto(desc)
    ]
    if exact:
        return exact, "exact"
    parcial = [p for p in activos if coincide_busqueda(p.nombre, desc)]
    if parcial:
        return parcial, "parcial"
    return [], ""


def _match_producto(
    data: AppData, codigo: str, descripcion: str
) -> tuple[str | None, str | None, str | None, str, list[str]]:
    """Empareja por nombre primero; el código verifica alineación.

    Returns:
        (producto_id, producto_nombre, producto_codigo, estado, avisos)
    """
    avisos: list[str] = []
    activos = [p for p in (data.productos or []) if getattr(p, "activo", True)]
    code = (codigo or "").strip()
    por_nombre, modo_nom = _hits_por_nombre(activos, descripcion)
    por_codigo = _hits_por_codigo(activos, code)

    # 1) Nombre primero
    if len(por_nombre) > 1:
        avisos.append(
            f"Varios productos coinciden con «{descripcion}». Elija o cree uno nuevo."
        )
        return None, None, None, MATCH_AMBIGUO, avisos

    if len(por_nombre) == 1:
        p = por_nombre[0]
        pcod = _codigo_prod(p)
        if modo_nom == "parcial":
            avisos.append("Emparejado por nombre parcial; verifique.")
        # Verificar código
        if not code:
            avisos.append("Sin código Noray; verifique el producto.")
            return p.id, p.nombre, pcod or None, MATCH_REVISAR, avisos
        if not pcod:
            avisos.append(
                f"Nombre OK («{p.nombre}») pero el producto no tiene código; "
                f"Noray trae {code}. Revise o cree producto nuevo."
            )
            return p.id, p.nombre, None, MATCH_REVISAR, avisos
        if pcod.casefold() == code.casefold():
            return p.id, p.nombre, pcod, MATCH_OK, avisos
        # Código distinto: ¿apunta a otro producto?
        if len(por_codigo) == 1 and por_codigo[0].id != p.id:
            otro = por_codigo[0]
            avisos.append(
                f"Conflicto: nombre → «{p.nombre}» [{pcod or '—'}]; "
                f"código {code} → «{otro.nombre}». Corrija, reasigne o cree nuevo."
            )
            return p.id, p.nombre, pcod, MATCH_CONFLICTO, avisos
        if len(por_codigo) > 1:
            avisos.append(
                f"Nombre → «{p.nombre}» pero código {code} es ambiguo en catálogo."
            )
            return p.id, p.nombre, pcod, MATCH_CONFLICTO, avisos
        avisos.append(
            f"Nombre → «{p.nombre}» [{pcod}] no coincide con código Noray {code}. "
            "Verifique, reasigne o cree producto nuevo."
        )
        return p.id, p.nombre, pcod, MATCH_CONFLICTO, avisos

    # 2) Sin nombre: el código solo sugiere, nunca confirma solo
    if len(por_codigo) == 1:
        p = por_codigo[0]
        avisos.append(
            f"Solo coincidencia por código {code} → «{p.nombre}». "
            f"Nombre Noray «{descripcion}» no emparejó; verifique."
        )
        return p.id, p.nombre, _codigo_prod(p) or None, MATCH_REVISAR, avisos
    if len(por_codigo) > 1:
        avisos.append(f"Código {code} ambiguo ({len(por_codigo)} productos).")
        return None, None, None, MATCH_AMBIGUO, avisos

    avisos.append(
        "Sin producto emparejado por nombre ni código. Puede crear uno nuevo."
    )
    return None, None, None, MATCH_SIN, avisos


def emparejar_lineas_noray(
    data: AppData, lineas: tuple[NorayLineaParseada, ...] | list[NorayLineaParseada]
) -> list[NorayLineaMatch]:
    out: list[NorayLineaMatch] = []
    for ln in lineas:
        pid, pnombre, pcod, estado, avisos = _match_producto(
            data, ln.codigo_articulo, ln.descripcion
        )
        uid, uetiq = _sugerir_ubicacion(data, ln.almacen)
        avisos = list(avisos)
        if not uid:
            avisos.append("Elija ubicación de almacenamiento.")
        out.append(
            NorayLineaMatch(
                linea=ln,
                producto_id=pid,
                producto_nombre=pnombre,
                producto_codigo=pcod,
                ubicacion_sugerida_id=uid,
                ubicacion_sugerida_etiqueta=uetiq,
                estado=estado,
                avisos=avisos,
            )
        )
    return out


def lineas_match_a_payload_compra(
    matches: list[NorayLineaMatch],
    *,
    ubicaciones_por_fila: dict[int, str] | None = None,
) -> list[dict]:
    """Payload para ``compra_registro_service.guardar_borrador``."""
    payload: list[dict] = []
    ubi_map = ubicaciones_por_fila or {}
    for i, m in enumerate(matches):
        if not m.producto_id:
            continue
        ln = m.linea
        ubi = ubi_map.get(ln.fila_excel) or m.ubicacion_sugerida_id
        payload.append(
            {
                "producto_id": m.producto_id,
                "client_line_key": f"noray-{ln.fila_excel}-{i}",
                "cantidad_compra": str(ln.cantidad),
                "unidad_compra": ln.unidad,
                "precio_unitario_compra": str(ln.coste_unitario),
                "impuesto_porcentaje": str(ln.igic_pct),
                "ubicacion_destino_id": ubi or None,
            }
        )
    return payload
