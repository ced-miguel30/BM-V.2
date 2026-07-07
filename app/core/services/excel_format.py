"""Formato profesional para exportaciones Excel."""

from __future__ import annotations

import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

NAVY = "0B1F3A"
WHITE = "FFFFFF"
LIGHT_BORDER = "E2E6EC"

_MONEDA_PALABRAS = (
    "coste", "precio", "consumo", "merma", "expiración", "expiracion",
    "total", "importe", "valor",
)


def _nombre_tabla_seguro(nombre: str) -> str:
    limpio = re.sub(r"[^A-Za-z0-9_]", "", nombre)
    if not limpio or limpio[0].isdigit():
        limpio = "T" + limpio
    return limpio[:240]


def _es_columna_moneda(cabecera: str) -> bool:
    texto = cabecera.strip().lower()
    return any(p in texto for p in _MONEDA_PALABRAS)


def _es_valor_numerico(valor) -> bool:
    return isinstance(valor, (int, float)) and not isinstance(valor, bool)


def ajustar_ancho_columnas(ws: Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor is not None:
                max_len = max(max_len, len(str(valor)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 52)


def ajustar_altura_filas(ws: Worksheet) -> None:
    if ws.max_row < 1:
        return
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, ws.max_row + 1):
        max_lineas = 1
        for col_idx in range(1, ws.max_column + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor is None:
                continue
            texto = str(valor)
            ancho = ws.column_dimensions[get_column_letter(col_idx)].width or 12
            lineas = max(1, len(texto) // max(int(ancho), 1) + 1)
            max_lineas = max(max_lineas, lineas)
        ws.row_dimensions[row_idx].height = min(16 + (max_lineas - 1) * 12, 48)


def formatear_cabecera(ws: Worksheet) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(color=WHITE, bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def aplicar_bordes_celdas(ws: Worksheet) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    borde = Border(
        left=Side(style="thin", color=LIGHT_BORDER),
        right=Side(style="thin", color=LIGHT_BORDER),
        top=Side(style="thin", color=LIGHT_BORDER),
        bottom=Side(style="thin", color=LIGHT_BORDER),
    )
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).border = borde


def aplicar_alineacion_datos(ws: Worksheet) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    cabeceras = [
        str(ws.cell(row=1, column=c).value or "")
        for c in range(1, ws.max_column + 1)
    ]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cabecera = cabeceras[col_idx - 1]
            if _es_columna_moneda(cabecera) and _es_valor_numerico(cell.value):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def aplicar_formato_moneda(ws: Worksheet) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    formato = '#,##0.00" €"'
    for col_idx in range(1, ws.max_column + 1):
        cabecera = str(ws.cell(row=1, column=col_idx).value or "")
        if not _es_columna_moneda(cabecera):
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if _es_valor_numerico(cell.value):
                cell.number_format = formato


def aplicar_tabla_excel(ws: Worksheet, nombre_tabla: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    nombre = _nombre_tabla_seguro(nombre_tabla)
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    if nombre in ws.tables:
        del ws.tables[nombre]
    tabla = Table(displayName=nombre, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def congelar_cabecera(ws: Worksheet) -> None:
    if ws.max_row >= 2:
        ws.freeze_panes = "A2"


def formatear_hoja_datos(ws: Worksheet, nombre_tabla: str) -> None:
    formatear_cabecera(ws)
    aplicar_tabla_excel(ws, nombre_tabla)
    aplicar_formato_moneda(ws)
    aplicar_alineacion_datos(ws)
    ajustar_ancho_columnas(ws)
    ajustar_altura_filas(ws)
    congelar_cabecera(ws)


def formatear_hoja_info(ws: Worksheet) -> None:
    formatear_cabecera(ws)
    aplicar_bordes_celdas(ws)
    aplicar_formato_moneda(ws)
    aplicar_alineacion_datos(ws)
    ajustar_ancho_columnas(ws)
    ajustar_altura_filas(ws)


def formatear_libro(writer, hojas: list[tuple[str, str, bool]]) -> None:
    """
    Aplica formato tras escribir DataFrames.
    hojas: [(nombre_hoja, nombre_tabla, es_datos), ...]
    """
    for nombre_hoja, nombre_tabla, es_datos in hojas:
        if nombre_hoja not in writer.sheets:
            continue
        ws = writer.sheets[nombre_hoja]
        if es_datos:
            formatear_hoja_datos(ws, nombre_tabla)
        else:
            formatear_hoja_info(ws)
