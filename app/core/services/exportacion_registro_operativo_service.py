"""Exportación semanal unificada del registro operativo (desayuno, comida, cena, buffet)."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.application.context import AppContext
from app.core.services import bebida_service, cena_service, comida_service, desayuno_service
from app.core.services.buffet_consumo_service import registros_exportables as buffet_exportables
from app.core.services.excel_bloques import RegistroExportable, escribir_hoja_info
from app.core.services.exportacion_semanal_service import ResultadoExportacion, _try_ctx
from app.core.storage.instance_paths import get_exports_root


def _dir_salida() -> Path:
    return get_exports_root(for_write=True) / "semanal" / "registro_operativo"


def _escribir_hoja_registros(ws, registros: list[RegistroExportable]) -> None:
    if not registros:
        ws.cell(1, 1, "Sin registros en el periodo.")
        return
    columnas = registros[0].columnas
    headers = ["Fecha", "Hora", "Ref.", "Usuario"] + columnas
    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True)
    fila = 2
    for reg in sorted(registros, key=lambda r: (r.fecha, r.hora or datetime.min.time())):
        hora_txt = reg.hora.strftime("%H:%M") if reg.hora else ""
        for sub in reg.filas or [[]]:
            ws.cell(fila, 1, reg.fecha.isoformat())
            ws.cell(fila, 2, hora_txt)
            ws.cell(fila, 3, reg.identificador)
            ws.cell(fila, 4, reg.usuario or "")
            for i, val in enumerate(sub):
                if i < len(columnas):
                    ws.cell(fila, 5 + i, val)
            fila += 1


def exportar_semana_registro_operativo(
    inicio: date,
    fin: date,
    *,
    ctx: AppContext | None = None,
) -> ResultadoExportacion:
    """Genera un único .xlsx con hojas Desayuno, BebidasDesayuno, Comida, Cena y ConsumoBuffet."""
    if fin < inicio:
        return ResultadoExportacion(False, "Rango inválido (fin < inicio).")

    context = _try_ctx(ctx)
    if context is None:
        return ResultadoExportacion(False, "No hay contexto de aplicación.")

    hasta_dt = datetime.combine(fin, datetime.max.time())
    data = context.data()
    desayuno = desayuno_service.registros_exportables(inicio, hasta_dt, ctx=context)
    bebidas_todas = bebida_service.registros_exportables(inicio, hasta_dt, ctx=context)
    ids_bebidas_desayuno = {
        r.id
        for r in data.registros_servicio
        if r.tipo_servicio == "bebidas"
        and not getattr(r, "anulado", False)
        and inicio <= r.fecha <= fin
        and str(getattr(r, "clave_idempotencia", "") or "").startswith("bebidas-desayuno-xlsx-")
    }
    bebidas_desayuno = [r for r in bebidas_todas if r.identificador in ids_bebidas_desayuno]
    comida = comida_service.registros_exportables(inicio, hasta_dt, ctx=context)
    cena = cena_service.registros_exportables(inicio, hasta_dt, ctx=context)
    buffet = buffet_exportables(inicio, hasta_dt, ctx=context)

    mermas_buffet: list[RegistroExportable] = []
    for m in data.mermas:
        if getattr(m, "anulado", False) or m.fecha < inicio or m.fecha > fin:
            continue
        for ln in m.lineas:
            com = (ln.comentario or "").lower()
            if "buffet" not in com:
                continue
            mermas_buffet.append(
                RegistroExportable(
                    fecha=m.fecha,
                    hora=m.hora,
                    tipo="Merma buffet",
                    identificador=m.id,
                    usuario=m.registrado_por,
                    columnas=["Producto", "Cantidad", "Motivo", "Comentario", "Coste"],
                    filas=[[
                        ln.producto_nombre_snapshot or ln.producto_id,
                        ln.cantidad,
                        ln.motivo.value if hasattr(ln.motivo, "value") else ln.motivo,
                        ln.comentario or "",
                        ln.coste,
                    ]],
                )
            )

    total = len(desayuno) + len(bebidas_desayuno) + len(comida) + len(cena) + len(buffet) + len(mermas_buffet)
    if total == 0:
        return ResultadoExportacion(False, "Sin registros exportables en el periodo.")

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Info"
    escribir_hoja_info(
        ws_info,
        titulo_documento="Registro operativo semanal",
        periodo_txt=f"{inicio.strftime('%d/%m/%Y')} — {fin.strftime('%d/%m/%Y')}",
        fecha_exportacion_txt=datetime.now().strftime("%d/%m/%Y %H:%M"),
        tipo_exportacion="Semanal unificada",
        total_registros=total,
    )

    hojas = (
        ("Desayuno", desayuno),
        ("BebidasDesayuno", bebidas_desayuno),
        ("Comida", comida),
        ("Cena", cena),
        ("ConsumoBuffet", buffet),
        ("MermasBuffet", mermas_buffet),
    )
    for nombre, regs in hojas:
        ws = wb.create_sheet(nombre)
        _escribir_hoja_registros(ws, regs)

    out_dir = _dir_salida()
    out_dir.mkdir(parents=True, exist_ok=True)
    nombre_archivo = f"Registro_operativo_{inicio.isoformat()}_{fin.isoformat()}.xlsx"
    dest = out_dir / nombre_archivo
    if dest.exists():
        suf = 2
        while dest.exists():
            dest = out_dir / f"Registro_operativo_{inicio.isoformat()}_{fin.isoformat()}_{suf}.xlsx"
            suf += 1
    wb.save(dest)
    return ResultadoExportacion(
        True,
        f"Exportado {total} registro(s).",
        ruta=dest,
        nombre_archivo=dest.name,
        filas_exportadas=total,
    )


def exportar_semana_cerrada(fecha_ref: date | None = None, *, ctx: AppContext | None = None) -> ResultadoExportacion:
    from datetime import timedelta

    ref = fecha_ref or date.today()
    lunes = ref - timedelta(days=ref.weekday())
    domingo = lunes + timedelta(days=6)
    return exportar_semana_registro_operativo(lunes, domingo, ctx=ctx)
