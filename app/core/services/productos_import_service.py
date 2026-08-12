"""Importación de productos + stock inicial desde Excel (Productos PRECIO).

Coste unitario del Excel = aproximación; los albaranes posteriores lo corrigen.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from app.core.services import catalogo_service, stock_service
from app.core.services.money import normalizar_codigo_funcional
from app.core.storage.session_store import get_data

COD_ECONOMATO = "ECO"
COD_COCINA = "COC"
COD_RESTAURANTE = "RES"

UBICACIONES_BASE: tuple[tuple[str, str], ...] = (
    ("Economato", COD_ECONOMATO),
    ("Cocina", COD_COCINA),
    ("Restaurante", COD_RESTAURANTE),
)

CATEGORIAS_COMIDA = frozenset({"101", "102", "103", "104", "105", "106", "108"})

MAPA_CATEGORIA_UBICACION: dict[str, str] = {
    "101": COD_COCINA,
    "102": COD_COCINA,
    "103": COD_COCINA,
    "104": COD_ECONOMATO,
    "105": COD_ECONOMATO,
    "106": COD_COCINA,
    "108": COD_COCINA,
    "": COD_ECONOMATO,
}

_PALABRAS_BEBIDA = (
    "agua",
    "vino",
    "cerveza",
    "whisky",
    "whiskey",
    "ron",
    "ginebra",
    "vodka",
    "licor",
    "zumo",
    "refresco",
    "coca cola",
    "cocacola",
    "fanta",
    "sprite",
    "bebida",
    "cava",
    "champagne",
    "sidra",
    "vermouth",
    "vermut",
    "anís",
    "anis",
    "amaretto",
    "baileys",
    "brandy",
    "cognac",
    "tequila",
    "sangría",
    "sangria",
    "aquabona",
    "royal bliss",
)

# Evita falsos positivos: «agua»⊂aguacate, «ron»⊂kronos, «cola»⊂cola de rape.
_RE_BEBIDA_NOMBRE = None


def _re_bebida_nombre():
    global _RE_BEBIDA_NOMBRE
    if _RE_BEBIDA_NOMBRE is None:
        import re

        alts = "|".join(
            re.escape(p) for p in sorted(_PALABRAS_BEBIDA, key=len, reverse=True)
        )
        _RE_BEBIDA_NOMBRE = re.compile(
            rf"(?<![a-záéíóúüñ])(?:{alts})(?![a-záéíóúüñ])",
            re.IGNORECASE,
        )
    return _RE_BEBIDA_NOMBRE


_PRECIO_MINIMO = 0.01
_MARCA_INICIAL = "Inventario inicial (coste aprox.)"


@dataclass
class ResumenImportProductos:
    ubicaciones_creadas: int = 0
    productos_creados: int = 0
    lotes_creados: int = 0
    omitidos_existentes: int = 0
    omitidos_filtro: int = 0
    omitidos_invalidos: int = 0
    errores: list[str] = field(default_factory=list)
    dry_run: bool = False


def mapear_unidad_excel(raw: object) -> str:
    t = str(raw or "").strip().upper()
    if t in {"UD", "UND", "UN", "UDS", "UNDS"}:
        return "Ud"
    if t in {"KG", "KILO", "KILOS"}:
        return "Kg"
    if t in {"LT", "L", "LITRO", "LITROS"}:
        return "L"
    if t in {"GR", "G", "GRAMO", "GRAMOS"}:
        return "gr"
    if t in {"BT", "PQ", "PAQ", "CAJA"}:
        return "Ud"
    return ""


def inferir_unidad_por_nombre(nombre: str) -> str:
    n = (nombre or "").casefold()
    if any(p in n for p in _PALABRAS_BEBIDA):
        return "L"
    if any(
        k in n
        for k in (
            "carne",
            "pollo",
            "cerdo",
            "ternera",
            "vaca",
            "pescado",
            "salmon",
            "salmón",
            "fruta",
            "manzana",
            "lechuga",
            "queso",
            "chorizo",
            "lomo",
            "entrecot",
            "solomillo",
        )
    ):
        return "Kg"
    if any(k in n for k in ("ud", "und", "croissant", "pan ", "baguette", "mini ")):
        return "Ud"
    return "Kg"


def resolver_unidad(unidad_excel: object, nombre: str) -> str:
    mapped = mapear_unidad_excel(unidad_excel)
    return mapped or inferir_unidad_por_nombre(nombre)


def es_bebida_por_nombre(nombre: str) -> bool:
    """Heurística por nombre con límites de palabra (no substring ciego)."""
    n = (nombre or "").strip()
    if not n:
        return False
    # Exclusiones explícitas de alimentos que contienen raíces de bebida.
    nl = n.casefold()
    if nl.startswith("aguacate") or "aguacate" in nl:
        return False
    if "cola de rape" in nl or "cola rape" in nl:
        return False
    return _re_bebida_nombre().search(n) is not None


def es_bebida_por_categoria_y_nombre(*, categoria: str, nombre: str) -> bool:
    """Regla canónica: categoría Excel 104 = bebidas; el nombre solo refuerza."""
    cat = str(categoria or "").strip()
    if cat == "104":
        return True
    if cat and cat != "104":
        # Fuera de 104 no se marca bebida por nombre (evita aguacate/queso…).
        return False
    return es_bebida_por_nombre(nombre)


def limpiar_nombre_producto(nombre: str) -> str:
    """Quita códigos C00000x embebidos en el nombre comercial (el código va en ``codigo``)."""
    import re

    n = (nombre or "").strip()
    if not n:
        return n
    # «NOMBRE (C00000603)» o «NOMBRE C00000603»
    n = re.sub(r"\s*\(\s*C0*\d+\s*\)\s*$", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+C0*\d{5,}\s*$", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s{2,}", " ", n).strip(" -–—")
    return n or (nombre or "").strip()


def map_categoria_a_ubicacion_codigo(categoria: str) -> str:
    return MAPA_CATEGORIA_UBICACION.get(str(categoria or "").strip(), COD_ECONOMATO)


def fila_incluida_comida(*, categoria: str, codigo: str, nombre: str) -> bool:
    cat = str(categoria or "").strip()
    if cat in CATEGORIAS_COMIDA:
        return True
    cod = str(codigo or "").strip().upper()
    nom = str(nombre or "").strip()
    if not cat and cod.startswith("C") and len(nom) >= 2:
        return True
    return False


def _ubicacion_id_por_codigo(codigo: str) -> str | None:
    codigo_n = normalizar_codigo_funcional(codigo)
    if not codigo_n:
        return None
    data = get_data()
    for u in getattr(data, "ubicaciones", []) or []:
        if normalizar_codigo_funcional(getattr(u, "codigo", None)) == codigo_n:
            return u.id
    return None


def _producto_id_por_codigo(codigo: str) -> str | None:
    codigo_n = normalizar_codigo_funcional(codigo)
    if not codigo_n:
        return None
    data = get_data()
    for p in data.productos:
        if normalizar_codigo_funcional(getattr(p, "codigo", None)) == codigo_n:
            return p.id
    return None


def ensure_ubicaciones_base(*, dry_run: bool = False) -> tuple[dict[str, str], int]:
    """Garantiza Economato/Cocina/Restaurante. Devuelve mapa codigo→id y altas."""
    creadas = 0
    mapa: dict[str, str] = {}
    for nombre, codigo in UBICACIONES_BASE:
        existente = _ubicacion_id_por_codigo(codigo)
        if existente:
            mapa[codigo] = existente
            continue
        if dry_run:
            mapa[codigo] = f"dry-{codigo}"
            creadas += 1
            continue
        r = catalogo_service.crear_ubicacion(nombre, codigo=codigo)
        if r.ok:
            creado = _ubicacion_id_por_codigo(codigo)
            if creado:
                mapa[codigo] = creado
                creadas += 1
                continue
        from app.core.services.catalogo_service import normalizar_nombre_catalogo

        data = get_data()
        clave = normalizar_nombre_catalogo(nombre)
        hit = next(
            (
                u
                for u in data.ubicaciones
                if normalizar_nombre_catalogo(u.nombre) == clave
            ),
            None,
        )
        if hit:
            mapa[codigo] = hit.id
        else:
            raise RuntimeError(f"No se pudo asegurar ubicación «{nombre}»: {r.mensaje}")
    return mapa, creadas


def _servicios_para(*, categoria: str, es_bebida: bool) -> list[str]:
    if es_bebida:
        return ["bebidas"]
    if str(categoria or "").strip() == "102":
        return ["desayuno"]
    return ["desayuno", "comida", "cena"]


def _float_safe(raw: object) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _leer_filas(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Productos" not in wb.sheetnames:
            raise ValueError("El Excel no tiene hoja «Productos».")
        ws = wb["Productos"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return []
        header = [str(c or "").strip().casefold() for c in rows[0]]
        ncols = max(len(header), max((len(r) for r in rows[1:]), default=0))

        # Índices por forma del fichero:
        # PRECIO (8 cols): Nº, Desc, Inv, Unidad, CosteAjust, CosteUnit, Cat, Desc2
        # Productos (4 cols): Nº, Desc, Inv, Cat
        if ncols >= 7:
            i_cod, i_nom, i_inv, i_uni, i_coste, i_cat = 0, 1, 2, 3, 5, 6
        else:
            i_cod, i_nom, i_inv, i_uni, i_coste, i_cat = 0, 1, 2, None, None, 3

        out: list[dict] = []
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def cell(idx: int | None):
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            out.append(
                {
                    "codigo": cell(i_cod),
                    "nombre": cell(i_nom),
                    "inventario": cell(i_inv),
                    "unidad": cell(i_uni),
                    "coste_unitario": cell(i_coste),
                    "categoria": cell(i_cat),
                }
            )
        return out
    finally:
        wb.close()


def _codigos_lista_productos_xlsx(precio_path: Path) -> set[str] | None:
    """Si existe docs/Productos.xlsx junto al PRECIO, limita el import a esos Nº."""
    sibling = precio_path.parent / "Productos.xlsx"
    if not sibling.is_file():
        return None
    try:
        filas = _leer_filas(sibling)
    except Exception:  # noqa: BLE001
        return None
    out: set[str] = set()
    for f in filas:
        cod = normalizar_codigo_funcional(str(f.get("codigo") or ""))
        if cod:
            out.add(cod)
    return out or None


def importar_productos_desde_excel(
    path: str | Path,
    *,
    dry_run: bool = False,
    solo_codigos: set[str] | None = None,
) -> ResumenImportProductos:
    """Importa productos comida/bebida + stock inicial aproximado.

    Si junto al Excel existe ``Productos.xlsx``, se limita a esos códigos
    (~lista operativa). Si no, filtra por categorías 101–108.
    """
    resumen = ResumenImportProductos(dry_run=dry_run)
    excel_path = Path(path)
    if not excel_path.is_file():
        resumen.errores.append(f"No se encuentra el Excel: {excel_path}")
        return resumen

    try:
        filas = _leer_filas(excel_path)
    except Exception as exc:  # noqa: BLE001
        resumen.errores.append(f"No se pudo leer el Excel: {exc}")
        return resumen

    allow = solo_codigos
    if allow is None:
        allow = _codigos_lista_productos_xlsx(excel_path)

    try:
        ubi_mapa, n_ubi = ensure_ubicaciones_base(dry_run=dry_run)
        resumen.ubicaciones_creadas = n_ubi
    except Exception as exc:  # noqa: BLE001
        resumen.errores.append(str(exc))
        return resumen

    for fila in filas:
        codigo_raw = str(fila.get("codigo") or "").strip()
        nombre = str(fila.get("nombre") or "").strip()
        cat = str(fila.get("categoria") or "").strip()
        if cat.endswith(".0"):
            cat = cat[:-2]

        codigo_n = normalizar_codigo_funcional(codigo_raw)
        if allow is not None:
            if not codigo_n or codigo_n not in allow:
                resumen.omitidos_filtro += 1
                continue
        elif not fila_incluida_comida(categoria=cat, codigo=codigo_raw, nombre=nombre):
            resumen.omitidos_filtro += 1
            continue

        if not codigo_n or len(nombre) < 2:
            resumen.omitidos_invalidos += 1
            continue

        if _producto_id_por_codigo(codigo_n):
            resumen.omitidos_existentes += 1
            continue

        unidad = resolver_unidad(fila.get("unidad"), nombre)
        nombre = limpiar_nombre_producto(nombre)
        es_bebida = es_bebida_por_categoria_y_nombre(categoria=cat, nombre=nombre)

        ubi_cod = map_categoria_a_ubicacion_codigo(cat)
        ubi_id = ubi_mapa.get(ubi_cod)
        if not ubi_id:
            resumen.errores.append(f"{codigo_n}: ubicación {ubi_cod} no disponible")
            continue

        qty = _float_safe(fila.get("inventario")) or 0.0
        coste_u = _float_safe(fila.get("coste_unitario"))
        if qty > 0 and coste_u is not None and coste_u > 0:
            precio_total = max(_PRECIO_MINIMO, round(coste_u * qty, 2))
        elif qty > 0:
            precio_total = _PRECIO_MINIMO
        else:
            precio_total = 0.0

        if dry_run:
            resumen.productos_creados += 1
            if qty > 0:
                resumen.lotes_creados += 1
            continue

        r_prod = stock_service.crear_producto(
            nombre,
            unidad,
            None,
            codigo=codigo_n,
            es_bebida=es_bebida,
            servicios_disponibles=_servicios_para(categoria=cat, es_bebida=es_bebida),
            categoria_inventario=cat or None,
            ubicacion_ids=[ubi_id],
            tipo_articulo="consumible",
        )
        if not r_prod.ok and "Ya existe" in (r_prod.mensaje or "") and "llamado" in (
            r_prod.mensaje or ""
        ):
            # Mismo nombre comercial, distinto código de factura.
            r_prod = stock_service.crear_producto(
                f"{nombre} ({codigo_n})",
                unidad,
                None,
                codigo=codigo_n,
                es_bebida=es_bebida,
                servicios_disponibles=_servicios_para(
                    categoria=cat, es_bebida=es_bebida
                ),
                categoria_inventario=cat or None,
                ubicacion_ids=[ubi_id],
                tipo_articulo="consumible",
            )
        if not r_prod.ok:
            resumen.errores.append(f"{codigo_n} «{nombre}»: {r_prod.mensaje}")
            continue

        resumen.productos_creados += 1
        prod_id = _producto_id_por_codigo(codigo_n)
        if not prod_id:
            resumen.errores.append(f"{codigo_n}: creado pero no localizado")
            continue

        if qty <= 0:
            continue

        r_lote = stock_service.registrar_lote(
            prod_id,
            precio_total,
            qty,
            marca_proveedor=_MARCA_INICIAL,
            ubicacion_destino_id=ubi_id,
        )
        if not r_lote.ok:
            resumen.errores.append(f"{codigo_n} stock: {r_lote.mensaje}")
            continue
        resumen.lotes_creados += 1

    return resumen


def ruta_excel_precio_default() -> Path:
    from app.core.storage.demo_files import PROJECT_ROOT

    return PROJECT_ROOT / "docs" / "Productos PRECIO.xlsx"


def mensaje_resumen(r: ResumenImportProductos) -> str:
    parts = [
        f"Ubicaciones nuevas: {r.ubicaciones_creadas}",
        f"Productos creados: {r.productos_creados}",
        f"Lotes iniciales: {r.lotes_creados}",
        f"Ya existían: {r.omitidos_existentes}",
        f"Fuera de filtro: {r.omitidos_filtro}",
        f"Inválidos: {r.omitidos_invalidos}",
    ]
    if r.dry_run:
        parts.insert(0, "DRY-RUN")
    msg = " · ".join(parts)
    if r.errores:
        preview = "; ".join(r.errores[:5])
        more = f" (+{len(r.errores) - 5} más)" if len(r.errores) > 5 else ""
        msg += f" · Errores: {preview}{more}"
    return msg
