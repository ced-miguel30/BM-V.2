"""Hojas Excel con varios registros separados dentro del mismo día.

Complementa (sin modificar) `excel_format.py`: aquella asume una única tabla
que empieza en la fila 1 de la hoja; aquí una misma hoja "día" puede contener
varios bloques de registro apilados (cada uno con su propia cabecera de
contexto y su propia tabla), con separación clara entre ellos.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, time
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from app.core.services.excel_format import (
    NAVY,
    WHITE,
    _es_columna_moneda,
    _es_valor_numerico,
    _nombre_tabla_seguro,
    formatear_hoja_info,
)

_DIAS_SEMANA = ("Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom")

# Separación (filas en blanco) entre el final de la tabla de un registro y el
# inicio del bloque del siguiente registro dentro de la misma hoja.
_FILAS_SEPARACION = 2


@dataclass(frozen=True)
class RegistroExportable:
    """Un registro individual exportable (un desayuno, una línea de merma,
    una compra, una entrada de actividad, ...)."""

    fecha: date
    hora: time | None
    tipo: str
    identificador: str
    usuario: str | None
    columnas: list[str]
    filas: list[list[Any]]
    resumen: list[tuple[str, str]] | None = None


def nombre_hoja_dia(fecha: date) -> str:
    """Nombre corto y válido para Excel, p. ej. 'Lun 20-07' (≤31 caracteres)."""
    dia = _DIAS_SEMANA[fecha.weekday()]
    return f"{dia} {fecha.strftime('%d-%m')}"


def _fila_titulo_dia(ws: Worksheet, fecha: date, num_registros: int) -> None:
    dia_nombre = ("lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo")[fecha.weekday()]
    texto = f"{dia_nombre.capitalize()} {fecha.strftime('%d/%m/%Y')} — {num_registros} registro(s)"
    celda = ws.cell(row=1, column=1, value=texto)
    celda.font = Font(color=WHITE, bold=True, size=12)
    celda.fill = PatternFill("solid", fgColor=NAVY)
    celda.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A3"


def _escribir_encabezado_registro(ws: Worksheet, fila: int, registro: RegistroExportable) -> int:
    hora_txt = registro.hora.strftime("%H:%M") if registro.hora else "—"
    partes = [f"Hora {hora_txt}", registro.tipo]
    if registro.usuario:
        partes.append(f"Usuario: {registro.usuario}")
    partes.append(f"Nº {registro.identificador}")
    if registro.resumen:
        partes.extend(f"{clave}: {valor}" for clave, valor in registro.resumen)
    texto = "  ·  ".join(partes)

    celda = ws.cell(row=fila, column=1, value=texto)
    celda.font = Font(color=NAVY, bold=True, size=10.5)
    celda.fill = PatternFill("solid", fgColor="EDF1F7")
    celda.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 20
    return fila + 1


def _escribir_tabla_registro(
    ws: Worksheet,
    fila_cabecera: int,
    registro: RegistroExportable,
    nombre_tabla: str,
) -> int:
    for col_idx, titulo in enumerate(registro.columnas, start=1):
        celda = ws.cell(row=fila_cabecera, column=col_idx, value=titulo)
        celda.font = Font(color=WHITE, bold=True, size=10.5)
        celda.fill = PatternFill("solid", fgColor=NAVY)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fila_datos_inicio = fila_cabecera + 1
    for offset, fila_valores in enumerate(registro.filas):
        fila_actual = fila_datos_inicio + offset
        for col_idx, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_actual, column=col_idx, value=valor)

    fila_fin = fila_datos_inicio + max(len(registro.filas), 1) - 1
    if not registro.filas:
        # Nunca dejamos una tabla sin filas de datos (rompería openpyxl.Table).
        ws.cell(row=fila_datos_inicio, column=1, value="— Sin líneas —")
        fila_fin = fila_datos_inicio

    _formatear_bloque_tabla(ws, fila_cabecera, fila_fin, len(registro.columnas), nombre_tabla)
    return fila_fin + 1


def _formatear_bloque_tabla(
    ws: Worksheet,
    fila_cabecera: int,
    fila_fin: int,
    num_columnas: int,
    nombre_tabla: str,
) -> None:
    cabeceras = [
        str(ws.cell(row=fila_cabecera, column=c).value or "") for c in range(1, num_columnas + 1)
    ]
    for fila_idx in range(fila_cabecera + 1, fila_fin + 1):
        for col_idx in range(1, num_columnas + 1):
            celda = ws.cell(row=fila_idx, column=col_idx)
            cabecera = cabeceras[col_idx - 1]
            if _es_columna_moneda(cabecera) and _es_valor_numerico(celda.value):
                celda.number_format = '#,##0.00" €"'
                celda.alignment = Alignment(horizontal="right", vertical="center")
            else:
                celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    nombre = _nombre_tabla_seguro(nombre_tabla)
    ref = f"A{fila_cabecera}:{get_column_letter(num_columnas)}{fila_fin}"
    if nombre in ws.tables:
        del ws.tables[nombre]
    tabla = Table(displayName=nombre, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def _ajustar_ancho_columnas_manual(ws: Worksheet) -> None:
    for col_idx in range(1, (ws.max_column or 1) + 1):
        max_len = 0
        for row_idx in range(1, (ws.max_row or 1) + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor is not None:
                max_len = max(max_len, len(str(valor)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 12), 52)


def escribir_hoja_dia(ws: Worksheet, fecha: date, registros: list[RegistroExportable], nombre_hoja: str) -> None:
    """Escribe en `ws` el título del día (fila 1, congelada) y, a partir de
    la fila 3, un bloque por registro: línea de contexto + tabla de desglose,
    con `_FILAS_SEPARACION` filas en blanco entre bloques."""
    _fila_titulo_dia(ws, fecha, len(registros))

    fila = 3
    for indice, registro in enumerate(registros):
        fila = _escribir_encabezado_registro(ws, fila, registro)
        nombre_tabla = _nombre_tabla_seguro(f"{nombre_hoja}_{registro.identificador}_{indice}")
        fila = _escribir_tabla_registro(ws, fila, registro, nombre_tabla)
        fila += _FILAS_SEPARACION

    _ajustar_ancho_columnas_manual(ws)


def escribir_hoja_info(
    ws: Worksheet,
    *,
    titulo_documento: str,
    periodo_txt: str,
    fecha_exportacion_txt: str,
    tipo_exportacion: str,
    total_registros: int,
) -> None:
    """Hoja de portada común: título, periodo, fecha/hora de exportación."""
    filas = [
        ("Documento", titulo_documento),
        ("Periodo exportado", periodo_txt),
        ("Fecha y hora de exportación", fecha_exportacion_txt),
        ("Tipo de exportación", tipo_exportacion),
        ("Registros incluidos", total_registros),
    ]
    ws.cell(row=1, column=1, value="Campo")
    ws.cell(row=1, column=2, value="Detalle")
    for offset, (campo, valor) in enumerate(filas, start=2):
        ws.cell(row=offset, column=1, value=campo)
        ws.cell(row=offset, column=2, value=valor)

    formatear_hoja_info(ws)
